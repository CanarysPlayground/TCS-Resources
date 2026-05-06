#!/usr/bin/env python3
"""
GitLab -> GitHub Mirror Migration Script
"""
from __future__ import annotations
import abc
import argparse
import base64
import calendar
import csv
import datetime
import json
import logging
from logging.handlers import RotatingFileHandler
import os
import random
import re
import shutil
import signal
import socket
import subprocess
import sys
import tempfile
import threading
import time
import urllib.error
import urllib.parse
import urllib.request
from collections.abc import Callable
from concurrent.futures import Future, ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from pathlib import Path

try:
    __import__("openpyxl")  # availability probe; used lazily in _write_xlsx
    _XLSX_AVAILABLE = True
except ImportError:
    _XLSX_AVAILABLE = False

# Git LFS availability probe -- checked once at startup, used to gate LFS migration.
# shutil.which short-circuits on PATH hit; subprocess fallback handles PATH oddities on
# some RHEL/Ubuntu installs where 'git-lfs' is a shell wrapper that's only on git's PATH.
_LFS_AVAILABLE: bool = shutil.which("git-lfs") is not None or (
    subprocess.run(
        ["git", "lfs", "version"],
        capture_output=True,
        timeout=10,
    ).returncode == 0
)

# ---------------------------------------------------------------------------
# Python version guard
# ---------------------------------------------------------------------------
if sys.version_info < (3, 8):
    sys.exit(
        f"Python 3.8 or later is required (you have {sys.version.split()[0]}). "
        "Please upgrade: https://www.python.org/downloads/"
    )


# ---------------------------------------------------------------------------
# Sentinel for user config errors (shown as WARNING, not ERROR)
# ---------------------------------------------------------------------------
class _MissingConfigError(ValueError):
    """Raised when required user configuration is missing (not a code bug)."""


# ---------------------------------------------------------------------------
# Script-relative paths -- resolved from the script's own directory so the
# script works correctly regardless of the current working directory.
# ---------------------------------------------------------------------------
_SCRIPT_DIR  = Path(__file__).resolve().parent
_RUN_TS      = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
_LOGS_DIR    = _SCRIPT_DIR / "logs"
_REPORTS_DIR = _SCRIPT_DIR / "reports"
_LOGS_DIR.mkdir(exist_ok=True)
_REPORTS_DIR.mkdir(exist_ok=True)
_LOG_FILE        = _LOGS_DIR    / f"mirror-migration-{_RUN_TS}.txt"
_RESULTS_JSON    = _REPORTS_DIR / f"mirror-migration-{_RUN_TS}.json"
_RESULTS_CSV     = _REPORTS_DIR / f"mirror-migration-{_RUN_TS}.csv"
_RESULTS_XLSX    = _REPORTS_DIR / f"mirror-migration-{_RUN_TS}.xlsx"
_CHECKPOINT_FILE = _SCRIPT_DIR  / "mirror-migration-checkpoint.json"
_CONFIG_FILENAME = "mirror-config.json"
_REPOS_CSV_FILENAME = "repos.csv"

# ---------------------------------------------------------------------------
# Internal constants
# ---------------------------------------------------------------------------
_GITHUB_API_URL = "https://api.github.com"
_GITHUB_URL = "https://github.com"

# Pause API calls when remaining quota falls below this threshold.
_RATE_LIMIT_BUFFER = 100

# GitHub hard limit per file object (non-LFS).  Pushes containing any blob
# larger than this are rejected by GitHub's receive-pack with an HTTP 400 error.
_GITHUB_FILE_SIZE_LIMIT_MB = 100

# Maximum repository rows displayed in the pre-flight preview table.
# Repos beyond this limit are summarised as "... and N more".
_PREVIEW_MAX_ROWS = 25

# Default migration settings (all overridable in mirror-config.json)
_DEFAULT_MAX_WORKERS = 5
_DEFAULT_CLONE_TIMEOUT = 7200        # seconds
_DEFAULT_PUSH_TIMEOUT = 7200         # seconds
_DEFAULT_GIT_RETRIES = 3
_DEFAULT_API_RETRIES = 5
_DEFAULT_MIN_FREE_DISK_GB = 10.0     # minimum free disk space in system temp dir (GB)
# Raise git's HTTP POST buffer to avoid "RPC failed; curl 55 / send-pack: unexpected
# disconnect" on large mirror pushes.  Configurable via mirror-config.json
# migration.git_http_post_buffer_bytes.  Default: 500 MB.
_DEFAULT_GIT_HTTP_POST_BUFFER = 500 * 1024 * 1024  # bytes
# Push refs in batches to avoid GitHub HTTP 500 on large repos.  A single
# git push --mirror creates one giant pack file; batching keeps each pack
# small enough for GitHub's receive-pack to accept.  Configurable via
# mirror-config.json migration.git_push_batch_size.  Default: 100 refs/batch.
_DEFAULT_PUSH_BATCH_SIZE = 100


_PLACEHOLDER_TOKENS = (
    "YOUR_", "xxxx", "glpat-xxxx", "ghp_xxxx",
    "fileserver.internal", "example.com",
)
_VALID_VISIBILITIES = frozenset({"private", "internal", "public"})
_VIS_ICONS = {"private": "\U0001f512", "internal": "\U0001f465", "public": "\U0001f310"}


def _fmt_ts() -> str:
    return datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _fmt_duration(seconds: float) -> str:
    """Return a human-readable duration: 'Xh Ym Zs', 'Ym Zs', or 'Zs'."""
    total = int(seconds)
    h, rem = divmod(total, 3600)
    m, s   = divmod(rem, 60)
    if h:
        return f"{h}h {m:02d}m {s:02d}s"
    if m:
        return f"{m}m {s:02d}s"
    return f"{s}s"


def _safe_dir_name(name: str) -> str:
    """Return a filesystem-safe directory name compatible with Linux and Windows.

    Replaces characters that are illegal on Windows NTFS (<>:"/\\|?*) and NUL/control
    bytes that cause issues on Linux.  Trims trailing dots and spaces (NTFS restriction).
    Falls back to '_repo' if the entire name is stripped away.
    """
    sanitized = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", name)
    sanitized = sanitized.rstrip(". ")
    return sanitized or "_repo"


def _is_lfs_repo(mirror_dir: Path) -> bool:
    """Return True if the bare mirror clone contains any Git LFS pointer objects.

    Uses 'git lfs ls-files --all' which scans all refs and reports LFS-tracked files.
    Returns False if git-lfs is not installed or produces no output.
    """
    code, out, _ = _run_git(
        ["git", "lfs", "ls-files", "--all"],
        cwd=mirror_dir,
        timeout=60,
    )
    return code == 0 and bool(out.strip())


def _git_lfs_fetch_all(mirror_dir: Path, label: str, timeout: int) -> tuple[bool, str]:
    """Download all LFS objects from the GitLab remote into the bare clone's LFS store.

    The bare mirror clone has 'origin' pointing to the authenticated GitLab URL (set
    at clone time), so no extra credential setup is needed here.
    """
    code, _, err = _run_git(
        ["git", "lfs", "fetch", "--all"],
        cwd=mirror_dir,
        timeout=timeout,
    )
    if code != 0:
        return False, f"git lfs fetch --all failed: {err}"
    log.info(f"[{label}] LFS fetch complete")
    return True, ""


def _git_lfs_push_all(
    mirror_dir: Path,
    push_url: str,
    safe_push: str,
    label: str,
    timeout: int,
    http_post_buffer: int,
) -> tuple[bool, str]:
    """Push all LFS objects from the bare clone's LFS store to GitHub.

    'git lfs push --all <url>' uploads every LFS object reachable from any ref,
    independently of the normal git push.  We use the same authenticated push URL.
    """
    code, _, err = _run_git(
        cmd=["git", "lfs", "push", "--all", push_url],
        log_cmd=["git", "lfs", "push", "--all", safe_push],
        cwd=mirror_dir,
        timeout=timeout,
        http_post_buffer=http_post_buffer,
    )
    if code != 0:
        return False, f"git lfs push --all failed: {err}"
    log.info(f"[{label}] LFS push complete")
    return True, ""


def _count_lfs_objects(mirror_dir: Path) -> int:
    """Return the number of distinct LFS objects in the bare clone."""
    code, out, _ = _run_git(
        ["git", "lfs", "ls-files", "--all"],
        cwd=mirror_dir,
        timeout=60,
    )
    return sum(1 for line in out.splitlines() if line.strip()) if code == 0 else 0

# ---------------------------------------------------------------------------
# Colored console logging
# ---------------------------------------------------------------------------

# Module-level flag: True only when stdout is an interactive TTY with ANSI support.
# Set by _build_logger(); checked by _c() so inline colorisation is
# unconditionally safe to call anywhere in the module.
_COLORS_ACTIVE: bool = False


def _c(text: str, *codes: int) -> str:
    """Wrap *text* in ANSI SGR escape codes when color output is active.

    Falls back to plain *text* when colors are disabled (non-TTY, CI, file
    redirect) so every call site is safe without an ``if`` guard.

    ANSI code reference (common codes used in this module):
      1=bold  2=dim  4=underline  22=normal-intensity
      30-37=standard fg  90-97=bright fg
      31=red  32=green  33=yellow  34=blue  35=magenta  36=cyan  37=white
      91=bright-red  92=bright-green  93=bright-yellow  96=bright-cyan  97=bright-white
    """
    if not _COLORS_ACTIVE or not codes:
        return text
    return f"\x1b[{';'.join(map(str, codes))}m{text}\x1b[0m"


class _ColorFormatter(logging.Formatter):
    """ANSI-colored log formatter for interactive terminal sessions.

    Three-zone coloring per record:
      Timestamp   → always dim dark-gray (metadata, unobtrusive)
      Level badge → bold + level color   (instant severity scan)
      Message body → level-specific base, with automatic inline pattern
                     recoloring on INFO lines: [labels], action verbs,
                     number+unit combos, durations, and URLs.

    Pattern recoloring is intentionally skipped when the message already
    contains ANSI escape sequences (i.e. colored via _c() at the call site),
    so the rich per-line colors in the summary banner are preserved as-is.

    No third-party dependencies; works on Windows 10+ via ctypes ANSI enable.
    """

    # (badge_color, body_base_color)
    # INFO body uses "\x1b[2;39m" (dim default-fg) so it renders as a
    # comfortable medium-gray on dark terminals instead of glaring bright white.
    _LEVEL_STYLES: dict[int, tuple[str, str]] = {
        logging.DEBUG:    ("\x1b[36m",    "\x1b[2;90m"),   # cyan badge,         dim dark-gray body
        logging.INFO:     ("\x1b[32m",    "\x1b[2;39m"),   # green badge,        dim default-fg body
        logging.WARNING:  ("\x1b[1;33m",  "\x1b[33m"),     # bold-yellow badge,  yellow body
        logging.ERROR:    ("\x1b[1;31m",  "\x1b[31m"),     # bold-red badge,     red body
        logging.CRITICAL: ("\x1b[1;35m",  "\x1b[1;35m"),  # bold-magenta badge + body
    }
    _TS_COLOR = "\x1b[2;90m"
    _RESET    = "\x1b[0m"

    # ---------------------------------------------------------------------------
    # Compiled patterns for INFO inline recoloring.
    # Applied in order so earlier substitutions don't interfere with later ones.
    # ---------------------------------------------------------------------------

    # [org/repo] or [label] prefix at the very start of the message.
    _RE_LABEL = re.compile(r'^\[([^\]]{1,80})\]')

    # Key action verbs/phrases that indicate what the script is doing.
    _RE_ACTION = re.compile(
        r'\b('
        r'Cloning|Pushing|Validating|Applying|Fetching|Retrying|Processing|'
        r'Creating|Setting|Scanning|Checking|Starting|Splitting|Migrating|'
        r'Already fully migrated|Push done|Clone done|'
        r'Custom properties applied|CI skeleton created'
        r')\b'
    )

    # Numeric value immediately followed by a git/migration unit word.
    _RE_NUM_UNIT = re.compile(
        r'\b(\d[\d,]*)\s+(branch(?:es)?|tag[s]?|commit[s]?|'
        r'repo[s]?|ref[s]?|file[s]?|worker[s]?|task[s]?|batch(?:es)?)\b',
        re.IGNORECASE,
    )

    # Duration strings: "2m 03s" or "1.5s"  (not plain "8s" — too broad).
    _RE_DURATION = re.compile(r'\b(\d+m\s+\d{1,2}s|\d+\.\d+s)\b')

    # HTTP/HTTPS URLs — minimum 8 chars after scheme to avoid false positives.
    _RE_URL = re.compile(r'https?://[^\s\x1b,;]{8,}')

    def _recolor_info(self, msg: str, base: str) -> str:
        """Apply inline color spans to a plain (no existing ANSI) INFO message."""
        R = self._RESET + base   # restore base color after every colored span

        # 1. [label] → bold bright-cyan  (stands out as the repo/context identifier)
        msg = self._RE_LABEL.sub(
            lambda m: f"\x1b[1;96m[{m.group(1)}]{R}", msg
        )
        # 2. Action words → bold bright-white  (pop against dim-gray base)
        msg = self._RE_ACTION.sub(
            lambda m: f"\x1b[1;97m{m.group()}{R}", msg
        )
        # 3. number + unit → bright-cyan number, unit stays in base color
        msg = self._RE_NUM_UNIT.sub(
            lambda m: f"\x1b[96m{m.group(1)}{R} {m.group(2)}", msg
        )
        # 4. Durations → bright-cyan  (time values deserve attention)
        msg = self._RE_DURATION.sub(
            lambda m: f"\x1b[96m{m.group()}{R}", msg
        )
        # 5. URLs → dim cyan  (visible but not distracting)
        msg = self._RE_URL.sub(
            lambda m: f"\x1b[2;36m{m.group()}{R}", msg
        )
        return msg

    def format(self, record: logging.LogRecord) -> str:
        badge_color, body_color = self._LEVEL_STYLES.get(record.levelno, ("", ""))

        msg = record.getMessage()
        # Apply pattern recoloring only on plain INFO messages.
        # Messages that already contain ANSI codes (from _c() at the call site)
        # are left untouched so their existing fine-grained colors are preserved.
        if record.levelno == logging.INFO and body_color and "\x1b[" not in msg:
            msg = self._recolor_info(msg, body_color)

        # Never mutate the shared LogRecord -- use a throwaway copy.
        rec      = logging.makeLogRecord(record.__dict__)
        rec.msg  = msg
        rec.args = ()

        # %-8s pads the level name so all columns align:
        # DEBUG=5, INFO=4, WARNING=7, ERROR=5, CRITICAL=8 chars → pad to 8.
        fmt = (
            f"{self._TS_COLOR}%(asctime)s{self._RESET} "
            f"{badge_color}\x1b[1m[%(levelname)-8s]{self._RESET} "
            f"{body_color}%(message)s{self._RESET}"
        )
        return logging.Formatter(fmt, datefmt="%Y-%m-%d %H:%M:%S").format(rec)


def _enable_windows_ansi() -> None:
    """Enable ANSI virtual terminal processing on Windows 10+. No-op elsewhere."""
    if sys.platform != "win32":
        return
    try:
        import ctypes
        import ctypes.wintypes
        kernel = ctypes.windll.kernel32  # type: ignore[attr-defined]
        handle = kernel.GetStdHandle(-11)  # STD_OUTPUT_HANDLE
        if handle == -1:
            return
        mode = ctypes.wintypes.DWORD()
        if not kernel.GetConsoleMode(handle, ctypes.byref(mode)):
            return
        kernel.SetConsoleMode(handle, mode.value | 0x0004)  # ENABLE_VIRTUAL_TERMINAL_PROCESSING
    except Exception:
        pass


def _reconfigure_stdout_utf8() -> None:
    """Force UTF-8 on stdout/stderr (avoids cp1252 UnicodeEncodeError on Windows)."""
    for stream in (sys.stdout, sys.stderr):
        if hasattr(stream, "reconfigure"):
            try:
                stream.reconfigure(encoding="utf-8", errors="replace")
            except Exception:
                pass


def _build_logger() -> logging.Logger:
    """Build a colored console + plain UTF-8 file logger."""
    global _COLORS_ACTIVE
    _reconfigure_stdout_utf8()
    _enable_windows_ansi()

    use_color = sys.stdout.isatty()
    _COLORS_ACTIVE = use_color
    console = logging.StreamHandler()
    console.setFormatter(
        _ColorFormatter()
        if use_color
        else logging.Formatter(
            "%(asctime)s [%(levelname)s] %(message)s", datefmt="%Y-%m-%d %H:%M:%S"
        )
    )

    file_handler = RotatingFileHandler(
        str(_LOG_FILE), encoding="utf-8",
        maxBytes=100 * 1024 * 1024,  # 100 MB per segment; prevents multi-GB log files on long runs
        backupCount=5,
    )
    file_handler.setFormatter(
        logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")
    )

    root = logging.getLogger()
    root.setLevel(logging.INFO)
    root.addHandler(console)
    root.addHandler(file_handler)
    return logging.getLogger(__name__)


log = _build_logger()


# ===========================================================================
# Authentication providers
# ===========================================================================

class _Auth(abc.ABC):
    """Abstract base for GitHub authentication."""

    @abc.abstractmethod
    def get_auth_header(self) -> str:
        """Return the value for the HTTP Authorization header."""

    @abc.abstractmethod
    def get_token_for_git(self) -> str:
        """Return a token suitable for embedding in a git HTTPS URL."""

    @property
    @abc.abstractmethod
    def mode_label(self) -> str:
        """Short label used in log messages."""


class PATAuth(_Auth):
    """Personal Access Token authentication (simple, stateless)."""

    def __init__(self, pat: str) -> None:
        self._pat = pat

    def get_auth_header(self) -> str:
        return f"Bearer {self._pat}"

    def get_token_for_git(self) -> str:
        return self._pat

    @property
    def mode_label(self) -> str:
        return "PAT"


class GitHubAppAuth(_Auth):

    _REFRESH_BUFFER_SECONDS = 300   # refresh 5 min before expiry
    _JWT_EXPIRY_SECONDS = 540       # 9 min (GitHub max is 10 min)

    def __init__(
        self,
        app_id: int,
        private_key_pem: str,
        installation_id: int,
        api_base: str,
    ) -> None:
        self._app_id = app_id
        self._private_key_pem = private_key_pem
        self._installation_id = installation_id
        self._api_base = api_base.rstrip("/")
        self._token: str = ""
        self._token_expires_at: float = 0.0
        self._lock = threading.Lock()
        self._check_dependencies()

    @staticmethod
    def _check_dependencies() -> None:
        try:
            __import__("jwt")
        except ImportError:
            raise RuntimeError(
                "GitHub App auth requires PyJWT and cryptography.\n"
                "Install with:  pip install PyJWT cryptography\n"
                "Then re-run the script."
            )

    def _generate_jwt(self) -> str:
        import jwt as pyjwt
        now = int(time.time())
        payload = {
            "iat": now - 60,                            # allow 60 s clock skew
            "exp": now + self._JWT_EXPIRY_SECONDS,
            "iss": str(self._app_id),
        }
        return pyjwt.encode(payload, self._private_key_pem, algorithm="RS256")

    def _refresh(self) -> None:

        jwt_token = self._generate_jwt()
        url = f"{self._api_base}/app/installations/{self._installation_id}/access_tokens"
        req = urllib.request.Request(
            url,
            data=b"{}",
            method="POST",
            headers={
                "Authorization": f"Bearer {jwt_token}",
                "Accept": "application/vnd.github+json",
                "X-GitHub-Api-Version": "2022-11-28",
            },
        )
        try:
            with urllib.request.urlopen(req, timeout=30) as resp:
                data = json.loads(resp.read())
        except urllib.error.HTTPError as exc:
            body = exc.read().decode(errors="replace")
            # Parse the body to give targeted, actionable guidance.
            try:
                err_msg = json.loads(body).get("message", "")
            except Exception:
                err_msg = body

            if exc.code == 401 and "must generate a public key" in err_msg:
                raise RuntimeError(
                    f"GitHub App authentication failed (HTTP 401): {err_msg}\n"
                    "  The App exists on GitHub but has no private key registered.\n"
                    "  The .pem file must be generated by GitHub, not created manually.\n\n"
                    "  Fix (3 steps):\n"
                    "    1. Open: GitHub → Settings → Developer settings → GitHub Apps → your app\n"
                    "    2. Scroll to ‘Private keys’ → click ‘Generate a private key’\n"
                    "    3. A real .pem file will download — replace the current file with it,\n"
                    "       then set auth.app.private_key_path in mirror-config.json accordingly."
                ) from exc

            if exc.code == 401:
                raise RuntimeError(
                    f"GitHub App authentication failed (HTTP 401): {err_msg}\n"
                    "  Verify auth.app.app_id and auth.app.installation_id in mirror-config.json."
                ) from exc

            raise RuntimeError(
                f"Failed to obtain GitHub App installation token (HTTP {exc.code}): {err_msg}"
            ) from exc

        self._token = data["token"]

        # expires_at: "2023-01-01T00:00:00Z"  -- strip Z for Python <3.11 compat
        expires_str = data["expires_at"].rstrip("Z")
        expires_dt = datetime.datetime.strptime(expires_str, "%Y-%m-%dT%H:%M:%S")
        # calendar.timegm interprets naive struct_time as UTC
        self._token_expires_at = (
            calendar.timegm(expires_dt.timetuple()) - self._REFRESH_BUFFER_SECONDS
        )
        log.info("GitHub App: installation token refreshed (valid ~55 min)")

    def _ensure_valid_token(self) -> str:
        with self._lock:
            if time.time() >= self._token_expires_at:
                self._refresh()
            return self._token

    def get_auth_header(self) -> str:
        return f"Bearer {self._ensure_valid_token()}"

    def get_token_for_git(self) -> str:
        # GitHub App tokens use "x-access-token" as the username for HTTPS auth.
        return f"x-access-token:{self._ensure_valid_token()}"

    def get_app_name(self) -> str:
        """Return the GitHub App name via a JWT-authenticated GET /app request.

        GET /app requires a JWT (not an installation token), so the shared
        GitHubClient cannot be used here -- it always sends the installation token.
        """
        try:
            jwt_token = self._generate_jwt()
            req = urllib.request.Request(
                f"{self._api_base}/app",
                headers={
                    "Authorization": f"Bearer {jwt_token}",
                    "Accept": "application/vnd.github+json",
                    "X-GitHub-Api-Version": "2022-11-28",
                },
            )
            with urllib.request.urlopen(req, timeout=15) as resp:
                data = json.loads(resp.read())
            return data.get("name") or data.get("slug", "")
        except Exception:
            return ""

    @property
    def mode_label(self) -> str:
        return "GitHub App"


# ===========================================================================
# Rate-limit-aware, retry-enabled GitHub REST client
# ===========================================================================

class GitHubClient:

    _BASE_BACKOFF = 2.0
    _MAX_BACKOFF = 600.0

    def __init__(self, auth: _Auth, api_base: str, max_retries: int = _DEFAULT_API_RETRIES) -> None:
        self._auth = auth
        self._api_base = api_base.rstrip("/")
        self._max_retries = max_retries
        self._rl_lock = threading.Lock()
        self._rl_remaining: int = 5000
        self._rl_reset_at: float = 0.0

    def _update_rate_limit(self, headers: dict) -> None:
        with self._rl_lock:
            try:
                self._rl_remaining = int(
                    headers.get("X-RateLimit-Remaining", self._rl_remaining)
                )
                self._rl_reset_at = float(
                    headers.get("X-RateLimit-Reset", self._rl_reset_at)
                )
            except (ValueError, TypeError):
                pass

    def _wait_if_rate_limited(self) -> None:
        with self._rl_lock:
            if self._rl_remaining > _RATE_LIMIT_BUFFER or self._rl_reset_at <= time.time():
                return
            wait = max(0.0, self._rl_reset_at - time.time() + 10)
        log.warning(
            f"API rate limit low ({self._rl_remaining} requests remaining). "
            f"Pausing {wait:.0f}s until quota resets..."
        )
        time.sleep(wait)

    def request(
        self,
        method: str,
        path: str,
        body: dict | None = None,
    ) -> tuple[int, dict]:
        """Execute a GitHub REST call. Returns (http_status, response_body_dict)."""
        url = f"{self._api_base}{path}"

        for attempt in range(self._max_retries):
            self._wait_if_rate_limited()

            data = json.dumps(body).encode() if body else None
            req = urllib.request.Request(
                url,
                data=data,
                method=method,
                headers={
                    "Authorization": self._auth.get_auth_header(),
                    "Accept": "application/vnd.github+json",
                    "Content-Type": "application/json",
                    "X-GitHub-Api-Version": "2022-11-28",
                },
            )

            try:
                with urllib.request.urlopen(req, timeout=30) as resp:
                    self._update_rate_limit(dict(resp.headers))
                    raw = resp.read()
                    # 204 No Content (and similar) return an empty body.
                    # json.loads(b"") raises JSONDecodeError, so guard here.
                    body_parsed: dict = json.loads(raw) if raw.strip() else {}
                    return resp.status, body_parsed

            except urllib.error.HTTPError as exc:
                resp_headers = dict(exc.headers) if exc.headers else {}
                self._update_rate_limit(resp_headers)
                try:
                    resp_body: dict = json.loads(exc.read())
                except Exception:
                    resp_body = {"message": str(exc)}

                msg = resp_body.get("message", "")

                # Secondary rate limit (403 with rate-limit body)
                if exc.code == 403 and (
                    "secondary rate limit" in msg.lower()
                    or "rate limit" in msg.lower()
                ):
                    backoff = min(
                        self._MAX_BACKOFF,
                        self._BASE_BACKOFF * (2 ** attempt) + random.uniform(0, 15),
                    )
                    log.warning(
                        f"Secondary rate limit hit "
                        f"(attempt {attempt + 1}/{self._max_retries}). "
                        f"Backing off {backoff:.0f}s..."
                    )
                    time.sleep(backoff)
                    continue

                # Primary rate limit
                if exc.code == 429:
                    retry_after = int(resp_headers.get("Retry-After", 60))
                    log.warning(f"429 Too Many Requests. Waiting {retry_after}s...")
                    time.sleep(retry_after)
                    continue

                # Transient server errors
                if exc.code in (500, 502, 503, 504):
                    backoff = min(self._MAX_BACKOFF, self._BASE_BACKOFF * (2 ** attempt))
                    log.warning(f"HTTP {exc.code} (transient). Retrying in {backoff:.0f}s...")
                    time.sleep(backoff)
                    continue

                # Non-retryable (400, 401, 404, 422, ...)
                return exc.code, resp_body

            except (urllib.error.URLError, OSError) as exc:
                backoff = min(self._MAX_BACKOFF, self._BASE_BACKOFF * (2 ** attempt))
                log.warning(f"Network error: {exc}. Retrying in {backoff:.0f}s...")
                time.sleep(backoff)

        raise RuntimeError(
            f"GitHub API call failed after {self._max_retries} retries: {method} {path}"
        )


# ===========================================================================
# Domain models
# ===========================================================================

@dataclass
class RepoSpec:
    namespace: str      # GitLab group / namespace
    project: str        # GitLab project name
    target_org: str     # GitHub organisation that will own this repo
    target_name: str    # Desired repo name on GitHub
    visibility: str = "private"
    branch_include: list[str] = field(default_factory=list)  # per-repo include patterns
    branch_exclude: list[str] = field(default_factory=list)  # per-repo exclude patterns
    branch_from_global: bool = False  # True when both columns were blank → inherited from global config
    ci_template: str = ""              # filename inside ci-templates/; blank → uses default.yml


@dataclass
class CiSkeletonConfig:
    """Settings for the CI skeleton workflow creation step (runs after each successful push)."""
    enabled: bool = True
    templates_dir: Path = field(default_factory=lambda: _SCRIPT_DIR / "ci-templates")
    target_path: str = ".github/workflows"   # folder path inside the repo (no trailing slash)
    commit_message: str = "chore: add CI skeleton workflow"
    branches: list[str] = field(default_factory=lambda: ["main", "master", "develop"])
    skip_if_exists: bool = True              # never overwrite an existing file at the resolved path


@dataclass
class MigrationConfig:
    auth: _Auth
    github_default_org: str  # fallback org when target_org is blank in repos.csv
    github_api_url: str
    github_url: str
    gitlab_url: str
    gitlab_pat: str = field(repr=False)  # never expose credential in repr / logs
    max_workers: int = 5
    clone_timeout: int = 7200
    push_timeout: int = 7200
    git_max_retries: int = 3
    dry_run: bool = False                              # log intent only; skip all writes
    rename_default_branch: bool = False                # if True, rename source default branch to 'main' on GitHub
    branch_include: list[str] = field(default_factory=list)  # regex include patterns; empty = all
    branch_exclude: list[str] = field(default_factory=list)  # regex exclude patterns; empty = none
    min_free_disk_gb: float = 10.0                     # min free temp-dir space before starting
    git_http_post_buffer: int = _DEFAULT_GIT_HTTP_POST_BUFFER  # bytes; 0 = git default (~1 MB)
    git_push_batch_size: int = _DEFAULT_PUSH_BATCH_SIZE        # refs per push; 0 = single push (not recommended for large repos)
    lfs_enabled: bool = True                           # fetch+push LFS objects when git-lfs is available
    detailed_commit_count: bool = False                # count commits via rev-list (slow on large repos)
    check_oversized_files: bool = True                 # scan for blobs > 100 MB before pushing (GitHub hard limit)
    ci_skeleton: CiSkeletonConfig | None = None        # CI skeleton creation config; None = disabled
    # "org/repo" -> {prop: raw_csv_val}; raw strings loaded from repo-properties.csv
    repo_custom_properties: dict[str, dict[str, str]] = field(default_factory=dict)
    # {org: {prop_name: value_type}}; fetched once per org in migrate_all, not from JSON
    _org_property_schemas: dict[str, dict[str, str]] = field(default_factory=dict, repr=False, compare=False)
    repos: list[RepoSpec] = field(default_factory=list)
    notification: "EmailConfig | None" = field(default=None, repr=False, compare=False)


@dataclass
class _RepoResult:
    source: str              # namespace/project
    target: str              # org/target_name
    status: str              # "succeeded" | "failed" | "skipped" | "dry-run"
    default_branch: str      # actual default branch (post-rename if applicable)
    error: str               # error message on failure (empty on success)
    duration_seconds: float  # wall-clock seconds for this repo
    completed_at: str        # "YYYY-MM-DD HH:MM:SS" human-readable timestamp
    visibility: str = ""              # visibility applied on GitHub
    gh_repo_url: str = ""             # clickable GitHub repo URL
    branch_count: int = 0             # branches pushed (post-filter)
    head_commit_sha: str = ""         # HEAD commit SHA from source bare clone
    gh_branch_count: int = 0          # branch count on GitHub after push
    gh_head_commit_sha: str = ""      # HEAD commit SHA on GitHub after push
    tag_count: int = 0                # tag count from source bare clone
    gh_tag_count: int = 0             # tag count on GitHub after push
    commit_count: int = 0             # commit count on default branch (source bare clone)
    gh_commit_count: int = 0          # commit count on default branch on GitHub after push
    default_branch_renamed: bool = False  # True if source default branch was renamed to 'main'
    validation_status: str = ""       # "match" | "mismatch" | "unknown" | "dry-run" | "empty"
    validation_notes: str = ""        # human-readable diff notes
    # Structured mismatch detail -- populated when validation_status == "mismatch"
    missing_branches: list[str] = field(default_factory=list)       # source branches absent on GitHub
    missing_tags: list[str] = field(default_factory=list)           # source tags absent on GitHub
    branch_sha_mismatches: list[str] = field(default_factory=list)  # "branch: src_sha != gh_sha"
    # LFS migration detail
    lfs_detected: bool = False          # True if LFS objects were found in the source repo
    lfs_object_count: int = 0           # number of LFS objects migrated
    # True when GitHub HEAD is strictly ahead of the migrated (GitLab) HEAD --
    # i.e. GitHub has extra commits such as .github CI skeleton on top.
    # Migration is still correct; HEAD SHA difference is expected and intentional.
    gh_head_is_ahead: bool = False
    # CI skeleton creation detail (populated after successful push when ci_skeleton is enabled)
    ci_skeleton_status: str = ""        # "created" | "partial" | "skipped_all" | "failed" | "not_configured" | "dry-run"
    ci_skeleton_branches_created: list[str] = field(default_factory=list)
    ci_skeleton_branches_skipped: list[str] = field(default_factory=list)
    ci_skeleton_error: str = ""
    # Custom properties detail
    custom_properties_status: str = ""   # "applied" | "failed" | "skipped" | "not_configured" | "dry-run"
    custom_properties_applied: dict = field(default_factory=dict)   # {prop: value} as applied
    custom_properties_error: str = ""


@dataclass
class _CiSkeletonResult:
    """Return type of _create_ci_skeleton."""
    status: str      # "created" | "partial" | "skipped_all" | "failed" | "not_configured" | "dry-run"
    branches_created: list[str] = field(default_factory=list)
    branches_skipped: list[str] = field(default_factory=list)
    error: str = ""


@dataclass
class _MirrorResult:
    """Return type of _mirror_repo -- avoids a wide positional tuple."""
    ok: bool
    default_branch: str = ""
    error: str = ""
    branch_count: int = 0
    head_commit_sha: str = ""
    gh_branch_count: int = 0
    gh_head_commit_sha: str = ""
    tag_count: int = 0
    gh_tag_count: int = 0
    commit_count: int = 0
    gh_commit_count: int = 0
    gh_repo_url: str = ""
    default_branch_renamed: bool = False
    validation_status: str = ""
    validation_notes: str = ""
    missing_branches: list[str] = field(default_factory=list)
    missing_tags: list[str] = field(default_factory=list)
    branch_sha_mismatches: list[str] = field(default_factory=list)
    lfs_detected: bool = False
    lfs_object_count: int = 0
    # True when GitHub HEAD is strictly *ahead* of GitLab HEAD on the default
    # branch, meaning GitHub has extra commits (e.g. .github CI skeleton) that
    # are not on GitLab. The migration is still correct -- GitLab history is
    # fully present on GitHub -- but the SHA comparison must not flag a mismatch.
    gh_head_is_ahead: bool = False


# ===========================================================================
# Checkpoint / Resume
# ===========================================================================

class MigrationState:

    def __init__(self, state_file: Path) -> None:
        self._file = state_file
        self._lock = threading.Lock()
        self._state: dict[str, str] = {}
        self._load()

    def _load(self) -> None:
        if self._file.exists():
            try:
                self._state = json.loads(self._file.read_text(encoding="utf-8"))
                n_done = sum(1 for v in self._state.values() if v == "succeeded")
                log.info(
                    f"Checkpoint found: {self._file.name} "
                    f"({n_done} succeeded, {len(self._state)} total entries)"
                )
            except Exception as exc:
                log.warning(f"Could not read checkpoint ({exc}). Starting fresh.")
                self._state = {}

    def _save(self) -> None:
        """Atomic write: write to .tmp then rename, so a crash never corrupts the file."""
        tmp = self._file.with_suffix(".tmp")
        tmp.write_text(json.dumps(self._state, indent=2), encoding="utf-8")
        tmp.replace(self._file)

    def is_succeeded(self, key: str) -> bool:
        with self._lock:
            # "partial" is NOT considered done -- re-run will retry the failed sub-steps.
            return self._state.get(key) == "succeeded"

    def record(self, key: str, status: str) -> None:
        with self._lock:
            self._state[key] = status
            self._save()


# ===========================================================================
# Git subprocess helper
# ===========================================================================

def _run_git(
    cmd: list[str],
    cwd: Path | None = None,
    timeout: int = 3600,
    log_cmd: list[str] | None = None,
    http_post_buffer: int = 0,
    stream_stderr: bool = False,
) -> tuple[int, str, str]:
    """Run a git command.

    When *stream_stderr* is True a background thread reads stderr line-by-line
    and logs each non-empty line at INFO level as it arrives, so long-running
    push/fetch operations show real-time progress instead of silence.
    """
    display = " ".join(log_cmd if log_cmd is not None else cmd)
    log.debug(f"git: {display}")

    env = os.environ.copy()
    # Prevent git from blocking on interactive credential prompts.
    env["GIT_TERMINAL_PROMPT"] = "0"
    env["GIT_ASKPASS"] = "echo"
    # Suppress OS credential managers (Keychain on macOS, Windows Credential Manager).
    # Optionally raise http.postBuffer to avoid "RPC failed; curl 55" on large pushes,
    # and disable the low-speed watchdog so slow-but-progressing transfers aren't killed.
    extra_cfg: list[tuple[str, str]] = [("credential.helper", "")]
    if http_post_buffer > 0:
        extra_cfg += [
            ("http.postBuffer", str(http_post_buffer)),
            ("http.lowSpeedLimit", "0"),
            ("http.lowSpeedTime", "999999"),
        ]
    env["GIT_CONFIG_COUNT"] = str(len(extra_cfg))
    for i, (key, val) in enumerate(extra_cfg):
        env[f"GIT_CONFIG_KEY_{i}"] = key
        env[f"GIT_CONFIG_VALUE_{i}"] = val

    # On Linux/macOS, launch git in a new session so it becomes its own process-group
    # leader (PGID == PID).  On timeout we can then send SIGKILL to the entire group,
    # killing git-remote-https, git-lfs workers, and any other helpers that would
    # otherwise become orphaned, hold sockets open, and exhaust resources when hundreds
    # of large repos are migrated concurrently.
    popen_kwargs: dict = {
        "stdout": subprocess.PIPE,
        "stderr": subprocess.PIPE,
        "env": env,
        **({"cwd": str(cwd.resolve())} if cwd is not None else {}),
        **(  # Windows: suppress console window; Linux/macOS: new session for group kill
            {"creationflags": 0x08000000}
            if sys.platform == "win32"
            else {"start_new_session": True}
        ),
    }

    try:
        with subprocess.Popen(cmd, **popen_kwargs) as proc:
            if stream_stderr:
                # Stream stderr line-by-line in a background thread so git's
                # progress output ("Writing objects: 45% ...") appears in the
                # log in real-time instead of being buffered until completion.
                # stdout is read in a separate thread so the polling loop is
                # never blocked waiting for it, and so a process that hangs
                # without ever closing stdout can still be killed on timeout.
                stderr_lines: list[str] = []
                stdout_chunks: list[bytes] = []

                def _read_stderr() -> None:
                    assert proc.stderr is not None
                    for raw in proc.stderr:
                        line = raw.decode("utf-8", errors="replace").rstrip()
                        if line:
                            stderr_lines.append(line)
                            log.info(f"  git> {line}")

                def _read_stdout() -> None:
                    assert proc.stdout is not None
                    while True:
                        chunk = proc.stdout.read(65536)
                        if not chunk:
                            break
                        stdout_chunks.append(chunk)

                reader = threading.Thread(target=_read_stderr, daemon=True)
                stdout_reader = threading.Thread(target=_read_stdout, daemon=True)
                reader.start()
                stdout_reader.start()
                try:
                    # Poll for process completion; never block on stdout/stderr directly.
                    deadline = time.monotonic() + timeout
                    while proc.poll() is None:
                        if time.monotonic() > deadline:
                            # Timeout: kill process group first, then reap.
                            if sys.platform != "win32":
                                try:
                                    os.killpg(proc.pid, signal.SIGKILL)
                                except (ProcessLookupError, PermissionError, OSError):
                                    pass
                            proc.kill()
                            # Join reader threads (pipe EOF after kill) then
                            # explicitly wait() to prevent zombie processes on Linux.
                            reader.join(timeout=5)
                            stdout_reader.join(timeout=5)
                            try:
                                proc.wait(timeout=5)
                            except subprocess.TimeoutExpired:
                                pass
                            return 1, "", f"git timed out after {timeout}s: {display}"
                        time.sleep(0.25)
                    reader.join(timeout=10)
                    stdout_reader.join(timeout=10)
                    stdout_b = b"".join(stdout_chunks)
                    stderr_str = "\n".join(stderr_lines)
                except Exception:
                    proc.kill()
                    reader.join(timeout=5)
                    stdout_reader.join(timeout=5)
                    raise
            else:
                try:
                    stdout_b, stderr_b = proc.communicate(timeout=timeout)
                except subprocess.TimeoutExpired:
                    # Kill the entire process group to reap all spawned git helpers.
                    if sys.platform != "win32":
                        try:
                            os.killpg(proc.pid, signal.SIGKILL)
                        except (ProcessLookupError, PermissionError, OSError):
                            pass
                    proc.kill()
                    proc.communicate()  # drain pipes; prevents zombie / resource leak
                    return 1, "", f"git timed out after {timeout}s: {display}"
                stderr_str = stderr_b.decode("utf-8", errors="replace").strip()
            return (
                proc.returncode,
                stdout_b.decode("utf-8", errors="replace").strip(),
                stderr_str,
            )
    except FileNotFoundError:
        return 1, "", "'git' not found on PATH. Install: https://git-scm.com/downloads"


# ===========================================================================
# GitHub repo operations
# ===========================================================================

def _create_github_repo(
    client: GitHubClient,
    org: str,
    name: str,
    visibility: str,
    custom_properties: "dict[str, str | list[str] | None] | None" = None,
) -> tuple[bool, bool]:
    """Create the GitHub repository.

    Returns (success, was_newly_created).
    was_newly_created=False means the repo already existed (re-run scenario).

    *custom_properties* must already have values coerced to the correct wire type
    (str for string/single_select/true_false/url, list[str] for multi_select,
    None for clearing a property). Injected verbatim into the creation payload
    so that orgs with required properties don't reject with HTTP 422.
    """
    _body: dict = {
        "name": name,
        "private": visibility != "public",
        "auto_init": False,
        "has_issues": False,
        "has_projects": False,
        "has_wiki": False,
    }
    if custom_properties:
        _body["custom_properties"] = custom_properties
        log.debug(
            f"[{name}] Injecting {len(custom_properties)} custom propert"
            f"{'y' if len(custom_properties) == 1 else 'ies'} into repo creation payload"
        )
    status, resp = client.request(
        "POST",
        f"/orgs/{org}/repos",
        body={**_body, "visibility": visibility},
    )
    if status in (200, 201):
        log.info(f"[{name}] GitHub repo created ({visibility})")
        return True, True

    msg = resp.get("message", "unknown")
    errors = resp.get("errors", [])

    # GitHub returns HTTP 422 with errors[].code == "already_exists" or
    # errors[].message containing "already" when the repo already exists.
    # The top-level message is "Repository creation failed." -- NOT "already exists" --
    # so we must also inspect the errors array.
    def _repo_already_exists(message: str, errs: list) -> bool:
        if "already exists" in message.lower():
            return True
        for e in errs:
            code = (e.get("code") or "").lower()
            emsg = (e.get("message") or "").lower()
            if code == "already_exists" or "already" in emsg:
                return True
        return False

    if status == 422 and _repo_already_exists(msg, errors):
        log.info(f"[{name}] GitHub repo already exists -- will push without deleting remote refs")
        return True, False

    # HTTP 422 "does not have permission to set custom properties" -- the token
    # lacks organization_custom_properties:write scope.  Retry WITHOUT the
    # custom_properties field; they will be applied separately via PATCH after
    # the push (--post-migration) or skipped with a warning if not configured.
    def _is_custom_properties_permission_error(message: str) -> bool:
        return "permission" in message.lower() and "custom properties" in message.lower()

    if status == 422 and custom_properties and _is_custom_properties_permission_error(msg):
        log.warning(
            f"[{name}] Token lacks permission to set custom properties at creation time "
            f"(HTTP 422: {msg}). Retrying repo creation without custom properties. "
            "Apply them afterwards with: python GitMirrorMigration.py --post-migration"
        )
        status, resp = client.request(
            "POST",
            f"/orgs/{org}/repos",
            body={k: v for k, v in {**_body, "visibility": visibility}.items()
                  if k != "custom_properties"},
        )
        if status in (200, 201):
            log.info(f"[{name}] GitHub repo created ({visibility}) [without custom properties]")
            return True, True
        if status == 422 and _repo_already_exists(
            resp.get("message", ""), resp.get("errors", [])
        ):
            log.info(f"[{name}] GitHub repo already exists -- will push without deleting remote refs")
            return True, False
        msg = resp.get("message", "unknown")

    # org endpoint 404 means the PAT belongs to a user account, not an org.
    if status == 404:
        log.debug(f"[{name}] Org endpoint returned 404 -- trying user-level creation")
        status, resp = client.request(
            "POST",
            "/user/repos",
            body={**_body, "visibility": "private" if visibility == "internal" else visibility},
        )
        if status in (200, 201):
            log.info(f"[{name}] GitHub repo created under user account ({visibility})")
            return True, True
        if status == 422 and _repo_already_exists(
            resp.get("message", ""), resp.get("errors", [])
        ):
            log.info(f"[{name}] GitHub user repo already exists")
            return True, False
        msg = resp.get("message", "unknown")

    log.error(f"[{name}] Failed to create GitHub repo (HTTP {status}): {msg}")
    return False, False


def _get_default_branch(mirror_dir: Path) -> str:
    """Read the HEAD symbolic-ref from a bare clone to get the source default branch.

    Returns the branch name exactly as set on the source (e.g. 'main', 'master',
    'develop', 'trunk').  Returns empty string if unresolvable.
    """
    code, out, _ = _run_git(
        ["git", "symbolic-ref", "HEAD"],
        cwd=mirror_dir,
        timeout=15,
    )
    if code != 0 or not out:
        return ""
    prefix = "refs/heads/"
    return out[len(prefix):] if out.startswith(prefix) else out


def _count_git_refs(mirror_dir: Path, prefix: str) -> int:
    """Count git refs under the given prefix (e.g. 'refs/heads/', 'refs/tags/')."""
    code, out, _ = _run_git(
        ["git", "for-each-ref", "--format=%(refname:short)", prefix],
        cwd=mirror_dir,
        timeout=30,
    )
    return sum(1 for r in out.splitlines() if r.strip()) if code == 0 else 0


def _count_git_commits(mirror_dir: Path) -> int:
    """Return the total commit count reachable from HEAD in a bare clone."""
    code, out, _ = _run_git(
        ["git", "rev-list", "--count", "HEAD"],
        cwd=mirror_dir,
        timeout=60,
    )
    try:
        return int(out.strip()) if code == 0 else 0
    except ValueError:
        return 0


def _get_head_commit_sha(mirror_dir: Path, branch: str) -> str:
    """Return the full commit SHA at refs/heads/<branch> in a bare clone."""
    if not branch:
        return ""
    code, out, _ = _run_git(
        ["git", "rev-parse", f"refs/heads/{branch}"],
        cwd=mirror_dir,
        timeout=15,
    )
    return out.strip() if code == 0 else ""


def _get_git_ref_names(mirror_dir: Path, prefix: str) -> set[str]:
    """Return the set of short ref names under *prefix* in a bare clone.

    E.g. prefix='refs/heads/' -> {'main', 'develop', 'release/1.0'}
         prefix='refs/tags/'  -> {'v1.0.0', 'v2.0.0-rc1'}
    Returns an empty set on any git error.
    """
    code, out, _ = _run_git(
        ["git", "for-each-ref", "--format=%(refname:short)", prefix],
        cwd=mirror_dir,
        timeout=30,
    )
    return {r.strip() for r in out.splitlines() if r.strip()} if code == 0 else set()


def _get_git_branch_shas(mirror_dir: Path) -> dict[str, str]:
    """Return {branch_name: full_commit_sha} for every branch in a bare clone."""
    code, out, _ = _run_git(
        ["git", "for-each-ref", "--format=%(refname:short) %(objectname)", "refs/heads/"],
        cwd=mirror_dir,
        timeout=30,
    )
    result: dict[str, str] = {}
    if code == 0:
        for line in out.splitlines():
            parts = line.strip().split(maxsplit=1)
            if len(parts) == 2:
                result[parts[0]] = parts[1]
    return result


def _set_github_default_branch(
    client: GitHubClient,
    org: str,
    name: str,
    branch: str,
) -> None:
    """Set the GitHub repo's default branch to match the source."""
    status, resp = client.request(
        "PATCH",
        f"/repos/{org}/{name}",
        body={"default_branch": branch},
    )
    if status == 200:
        log.info(f"[{name}] Default branch set to '{branch}'")
    else:
        log.warning(
            f"[{name}] Could not set default branch to '{branch}' "
            f"(HTTP {status}: {resp.get('message', 'unknown')})"
        )


def _get_github_actor(client: GitHubClient, auth: _Auth) -> str:
    """Return the login/name of the token owner.

    For PAT auth  → GET /user  (installation token) → login
    For App auth  → GET /app   (JWT required)         → app name / slug
    Returns an empty string on any failure so it never blocks the migration.
    """
    if isinstance(auth, GitHubAppAuth):
        return auth.get_app_name()
    st, resp = client.request("GET", "/user")
    if st == 200 and isinstance(resp, dict):
        return resp.get("login", "")
    return ""


def _get_github_list_count(client: GitHubClient, path: str) -> int:
    """Return the paginated item count at the given GitHub REST list endpoint."""
    page, total = 1, 0
    while True:
        st, resp = client.request("GET", f"{path}?per_page=100&page={page}")
        if st != 200 or not isinstance(resp, list):
            break
        total += len(resp)
        if len(resp) < 100:
            break
        page += 1
    return total


def _get_github_head_commit_sha(
    client: GitHubClient, org: str, repo: str, branch: str
) -> str:
    """Return the HEAD commit SHA of a branch via the GitHub Git refs API."""
    if not branch or branch == "dry-run":
        return ""
    st, resp = client.request("GET", f"/repos/{org}/{repo}/git/ref/heads/{branch}")
    if st == 200 and isinstance(resp, dict):
        return resp.get("object", {}).get("sha", "")
    return ""


def _get_github_commit_count(client: GitHubClient, org: str, repo: str, sha: str) -> int:
    """Return the commit count on *sha* via the GitHub commits API.

    Strategy: GET /repos/{org}/{repo}/commits?sha={sha}&per_page=1
    GitHub returns a ``Link`` header whose ``rel="last"`` URL contains
    ``page=N`` — that N is the total commit count (one commit per page).
    Falls back to 0 on any error (the caller treats 0 as 'unknown').

    Note: GitHub caps this at ~10 000 for very large histories, but it is
    accurate enough for display purposes.
    """
    if not sha:
        return 0
    url = f"{client._api_base}/repos/{org}/{repo}/commits?sha={sha}&per_page=1"
    req = urllib.request.Request(
        url,
        headers={
            "Authorization": client._auth.get_auth_header(),
            "Accept": "application/vnd.github+json",
            "X-GitHub-Api-Version": "2022-11-28",
        },
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            link = resp.headers.get("Link", "")
            # Link header format (RFC 5988):
            # <https://api.github.com/...?page=347>; rel="last"
            import re as _re
            m = _re.search(r'[?&]page=(\d+)>[^,]*rel="last"', link)
            if m:
                return int(m.group(1))
            # No Link header means the entire history fits on one page.
            body = json.loads(resp.read())
            return len(body) if isinstance(body, list) else 0
    except Exception:
        return 0


def _get_github_ref_names(client: GitHubClient, path: str) -> set[str]:
    """Return all 'name' values from a paginated GitHub list endpoint.

    Uses the same pagination logic as _get_github_list_count but returns names
    instead of a count, enabling set-based comparisons against source refs.
    """
    page, names = 1, set()
    while True:
        st, resp = client.request("GET", f"{path}?per_page=100&page={page}")
        if st != 200 or not isinstance(resp, list):
            break
        for item in resp:
            n = item.get("name", "")
            if n:
                names.add(n)
        if len(resp) < 100:
            break
        page += 1
    return names


def _get_github_branch_shas(client: GitHubClient, org: str, repo: str) -> dict[str, str]:
    """Return {branch_name: commit_sha} for every branch on the GitHub repo.

    The GitHub branches list endpoint includes 'commit.sha' for each entry,
    so we get both names and SHAs in a single paginated pass.
    """
    page, result = 1, {}
    while True:
        st, resp = client.request(
            "GET", f"/repos/{org}/{repo}/branches?per_page=100&page={page}"
        )
        if st != 200 or not isinstance(resp, list):
            break
        for item in resp:
            branch_name = item.get("name", "")
            sha = item.get("commit", {}).get("sha", "")
            if branch_name and sha:
                result[branch_name] = sha
        if len(resp) < 100:
            break
        page += 1
    return result


def _compute_validation(
    branch_count: int,
    head_sha: str,
    gh_branch_count: int,
    gh_head_sha: str,
    tag_count: int = 0,
    gh_tag_count: int = 0,
    commit_count: int = 0,
    gh_commit_count: int = 0,
    missing_branches: list[str] | None = None,
    missing_tags: list[str] | None = None,
    branch_sha_mismatches: list[str] | None = None,
) -> tuple[str, str]:
    """Compare source bare-clone data against GitHub post-push data.

    Name-level checks (missing_branches, missing_tags, branch_sha_mismatches) are
    the primary, authoritative criteria.  Count-level checks serve as a fallback
    when the name-level data could not be gathered (API failure).

    Returns (validation_status, notes):
        "match"    -- all checks pass
        "mismatch" -- at least one value differs (notes describes what)
        "unknown"  -- GitHub data was not available (API error)
    """
    _mb  = missing_branches or []
    _mt  = missing_tags or []
    _bsm = branch_sha_mismatches or []

    # If no GitHub data came back at all, we can't validate.
    api_unavailable = (
        not gh_branch_count
        and not gh_head_sha
        and not _mb
        and not _mt
        and not _bsm
    )
    if api_unavailable:
        return "unknown", "GitHub validation data unavailable -- API may have failed"

    notes: list[str] = []

    # -- Primary: name-level checks -----------------------------------------
    if _mb:
        sample = ", ".join(_mb[:8])
        suffix = f" (+{len(_mb) - 8} more)" if len(_mb) > 8 else ""
        notes.append(f"missing branches ({len(_mb)}): {sample}{suffix}")

    if _mt:
        sample = ", ".join(_mt[:8])
        suffix = f" (+{len(_mt) - 8} more)" if len(_mt) > 8 else ""
        notes.append(f"missing tags ({len(_mt)}): {sample}{suffix}")

    if _bsm:
        sample = "; ".join(_bsm[:5])
        suffix = f" (+{len(_bsm) - 5} more)" if len(_bsm) > 5 else ""
        notes.append(f"branch SHA mismatches ({len(_bsm)}): {sample}{suffix}")

    # -- Fallback: count/SHA checks (when name-level data is empty) ----------
    if not _mb and branch_count and gh_branch_count and branch_count != gh_branch_count:
        notes.append(f"branch count: source={branch_count} github={gh_branch_count}")

    if not _mt and tag_count and gh_tag_count and tag_count != gh_tag_count:
        notes.append(f"tag count: source={tag_count} github={gh_tag_count}")

    if head_sha and gh_head_sha and head_sha != gh_head_sha:
        notes.append(f"default-branch HEAD SHA: source={head_sha[:8]}… github={gh_head_sha[:8]}…")

    # Commit count is informational only: if HEAD SHAs match the full history is
    # present (git is content-addressed), so a count difference is never grounds
    # for failure but we surface it in notes for transparency.
    if commit_count and gh_commit_count and commit_count != gh_commit_count:
        notes.append(
            f"commit count on default branch: source={commit_count} github={gh_commit_count}"
            " (informational -- HEAD SHA is the authoritative check)"
        )

    return ("mismatch", "; ".join(notes)) if notes else ("match", "")


# ===========================================================================
# Git retry helper and branch filter
# ===========================================================================

def _git_with_retry(
    cmd: list[str],
    log_cmd: list[str],
    timeout: int,
    max_retries: int,
    label: str,
    action: str,
    cwd: Path | None = None,
    cleanup_dir: Path | None = None,
    http_post_buffer: int = 0,
    cmd_factory: Callable[[], tuple[list[str], list[str]]] | None = None,
    stream_stderr: bool = False,
) -> tuple[bool, str]:
    """Run a git command with retries and capped exponential backoff.

    If *cmd_factory* is provided it is called fresh before every attempt to
    produce a new (cmd, log_cmd) pair.  This is used for push operations so the
    GitHub auth token embedded in the push URL is always current (GitHub App
    installation tokens expire in 55 min; a large repo push with retries can
    exceed this window).

    Set *stream_stderr* to True for push/fetch operations to relay git's
    progress output to the log in real-time.
    """
    err = ""
    for attempt in range(max_retries):
        # Always get a fresh cmd from the factory (re-fetches expiring tokens).
        _cmd, _log_cmd = cmd_factory() if cmd_factory is not None else (cmd, log_cmd)
        code, _, err = _run_git(
            _cmd, log_cmd=_log_cmd, cwd=cwd, timeout=timeout,
            http_post_buffer=http_post_buffer, stream_stderr=stream_stderr,
        )
        if code == 0:
            return True, ""
        if attempt < max_retries - 1:
            backoff = min(300, 30 * (2 ** attempt))  # 30s → 60s → 120s … capped at 5 min
            # GitHub HTTP 500 ("Internal Server Error") means server-side pack-processing
            # failure or transient overload -- needs longer recovery time than network blips.
            if "internal server error" in err.lower():
                backoff = max(180, backoff)
            log.warning(
                f"[{label}] {action} failed (attempt {attempt + 1}/{max_retries}). "
                f"Retrying in {backoff}s: {err}"
            )
            if cleanup_dir is not None:
                shutil.rmtree(str(cleanup_dir), ignore_errors=True)
            time.sleep(backoff)
    return False, f"{action} failed after {max_retries} attempts: {err}"


def _filter_branches(
    mirror_dir: Path,
    include_patterns: list[str],
    exclude_patterns: list[str],
    label: str,
) -> None:
    if not include_patterns and not exclude_patterns:
        return
    code, out, _ = _run_git(
        ["git", "for-each-ref", "--format=%(refname:short)", "refs/heads/"],
        cwd=mirror_dir,
        timeout=30,
    )
    if code != 0 or not out:
        return
    branches = [b.strip() for b in out.splitlines() if b.strip()]
    if not branches:
        return

    default = _get_default_branch(mirror_dir)
    kept = list(branches)
    if include_patterns:
        kept = [b for b in kept if any(re.search(p, b) for p in include_patterns)]
    if exclude_patterns:
        kept = [b for b in kept if not any(re.search(p, b) for p in exclude_patterns)]

    # Always preserve the default branch regardless of filter outcome.
    if default and default not in kept:
        kept.append(default)
        log.debug(f"[{label}] Default branch '{default}' preserved despite filter rules")

    to_delete = set(branches) - set(kept)
    if not to_delete:
        return

    sample = ", ".join(sorted(to_delete)[:8])
    suffix = f", +{len(to_delete) - 8} more" if len(to_delete) > 8 else ""
    log.info(f"[{label}] Branch filter: {len(kept)} kept, {len(to_delete)} removed ({sample}{suffix})")
    for branch in to_delete:
        c, _, e = _run_git(
            ["git", "update-ref", "-d", f"refs/heads/{branch}"],
            cwd=mirror_dir,
            timeout=15,
        )
        if c != 0:
            log.warning(f"[{label}] Could not remove branch '{branch}': {e}")


# ===========================================================================
# CI skeleton -- config validation and per-repo file creation
# ===========================================================================

def _validate_ci_skeleton_config(
    ci_cfg: CiSkeletonConfig,
    repos: list[RepoSpec],
) -> list[str]:
    """Return a list of human-readable problems found in the CI skeleton configuration.

    Checks performed:
    - ci-templates/ directory exists
    - Every ci_template filename referenced in repos.csv exists in that directory
    - default.yml exists when any repo has a blank ci_template column
    """
    problems: list[str] = []
    if not ci_cfg.templates_dir.is_dir():
        problems.append(
            f"ci_skeleton.templates_dir does not exist: {ci_cfg.templates_dir}. "
            "Create the directory and add your CI template YAML files."
        )
        return problems  # no point checking individual files if the dir is absent

    templates_referenced: set[str] = set()
    needs_default = False
    for spec in repos:
        fname = (spec.ci_template or "").strip()
        if fname:
            templates_referenced.add(fname)
        else:
            needs_default = True

    for fname in sorted(templates_referenced):
        if not (ci_cfg.templates_dir / fname).exists():
            problems.append(
                f"CI template file not found: {ci_cfg.templates_dir.name}/{fname}. "
                f"Add the file or correct the ci_template column in repos.csv."
            )

    if needs_default and not (ci_cfg.templates_dir / "default.yml").exists():
        problems.append(
            f"One or more repos have no ci_template set in repos.csv, but "
            f"{ci_cfg.templates_dir.name}/default.yml does not exist. "
            "Create default.yml or set ci_template for every row."
        )

    return problems


def _create_ci_skeleton(
    spec: RepoSpec,
    config: MigrationConfig,
    client: GitHubClient,
    default_branch: str,
) -> _CiSkeletonResult:
    """Create the CI skeleton workflow file on each configured branch of the GitHub repo.

    Resolution order for the template file:
      1. spec.ci_template (from repos.csv ci_template column) -- if non-blank
      2. default.yml in ci_cfg.templates_dir -- fallback when ci_template is blank

    Placeholders substituted in the template:
      {{repo_name}}       -- target repo name on GitHub
      {{org}}             -- target GitHub organisation
      {{default_branch}}  -- actual default branch of the repo

    The resolved file path inside the repo is:
      {ci_cfg.target_path}/{template_filename}
    e.g. .github/workflows/java-maven.yml

    When skip_if_exists=true, the file is only created on branches where
    the exact resolved path does not yet exist (checked via GET /contents).
    """
    ci_cfg = config.ci_skeleton
    if ci_cfg is None or not ci_cfg.enabled:
        return _CiSkeletonResult(status="not_configured")

    label = spec.target_name

    # Resolve template filename and load content.
    template_filename = (spec.ci_template or "").strip() or "default.yml"
    template_path = ci_cfg.templates_dir / template_filename
    if not template_path.exists():
        err = (
            f"Template file not found: {ci_cfg.templates_dir.name}/{template_filename}. "
            "Add the file or correct the ci_template column."
        )
        log.error(f"[{label}] CI skeleton: {err}")
        return _CiSkeletonResult(status="failed", error=err)

    try:
        raw_content = template_path.read_text(encoding="utf-8")
    except OSError as exc:
        err = f"Cannot read template '{template_filename}': {exc}"
        log.error(f"[{label}] CI skeleton: {err}")
        return _CiSkeletonResult(status="failed", error=err)

    # Substitute known placeholders.
    content = (
        raw_content
        .replace("{{repo_name}}", spec.target_name)
        .replace("{{org}}", spec.target_org)
        .replace("{{default_branch}}", default_branch or "main")
    )
    content_b64 = base64.b64encode(content.encode("utf-8")).decode("ascii")

    # Target file path inside the repo (no leading slash).
    target_file = f"{ci_cfg.target_path.strip('/')}/{template_filename}"

    if config.dry_run:
        log.info(
            f"[{label}] CI skeleton: DRY RUN -- would create '{target_file}' "
            f"using template '{template_filename}' on branches: {ci_cfg.branches}"
        )
        return _CiSkeletonResult(status="dry-run")

    # Determine which configured branches actually exist in this repo.
    gh_branch_names = _get_github_ref_names(
        client, f"/repos/{spec.target_org}/{spec.target_name}/branches"
    )
    target_branches = [b for b in ci_cfg.branches if b in gh_branch_names]

    if not target_branches:
        log.info(
            f"[{label}] CI skeleton: none of the configured branches "
            f"({', '.join(ci_cfg.branches)}) exist in this repo -- skipped"
        )
        return _CiSkeletonResult(status="skipped_all")

    log.info(
        f"[{label}] CI skeleton: creating '{target_file}' "
        f"(template: {template_filename}) on {len(target_branches)} branch(es): "
        f"{', '.join(target_branches)}"
    )

    branches_created: list[str] = []
    branches_skipped: list[str] = []
    first_error: str = ""

    # URL-encode the file path once (forward slashes are valid in GitHub's
    # contents API path segment so we preserve them; only special chars encoded).
    file_path_enc = "/".join(
        urllib.parse.quote(part, safe="") for part in target_file.split("/")
    )
    contents_base = f"/repos/{spec.target_org}/{spec.target_name}/contents/{file_path_enc}"

    for branch in target_branches:
        branch_enc = urllib.parse.quote(branch, safe="")

        if ci_cfg.skip_if_exists:
            st, _ = client.request("GET", f"{contents_base}?ref={branch_enc}")
            if st == 200:
                log.info(
                    f"[{label}] CI skeleton: '{target_file}' already exists on "
                    f"branch '{branch}' -- skipped (skip_if_exists=true)"
                )
                branches_skipped.append(branch)
                continue
            # 404 = file absent, proceed; any other status is unexpected but we proceed anyway.

        st, resp = client.request(
            "PUT",
            contents_base,
            body={
                "message": ci_cfg.commit_message,
                "content": content_b64,
                "branch": branch,
            },
        )
        if st in (200, 201):
            log.info(
                f"[{label}] CI skeleton: created '{target_file}' on branch '{branch}'"
            )
            branches_created.append(branch)
        else:
            msg = resp.get("message", "unknown") if isinstance(resp, dict) else str(resp)
            log.warning(
                f"[{label}] CI skeleton: failed to create '{target_file}' "
                f"on branch '{branch}' (HTTP {st}): {msg}"
            )
            if not first_error:
                first_error = f"HTTP {st} on branch '{branch}': {msg}"

    # Aggregate status: created | partial | skipped_all | failed.
    if branches_created and not first_error:
        status = "created"
    elif branches_created and (branches_skipped or first_error):
        status = "partial"
    elif branches_skipped and not branches_created and not first_error:
        status = "skipped_all"
    else:
        status = "failed"

    log.info(
        f"[{label}] CI skeleton summary: "
        f"template='{template_filename}', "
        f"created={len(branches_created)}, "
        f"skipped(exists)={len(branches_skipped)}"
        + (f", error='{first_error}'" if first_error else "")
    )

    return _CiSkeletonResult(
        status=status,
        branches_created=branches_created,
        branches_skipped=branches_skipped,
        error=first_error,
    )


# ===========================================================================
# Oversized blob detector
# ===========================================================================

def _find_oversized_blobs(mirror_dir: Path, label: str, timeout: int = 180) -> int:
    """Return the count of git blob objects that exceed GitHub's per-file size limit.

    Uses 'git cat-file --batch-check --batch-all-objects' to inspect every
    object in the bare clone without checking it out.  Only blob (file content)
    objects are examined; commits, trees, and tags are skipped.

    A non-zero return value means the push WILL be rejected by GitHub; the
    repository must be rewritten before migration can proceed.
    Returns 0 on any git error (conservative: lets the push attempt continue).
    """
    limit_bytes = _GITHUB_FILE_SIZE_LIMIT_MB * 1024 * 1024
    code, out, _ = _run_git(
        [
            "git", "cat-file",
            "--batch-check=%(objecttype) %(objectsize)",
            "--batch-all-objects",
        ],
        cwd=mirror_dir,
        timeout=timeout,
    )
    if code != 0 or not out:
        return 0
    count = 0
    for line in out.splitlines():
        parts = line.split(None, 1)
        if len(parts) == 2 and parts[0] == "blob":
            try:
                if int(parts[1]) > limit_bytes:
                    count += 1
            except ValueError:
                pass
    if count:
        log.warning(
            f"[{label}] {count} blob object(s) exceed {_GITHUB_FILE_SIZE_LIMIT_MB} MB -- "
            "GitHub will reject the push. "
            "Rewrite history with 'git filter-repo --strip-blobs-bigger-than 100M'."
        )
    return count


# ===========================================================================
# Core per-repo mirror operation
# ===========================================================================

def _github_is_ahead(
    client: "GitHubClient",
    org: str,
    repo_name: str,
    base_sha: str,
    head_sha: str,
) -> bool:
    """Return True if *head_sha* (GitHub) is strictly ahead of *base_sha* (GitLab).

    Uses GitHub's compare endpoint: GET /repos/{org}/{repo}/compare/{base}...{head}.
    ``status == "ahead"`` means GitHub has extra commits on top of the GitLab
    history and those GitLab commits are all present -- safe to skip re-push.
    Any other status ("behind", "diverged") or an API error returns False so the
    caller proceeds with the normal clone+push path (conservative fall-through).
    """
    if not base_sha or not head_sha or base_sha == head_sha:
        return False
    try:
        status_code, data = client.request(
            "GET", f"/repos/{org}/{repo_name}/compare/{base_sha}...{head_sha}"
        )
        if status_code == 200 and isinstance(data, dict):
            return data.get("status") == "ahead"
    except Exception:
        pass
    return False


def _ls_remote_refs(
    clone_url: str,
    log_url: str,
    label: str,
    timeout: int = 120,
) -> tuple[str, dict[str, str], set[str]]:
    """Fetch only ref metadata from a remote via git ls-remote (no object transfer).

    Returns (default_branch, {branch_name: sha}, {tag_name}).
    All three values will be empty/falsy on failure.
    """
    rc, out, _ = _run_git(
        ["git", "ls-remote", "--symref", clone_url, "HEAD", "refs/heads/*", "refs/tags/*"],
        log_cmd=["git", "ls-remote", "--symref", log_url, "HEAD", "refs/heads/*", "refs/tags/*"],
        timeout=timeout,
    )
    if rc != 0:
        log.debug(f"[{label}] git ls-remote failed (rc={rc}) -- will fall back to full clone")
        return "", {}, set()

    default_branch = ""
    branch_shas: dict[str, str] = {}
    tag_names: set[str] = set()

    for line in out.splitlines():
        line = line.strip()
        if not line:
            continue
        if line.startswith("ref:"):
            # "ref: refs/heads/main\tHEAD"
            ref_target = line[len("ref:"):].strip().split("\t", 1)[0].strip()
            if ref_target.startswith("refs/heads/"):
                default_branch = ref_target[len("refs/heads/"):]
        else:
            parts = line.split("\t", 1)
            if len(parts) != 2:
                continue
            sha, ref = parts[0].strip(), parts[1].strip()
            if ref.endswith("^{}"):
                continue  # dereferenced tag object -- skip
            if ref.startswith("refs/heads/"):
                branch_shas[ref[len("refs/heads/"):]] = sha
            elif ref.startswith("refs/tags/"):
                tag_names.add(ref[len("refs/tags/"):])

    return default_branch, branch_shas, tag_names


def _mirror_repo(
    spec: RepoSpec,
    config: MigrationConfig,
    client: GitHubClient,
) -> _MirrorResult:

    gitlab_base = config.gitlab_url.rstrip("/")
    github_base = config.github_url.rstrip("/")
    name = spec.target_name

    # Authenticated URLs -- NEVER passed to log or log_cmd.
    gl_scheme, gl_rest = gitlab_base.split("://", 1)
    clone_url = (
        f"{gl_scheme}://oauth2:{config.gitlab_pat}@{gl_rest}"
        f"/{spec.namespace}/{spec.project}.git"
    )
    safe_clone = f"{gitlab_base}/{spec.namespace}/{spec.project}.git"
    safe_push  = f"{github_base}/{spec.target_org}/{name}.git"

    if config.dry_run:
        log.info(f"[{name}] DRY RUN -- would clone {safe_clone} -> {safe_push}")
        return _MirrorResult(ok=True, default_branch="dry-run", validation_status="dry-run")

    # Sanitize project name for safe temp-dir path (Windows + Linux safe).
    safe_dir = _safe_dir_name(spec.project)
    tmp_dir = Path(tempfile.mkdtemp(prefix="gitmirror_"))
    mirror_dir = tmp_dir / f"{safe_dir}.git"

    try:
        # Per-repo disk space check: concurrent workers can drain disk mid-run.
        with _disk_lock:
            free_gb = shutil.disk_usage(tmp_dir).free / (1024 ** 3)
            _disk_ok = free_gb >= config.min_free_disk_gb
        if not _disk_ok:
            return _MirrorResult(
                ok=False,
                error=(
                    f"Insufficient disk space before cloning: {free_gb:.1f} GB free, "
                    f"need at least {config.min_free_disk_gb:.1f} GB. "
                    "Free up space or lower migration.min_free_disk_gb."
                ),
            )

        # Create GitHub repo if it doesn't exist yet.
        _raw_props = config.repo_custom_properties.get(f"{spec.target_org}/{spec.target_name}")
        _creation_props: "dict[str, str | list[str] | None] | None" = None
        if _raw_props:
            _schema = config._org_property_schemas.get(spec.target_org, {})
            _creation_props = {
                k: _coerce_property_value(v, _schema.get(k, "string"), prop_name=k)
                for k, v in _raw_props.items()
            }
        repo_created, newly_created = _create_github_repo(
            client, spec.target_org, name, spec.visibility,
            custom_properties=_creation_props,
        )
        if not repo_created:
            return _MirrorResult(ok=False, error="Failed to create GitHub repository")

        # Fast-path: if the GitHub repo already existed, check whether migration is
        # already complete using git ls-remote (metadata only -- no object transfer).
        # This avoids a full clone+push for repos that are already fully migrated,
        # even when the checkpoint file has been deleted.
        if not newly_created:
            log.info(
                f"[{name}] GitHub repo already exists -- checking via ls-remote "
                "whether migration is already complete..."
            )
            _ls_db, _ls_br_shas, _ls_tag_names = _ls_remote_refs(
                clone_url, safe_clone, name,
                timeout=min(config.clone_timeout, 120),
            )
            if _ls_br_shas or _ls_tag_names:
                _gh_br_shas = _get_github_branch_shas(client, spec.target_org, name)
                _gh_tag_names = _get_github_ref_names(
                    client, f"/repos/{spec.target_org}/{name}/tags"
                )
                _missing_br = sorted(set(_ls_br_shas) - set(_gh_br_shas))
                _missing_tg = sorted(_ls_tag_names - _gh_tag_names)
                # A branch where GitHub is strictly *ahead* of GitLab means GitHub
                # has extra commits (e.g. .github CI skeleton) that are not on GitLab.
                # Treating those as a mismatch would trigger a force-push that wipes
                # those extra commits.  Use the compare API to distinguish "ahead"
                # (safe to skip) from "behind/diverged" (needs re-push).
                _sha_mismatch = []
                _gh_ahead_branches = []
                for _b in _ls_br_shas:
                    if _b not in _gh_br_shas:
                        continue  # counted in _missing_br already
                    if _ls_br_shas[_b] == _gh_br_shas[_b]:
                        continue  # SHAs match -- fine
                    if _github_is_ahead(client, spec.target_org, name, _ls_br_shas[_b], _gh_br_shas[_b]):
                        _gh_ahead_branches.append(_b)
                    else:
                        _sha_mismatch.append(_b)
                if _gh_ahead_branches:
                    log.info(
                        f"[{name}] GitHub is ahead of GitLab on "
                        f"{len(_gh_ahead_branches)} branch(es) "
                        f"({', '.join(_gh_ahead_branches[:5])}{'...' if len(_gh_ahead_branches) > 5 else ''}) "
                        "-- likely .github CI skeleton commits; excluded from mismatch check"
                    )
                if not _missing_br and not _missing_tg and not _sha_mismatch:
                    _db = _ls_db or next(iter(_ls_br_shas), "")
                    _db_sha = _ls_br_shas.get(_db, "")
                    # Fetch commit count from GitHub for the default branch so the
                    # progress line can show [Nb Nt Nc] even on the fast-exit path.
                    _commit_count = _get_github_commit_count(
                        client, spec.target_org, name, _db_sha
                    ) if _db_sha else 0
                    _commit_info = f", {_commit_count:,} commit(s) on '{_db}'" if _commit_count else ""
                    # Determine whether GitHub HEAD is ahead of GitLab HEAD
                    # (i.e. GitHub has .github CI skeleton commits on top).
                    # If so, record gh_head_is_ahead=True so the report can
                    # show a qualified match instead of a false mismatch.
                    _current_gh_sha = _gh_br_shas.get(_db, "")
                    _gh_ahead = _db_sha != _current_gh_sha and _github_is_ahead(
                        client, spec.target_org, name, _db_sha, _current_gh_sha
                    )
                    if _gh_ahead:
                        log.info(
                            f"[{name}] Already fully migrated: {len(_ls_br_shas)} branch(es), "
                            f"{len(_ls_tag_names)} tag(s), all GitLab SHAs present"
                            f"{_commit_info} -- GitHub HEAD is ahead (extra .github commit(s)); skipping clone+push"
                        )
                    else:
                        log.info(
                            f"[{name}] Already fully migrated: {len(_ls_br_shas)} branch(es), "
                            f"{len(_ls_tag_names)} tag(s), all SHAs match"
                            f"{_commit_info} -- skipping clone+push"
                        )
                    return _MirrorResult(
                        ok=True,
                        default_branch=_db,
                        branch_count=len(_ls_br_shas),
                        head_commit_sha=_db_sha,
                        gh_branch_count=len(_gh_br_shas),
                        gh_head_commit_sha=_current_gh_sha,
                        tag_count=len(_ls_tag_names),
                        gh_tag_count=len(_gh_tag_names),
                        commit_count=_commit_count,
                        gh_repo_url=f"{config.github_url.rstrip('/')}/{spec.target_org}/{name}",
                        validation_status="match",
                        validation_notes=(
                            "already migrated -- GitHub HEAD ahead by .github commit(s); GitLab history fully present"
                            if _gh_ahead else
                            "already migrated -- skipped clone+push"
                        ),
                        gh_head_is_ahead=_gh_ahead,
                    )
                log.info(
                    f"[{name}] Incomplete migration detected: "
                    f"{len(_missing_br)} missing branch(es), "
                    f"{len(_missing_tg)} missing tag(s), "
                    f"{len(_sha_mismatch)} SHA mismatch(es) -- proceeding with full clone+push"
                )
            # ls-remote returned empty refs (source empty or network error) -- fall through

        log.info(f"[{name}] Cloning {safe_clone}")
        ok, err = _git_with_retry(
            cmd=["git", "clone", "--mirror", clone_url, str(mirror_dir)],
            log_cmd=["git", "clone", "--mirror", safe_clone, str(mirror_dir)],
            timeout=config.clone_timeout,
            max_retries=config.git_max_retries,
            label=name,
            action="Clone",
            cleanup_dir=mirror_dir,
        )
        if not ok:
            return _MirrorResult(ok=False, error=err)

        # Empty repo: git clone --mirror exits 0 but produces no refs.
        branch_count_raw = _count_git_refs(mirror_dir, "refs/heads/")
        tag_count_raw    = _count_git_refs(mirror_dir, "refs/tags/")
        if branch_count_raw == 0 and tag_count_raw == 0:
            log.info(
                f"[{name}] Source repository is empty (no branches, no tags). "
                "GitHub repo created; nothing to push."
            )
            return _MirrorResult(
                ok=True,
                default_branch="",
                validation_status="match",
                validation_notes="empty repository -- no branches or tags to migrate",
                gh_repo_url=f"{config.github_url.rstrip('/')}/{spec.target_org}/{name}",
            )

        # Pre-push oversized blob scan: GitHub rejects blobs > 100 MB.
        if config.check_oversized_files:
            oversized = _find_oversized_blobs(mirror_dir, name)
            if oversized > 0:
                return _MirrorResult(
                    ok=False,
                    error=(
                        f"{oversized} blob object(s) exceed GitHub's "
                        f"{_GITHUB_FILE_SIZE_LIMIT_MB} MB file size limit and will cause "
                        "the push to be rejected. Rewrite history with "
                        "'git filter-repo --strip-blobs-bigger-than 100M', then retry."
                    ),
                )

        # LFS: --mirror only copies pointer files; fetch + push binary objects separately.
        lfs_detected = False
        lfs_object_count = 0
        if config.lfs_enabled and _LFS_AVAILABLE:
            lfs_detected = _is_lfs_repo(mirror_dir)
            if lfs_detected:
                log.info(f"[{name}] Git LFS objects detected -- fetching from GitLab source...")
                lfs_ok, lfs_err = _git_lfs_fetch_all(mirror_dir, name, config.clone_timeout)
                if not lfs_ok:
                    # Hard failure: the source has LFS objects but we couldn't download
                    # them.  A soft-continue would silently leave dangling LFS pointers
                    # on GitHub -- exactly the data-loss scenario we want to prevent.
                    return _MirrorResult(
                        ok=False,
                        error=f"LFS fetch failed (data would be lost on GitHub): {lfs_err}",
                        lfs_detected=True,
                    )
                lfs_object_count = _count_lfs_objects(mirror_dir)
                log.info(f"[{name}] LFS fetch: {lfs_object_count} object(s) ready to push")

        _filter_branches(mirror_dir, spec.branch_include, spec.branch_exclude, name)

        # Snapshot source state after filtering (used for validation later).
        src_branch_names: set[str] = _get_git_ref_names(mirror_dir, "refs/heads/")
        src_tag_names:    set[str] = _get_git_ref_names(mirror_dir, "refs/tags/")
        src_branch_shas:  dict[str, str] = _get_git_branch_shas(mirror_dir)

        default_branch = _get_default_branch(mirror_dir)
        head_sha       = _get_head_commit_sha(mirror_dir, default_branch)
        commit_count   = _count_git_commits(mirror_dir) if config.detailed_commit_count else 0

        if default_branch:
            log.info(
                f"[{name}] Default branch: '{default_branch}' | "
                f"{len(src_branch_names)} branch(es) | {len(src_tag_names)} tag(s)"
                + (f" | {commit_count} commit(s)" if commit_count else "")
            )
        else:
            log.warning(f"[{name}] Could not determine source default branch")

        # Optional: rename the source default branch to 'main' before pushing.
        renamed = False
        if config.rename_default_branch and default_branch and default_branch != "main":
            rc1, _, e1 = _run_git(
                ["git", "update-ref", "refs/heads/main", f"refs/heads/{default_branch}"],
                cwd=mirror_dir, timeout=15,
            )
            rc2, _, e2 = _run_git(
                ["git", "update-ref", "-d", f"refs/heads/{default_branch}"],
                cwd=mirror_dir, timeout=15,
            )
            if rc1 == 0 and rc2 == 0:
                log.info(f"[{name}] Branch renamed: '{default_branch}' -> 'main' (pre-push)")
                default_branch = "main"
                renamed = True
                # Refresh snapshots after rename
                src_branch_names = _get_git_ref_names(mirror_dir, "refs/heads/")
                src_branch_shas  = _get_git_branch_shas(mirror_dir)
            else:
                log.warning(f"[{name}] Could not rename branch to 'main': {e1 or e2} -- pushing as-is")

        gh_base_no_scheme = github_base.split("://", 1)[-1]

        # Batched push: single --mirror on large repos triggers GitHub HTTP 500;
        # explicit refspecs in batches keep each pack within receive-pack limits.
        ref_code, ref_out, ref_err = _run_git(
            ["git", "for-each-ref", "--format=%(refname)"],
            cwd=mirror_dir, timeout=60,
        )
        if ref_code != 0:
            return _MirrorResult(
                ok=False,
                error=f"Failed to enumerate refs before push: {ref_err}",
                lfs_detected=lfs_detected,
                lfs_object_count=lfs_object_count,
            )
        all_push_refs: list[str] = [r.strip() for r in ref_out.splitlines() if r.strip()]
        if not all_push_refs:
            log.warning(f"[{name}] No refs remain after branch filtering -- skipping push")
        else:
            batch_size = config.git_push_batch_size
            if batch_size > 0 and len(all_push_refs) > batch_size:
                ref_batches = [
                    all_push_refs[i : i + batch_size]
                    for i in range(0, len(all_push_refs), batch_size)
                ]
                log.info(
                    f"[{name}] Splitting {len(all_push_refs)} refs into "
                    f"{len(ref_batches)} push batches ({batch_size} refs each) "
                    f"to avoid GitHub 500 on large packs"
                )
            else:
                ref_batches = [all_push_refs]

            for batch_idx, ref_batch in enumerate(ref_batches, 1):
                n_batches = len(ref_batches)
                batch_label = f" (batch {batch_idx}/{n_batches})" if n_batches > 1 else ""
                refspecs = [f"+{r}:{r}" for r in ref_batch]

                # Log a preview of which refs are in this batch so operators
                # can see exactly what is being pushed.
                _preview = ref_batch[:3] + ([f"... +{len(ref_batch) - 3} more"] if len(ref_batch) > 3 else [])
                log.info(
                    f"[{name}] Push{batch_label}: {len(ref_batch)} refs -- "
                    + ", ".join(_preview)
                )

                def _push_cmd_factory(  # type: ignore[misc]
                    _refspecs: list[str] = refspecs,
                ) -> tuple[list[str], list[str]]:
                    fresh_token = config.auth.get_token_for_git()
                    push_url = f"https://{fresh_token}@{gh_base_no_scheme}/{spec.target_org}/{name}.git"
                    return (
                        ["git", "push", "--no-thin", push_url] + _refspecs,
                        ["git", "push", "--no-thin", safe_push] + _refspecs,
                    )

                _batch_start = time.monotonic()
                ok, err = _git_with_retry(
                    cmd=[], log_cmd=[],
                    cmd_factory=_push_cmd_factory,
                    timeout=config.push_timeout,
                    max_retries=config.git_max_retries,
                    label=name,
                    action=f"Push{batch_label}",
                    cwd=mirror_dir,
                    http_post_buffer=config.git_http_post_buffer,
                    stream_stderr=True,
                )
                if not ok:
                    return _MirrorResult(
                        ok=False, error=err, lfs_detected=lfs_detected,
                        lfs_object_count=lfs_object_count,
                    )
                log.info(
                    f"[{name}] Push{batch_label} done in "
                    f"{_fmt_duration(time.monotonic() - _batch_start)}"
                )

        # LFS push (after git objects are on GitHub).
        if lfs_detected and lfs_object_count > 0:
            log.info(f"[{name}] Pushing {lfs_object_count} LFS object(s) to GitHub...")
            fresh_token = config.auth.get_token_for_git()
            lfs_push_url = f"https://{fresh_token}@{gh_base_no_scheme}/{spec.target_org}/{name}.git"
            lfs_ok, lfs_err = _git_lfs_push_all(
                mirror_dir, lfs_push_url, safe_push, name,
                config.push_timeout, config.git_http_post_buffer,
            )
            if not lfs_ok:
                return _MirrorResult(
                    ok=False,
                    error=f"LFS push failed: {lfs_err}",
                    lfs_detected=lfs_detected,
                    lfs_object_count=lfs_object_count,
                )

        if default_branch:
            _set_github_default_branch(client, spec.target_org, name, default_branch)

        gh_repo_url = f"{config.github_url.rstrip('/')}/{spec.target_org}/{name}"

        # Post-push validation: compare branch/tag name-sets and SHAs.
        log.info(f"[{name}] Validating migration against GitHub...")

        gh_branch_shas: dict[str, str] = _get_github_branch_shas(client, spec.target_org, name)
        gh_branch_names: set[str] = set(gh_branch_shas.keys())
        gh_tag_names:    set[str] = _get_github_ref_names(
            client, f"/repos/{spec.target_org}/{name}/tags"
        )
        gh_head_sha: str = gh_branch_shas.get(default_branch, "")

        missing_branches: list[str] = sorted(src_branch_names - gh_branch_names)
        missing_tags:     list[str] = sorted(src_tag_names - gh_tag_names)

        branch_sha_mismatches: list[str] = [
            f"{b}: src={src_branch_shas[b][:8]}…{src_branch_shas[b][-4:]} "
            f"!= gh={gh_branch_shas[b][:8]}…{gh_branch_shas[b][-4:]}"
            for b in sorted(src_branch_names & gh_branch_names)
            if src_branch_shas.get(b) and gh_branch_shas.get(b)
            and src_branch_shas[b] != gh_branch_shas[b]
        ]

        v_status, v_notes = _compute_validation(
            branch_count=len(src_branch_names),
            head_sha=head_sha,
            gh_branch_count=len(gh_branch_names),
            gh_head_sha=gh_head_sha,
            tag_count=len(src_tag_names),
            gh_tag_count=len(gh_tag_names),
            commit_count=commit_count,
            gh_commit_count=0,
            missing_branches=missing_branches,
            missing_tags=missing_tags,
            branch_sha_mismatches=branch_sha_mismatches,
        )

        if v_status == "mismatch":
            log.warning(f"[{name}] Initial push incomplete -- attempting targeted remediation:")
            if missing_branches:
                sample = ", ".join(missing_branches[:10])
                extra  = f" (+{len(missing_branches)-10} more)" if len(missing_branches) > 10 else ""
                log.warning(f"[{name}]   Missing branches ({len(missing_branches)}): {sample}{extra}")
            if missing_tags:
                sample = ", ".join(missing_tags[:10])
                extra  = f" (+{len(missing_tags)-10} more)" if len(missing_tags) > 10 else ""
                log.warning(f"[{name}]   Missing tags ({len(missing_tags)}): {sample}{extra}")
            if branch_sha_mismatches:
                sample = "; ".join(branch_sha_mismatches[:5])
                extra  = f" (+{len(branch_sha_mismatches)-5} more)" if len(branch_sha_mismatches) > 5 else ""
                log.warning(f"[{name}]   Branch SHA mismatches ({len(branch_sha_mismatches)}): {sample}{extra}")

            # Targeted remediation: re-push only missing/diverged refs (non-destructive).
            mismatch_branches = [m.split(":")[0] for m in branch_sha_mismatches]
            rem_refspecs: list[str] = (
                [f"+refs/heads/{b}:refs/heads/{b}" for b in missing_branches]
                + [f"+refs/heads/{b}:refs/heads/{b}" for b in mismatch_branches]
                + [f"+refs/tags/{t}:refs/tags/{t}" for t in missing_tags]
            )

            if rem_refspecs:
                _rem_batch_size = config.git_push_batch_size
                _rem_batches = (
                    [
                        rem_refspecs[i: i + _rem_batch_size]
                        for i in range(0, len(rem_refspecs), _rem_batch_size)
                    ]
                    if _rem_batch_size > 0 and len(rem_refspecs) > _rem_batch_size
                    else [rem_refspecs]
                )
                log.info(
                    f"[{name}] Remediation: pushing {len(rem_refspecs)} missing/diverged "
                    f"ref(s) in {len(_rem_batches)} batch(es)..."
                )

                rem_ok, rem_err = True, ""
                for _rbatch_idx, _rbatch in enumerate(_rem_batches, 1):
                    _rn = len(_rem_batches)
                    _rlabel = f" (batch {_rbatch_idx}/{_rn})" if _rn > 1 else ""

                    def _remediation_factory(  # type: ignore[misc]
                        _rs: list[str] = _rbatch,
                    ) -> tuple[list[str], list[str]]:
                        fresh_token = config.auth.get_token_for_git()
                        push_url = f"https://{fresh_token}@{gh_base_no_scheme}/{spec.target_org}/{name}.git"
                        return (
                            ["git", "push", push_url] + _rs,
                            ["git", "push", safe_push] + _rs,
                        )

                    rem_ok, rem_err = _git_with_retry(
                        cmd=[], log_cmd=[],
                        cmd_factory=_remediation_factory,
                        timeout=config.push_timeout,
                        max_retries=config.git_max_retries,
                        label=name,
                        action=f"Remediation push{_rlabel}",
                        cwd=mirror_dir,
                        http_post_buffer=config.git_http_post_buffer,
                    )
                    if not rem_ok:
                        break

                if rem_ok:
                    log.info(f"[{name}] Remediation push succeeded -- re-validating...")

                    # Re-fetch GitHub state after remediation.
                    gh_branch_shas = _get_github_branch_shas(client, spec.target_org, name)
                    gh_branch_names = set(gh_branch_shas.keys())
                    gh_tag_names = _get_github_ref_names(
                        client, f"/repos/{spec.target_org}/{name}/tags"
                    )
                    gh_head_sha = gh_branch_shas.get(default_branch, "")

                    missing_branches = sorted(src_branch_names - gh_branch_names)
                    missing_tags     = sorted(src_tag_names - gh_tag_names)
                    branch_sha_mismatches = [
                        f"{b}: src={src_branch_shas[b][:8]}…{src_branch_shas[b][-4:]} "
                        f"!= gh={gh_branch_shas[b][:8]}…{gh_branch_shas[b][-4:]}"
                        for b in sorted(src_branch_names & gh_branch_names)
                        if src_branch_shas.get(b) and gh_branch_shas.get(b)
                        and src_branch_shas[b] != gh_branch_shas[b]
                    ]

                    v_status, v_notes = _compute_validation(
                        branch_count=len(src_branch_names),
                        head_sha=head_sha,
                        gh_branch_count=len(gh_branch_names),
                        gh_head_sha=gh_head_sha,
                        tag_count=len(src_tag_names),
                        gh_tag_count=len(gh_tag_names),
                        commit_count=commit_count,
                        gh_commit_count=0,
                        missing_branches=missing_branches,
                        missing_tags=missing_tags,
                        branch_sha_mismatches=branch_sha_mismatches,
                    )

                    if v_status == "match":
                        log.info(
                            f"[{name}] Validation PASSED after remediation -- "
                            f"{len(src_branch_names)} branch(es), {len(src_tag_names)} tag(s), "
                            "all HEAD SHAs match"
                            + (f" | LFS: {lfs_object_count} object(s) migrated" if lfs_detected else "")
                        )
                        return _MirrorResult(
                            ok=True,
                            default_branch=default_branch,
                            branch_count=len(src_branch_names),
                            head_commit_sha=head_sha,
                            gh_branch_count=len(gh_branch_names),
                            gh_head_commit_sha=gh_head_sha,
                            tag_count=len(src_tag_names),
                            gh_tag_count=len(gh_tag_names),
                            commit_count=commit_count,
                            gh_commit_count=0,
                            gh_repo_url=gh_repo_url,
                            default_branch_renamed=renamed,
                            validation_status="match",
                            validation_notes="passed after remediation push",
                            missing_branches=[],
                            missing_tags=[],
                            branch_sha_mismatches=[],
                            lfs_detected=lfs_detected,
                            lfs_object_count=lfs_object_count,
                        )

                    log.error(
                        f"[{name}] Still incomplete after remediation: {v_notes}"
                    )
                else:
                    log.error(f"[{name}] Remediation push failed: {rem_err}")

            # Remediation could not resolve all gaps -- fail the repo with full detail
            # so the operator knows exactly what to investigate.
            log.error(f"[{name}] Validation FAILED -- repo marked failed for re-run:")
            if missing_branches:
                sample = ", ".join(missing_branches[:10])
                extra  = f" (+{len(missing_branches)-10} more)" if len(missing_branches) > 10 else ""
                log.error(f"[{name}]   Missing branches ({len(missing_branches)}): {sample}{extra}")
            if missing_tags:
                sample = ", ".join(missing_tags[:10])
                extra  = f" (+{len(missing_tags)-10} more)" if len(missing_tags) > 10 else ""
                log.error(f"[{name}]   Missing tags ({len(missing_tags)}): {sample}{extra}")
            if branch_sha_mismatches:
                sample = "; ".join(branch_sha_mismatches[:5])
                extra  = f" (+{len(branch_sha_mismatches)-5} more)" if len(branch_sha_mismatches) > 5 else ""
                log.error(f"[{name}]   Branch SHA mismatches ({len(branch_sha_mismatches)}): {sample}{extra}")
            return _MirrorResult(
                ok=False,
                error=f"Validation mismatch (persisted after remediation): {v_notes}",
                default_branch=default_branch,
                branch_count=len(src_branch_names),
                head_commit_sha=head_sha,
                gh_branch_count=len(gh_branch_names),
                gh_head_commit_sha=gh_head_sha,
                tag_count=len(src_tag_names),
                gh_tag_count=len(gh_tag_names),
                commit_count=commit_count,
                gh_commit_count=0,
                gh_repo_url=gh_repo_url,
                default_branch_renamed=renamed,
                validation_status=v_status,
                validation_notes=v_notes,
                missing_branches=missing_branches,
                missing_tags=missing_tags,
                branch_sha_mismatches=branch_sha_mismatches,
                lfs_detected=lfs_detected,
                lfs_object_count=lfs_object_count,
            )

        if v_status == "unknown":
            log.warning(
                f"[{name}] Validation INCONCLUSIVE -- GitHub API returned no data. "
                "Repo is marked failed; re-run to retry."
            )
            return _MirrorResult(
                ok=False,
                error=f"Validation inconclusive (GitHub API returned no data): {v_notes}",
                default_branch=default_branch,
                branch_count=len(src_branch_names),
                head_commit_sha=head_sha,
                gh_repo_url=gh_repo_url,
                default_branch_renamed=renamed,
                validation_status=v_status,
                validation_notes=v_notes,
                lfs_detected=lfs_detected,
                lfs_object_count=lfs_object_count,
            )

        log.info(
            f"[{name}] Validation PASSED -- "
            f"{len(src_branch_names)} branch(es), {len(src_tag_names)} tag(s), "
            "all HEAD SHAs match"
            + (f" | LFS: {lfs_object_count} object(s) migrated" if lfs_detected else "")
        )
        return _MirrorResult(
            ok=True,
            default_branch=default_branch,
            branch_count=len(src_branch_names),
            head_commit_sha=head_sha,
            gh_branch_count=len(gh_branch_names),
            gh_head_commit_sha=gh_head_sha,
            tag_count=len(src_tag_names),
            gh_tag_count=len(gh_tag_names),
            commit_count=commit_count,
            gh_commit_count=0,
            gh_repo_url=gh_repo_url,
            default_branch_renamed=renamed,
            validation_status=v_status,
            validation_notes=v_notes,
            missing_branches=[],
            missing_tags=[],
            branch_sha_mismatches=[],
            lfs_detected=lfs_detected,
            lfs_object_count=lfs_object_count,
        )

    except Exception as exc:
        return _MirrorResult(ok=False, error=str(exc))

    finally:
        shutil.rmtree(str(tmp_dir), ignore_errors=True)


# ===========================================================================
# repos.csv loader
# ===========================================================================

def load_repos_csv(
    csv_path: Path,
    default_org: str = "",
    default_branch_include: list[str] | None = None,
    default_branch_exclude: list[str] | None = None,
) -> list[RepoSpec]:

    _global_inc = default_branch_include or []
    _global_exc = default_branch_exclude or []

    def _parse_patterns(raw_cell: str, defaults: list[str]) -> list[str]:
        """Split a semicolon-separated cell into a list of regex patterns."""
        raw_cell = (raw_cell or "").strip()
        return [p.strip() for p in raw_cell.split(";") if p.strip()] if raw_cell else defaults

    required_cols = {"namespace", "project", "target_name"}
    repos: list[RepoSpec] = []
    skipped = 0
    seen_sources: set[str] = set()   # "namespace/project" keys already queued
    seen_targets: set[str] = set()   # "target_org/target_name" keys already queued

    try:
        fh = csv_path.open(encoding="utf-8-sig", newline="")
    except OSError as exc:
        raise FileNotFoundError(f"Cannot open {csv_path.name}: {exc}") from exc

    with fh:
        reader = csv.DictReader(fh)
        if not reader.fieldnames:
            raise _MissingConfigError(f"{csv_path.name} is empty or has no header row.")

        actual = {c.strip().lower() for c in reader.fieldnames}
        missing = required_cols - actual
        if missing:
            raise ValueError(
                f"{csv_path.name} is missing required columns: {missing}. "
                "First row must be: namespace,project,target_name[,visibility]"
            )

        for i, row in enumerate(reader, start=2):
            ns = (row.get("namespace") or "").strip()
            proj = (row.get("project") or "").strip()
            tgt_org = (row.get("target_org") or "").strip() or default_org
            tgt = (row.get("target_name") or "").strip()
            vis = (row.get("visibility") or "private").strip().lower()

            if not ns or not proj or not tgt:
                log.warning(f"repos.csv row {i}: blank required field (namespace/project/target_name) -- skipped")
                skipped += 1
                continue

            if not tgt_org:
                log.warning(
                    f"repos.csv row {i}: [{ns}/{proj}] has no target_org in CSV "
                    f"and no github.default_org set in mirror-config.json -- skipped"
                )
                skipped += 1
                continue

            if vis not in _VALID_VISIBILITIES:
                log.warning(
                    f"repos.csv row {i}: invalid visibility '{vis}' -- defaulting to 'private'"
                )
                vis = "private"

            inc_raw = (row.get("branch_include") or "").strip()
            exc_raw = (row.get("branch_exclude") or "").strip()
            inc = _parse_patterns(inc_raw, _global_inc)
            exc = _parse_patterns(exc_raw, _global_exc)
            from_global = not inc_raw and not exc_raw
            ci_template = (row.get("ci_template") or "").strip()

            # Duplicate detection: same source/target across rows corrupts checkpoint or GitHub repo.
            source_key = f"{ns}/{proj}"
            target_key = f"{tgt_org}/{tgt}"

            if source_key in seen_sources:
                log.warning(
                    f"repos.csv row {i}: duplicate source '{source_key}' -- "
                    "skipped to prevent double migration"
                )
                skipped += 1
                continue

            if target_key in seen_targets:
                log.warning(
                    f"repos.csv row {i}: target '{target_key}' is already used by another row -- "
                    "two different sources would overwrite each other on GitHub; row skipped"
                )
                skipped += 1
                continue

            seen_sources.add(source_key)
            seen_targets.add(target_key)

            repos.append(RepoSpec(ns, proj, tgt_org, tgt, vis, inc, exc, from_global, ci_template))

    if skipped:
        log.warning(f"Skipped {skipped} invalid row(s) in repos.csv")

    if not repos:
        raise _MissingConfigError(
            f"repos.csv is empty -- no repositories to migrate.\n"
            f"  Open {csv_path} and add at least one row, e.g.:\n"
            f"    namespace,project,target_org,target_name,visibility\n"
            f"    my-group,my-repo,my-github-org,my-repo,private"
        )

    log.info(f"Loaded {len(repos)} repo(s) from {csv_path.name}")
    return repos


# ===========================================================================
# Config loading and validation
# ===========================================================================

def _is_placeholder(value: str) -> bool:
    return not value or any(tok in value for tok in _PLACEHOLDER_TOKENS)


def _check_placeholders(raw: dict) -> None:
    """Scan mirror-config.json for unfilled placeholder values. Reports all issues at once."""
    issues: list[str] = []
    auth = raw.get("auth", {})
    mode = str(auth.get("mode", "pat")).lower()

    if mode == "pat":
        if _is_placeholder(auth.get("pat", "")):
            issues.append(
                "auth.pat -- replace with your GitHub Personal Access Token "
                "(GitHub -> Settings -> Developer settings -> Personal access tokens)"
            )
    elif mode == "app":
        app = auth.get("app", {})
        app_id = app.get("app_id")
        if not isinstance(app_id, int) or app_id <= 0:
            issues.append("auth.app.app_id -- replace with your numeric GitHub App ID")
        key_path_val = app.get("private_key_path", "")
        if not key_path_val or _is_placeholder(key_path_val):
            issues.append(
                "auth.app.private_key_path -- path to your GitHub App .pem private key file"
            )
        inst_id = app.get("installation_id")
        if not isinstance(inst_id, int) or inst_id <= 0:
            issues.append(
                "auth.app.installation_id -- replace with your numeric GitHub App Installation ID"
            )
    else:
        issues.append(f"auth.mode -- invalid value '{mode}': must be 'pat' or 'app'")

    gl = raw.get("gitlab", {})
    if _is_placeholder(gl.get("url", "")):
        issues.append(
            "gitlab.url -- replace with your GitLab instance URL "
            "(e.g. https://gitlab.com or https://gitlab.your-company.com)"
        )
    if _is_placeholder(gl.get("pat", "")):
        issues.append(
            "gitlab.pat -- replace with your GitLab Personal Access Token "
            "(GitLab -> User Settings -> Access Tokens, scope: read_repository)"
        )

    if issues:
        numbered = "\n".join(f"  {i + 1}. {m}" for i, m in enumerate(issues))
        raise _MissingConfigError(
            f"mirror-config.json has {len(issues)} unfilled value(s):\n"
            f"{numbered}\n\n"
            "Fill in all values and run again."
        )


def _load_repo_properties_for_creation(script_dir: Path) -> dict[str, dict[str, str]]:
    """Load repo-properties.csv for injection at repo-creation time.

    Returns an empty dict (silently) when the file does not exist, so the
    feature is purely opt-in -- no file, no properties injected.
    Logs a warning and returns an empty dict on any parse error so a bad
    CSV never blocks the entire migration.
    """
    csv_path = script_dir / _REPO_PROPERTIES_CSV_FILENAME
    if not csv_path.exists():
        return {}
    try:
        props = _load_repo_properties_csv(csv_path)
        if props:
            log.info(
                f"Repo custom properties loaded from {csv_path.name} "
                f"({len(props)} repo(s)) -- will be injected at repo-creation time"
            )
        return props
    except Exception as exc:
        log.warning(
            f"Could not load {csv_path.name} for repo-creation properties: {exc}. "
            "Continuing without custom properties at creation time."
        )
        return {}


def load_config(config_path: Path, repos_csv_path: Path) -> MigrationConfig:
    """Parse mirror-config.json + repos.csv. Returns a validated MigrationConfig."""
    if not config_path.exists():
        raise FileNotFoundError(
            f"Config file not found: {config_path}\n"
            "Create mirror-config.json next to this script."
        )

    try:
        raw = json.loads(config_path.read_text(encoding="utf-8-sig"))
    except json.JSONDecodeError as exc:
        raise ValueError(
            f"mirror-config.json: invalid JSON at line {exc.lineno}: {exc.msg}"
        ) from exc

    for section in ("auth", "github", "gitlab"):
        if section not in raw:
            raise ValueError(f"mirror-config.json is missing required section: '{section}'")

    _check_placeholders(raw)

    gh = raw["github"]
    gl = raw["gitlab"]
    auth_raw = raw["auth"]
    mig = raw.get("migration", {})

    mode = str(auth_raw.get("mode", "pat")).lower()
    api_base = gh.get("api_url", _GITHUB_API_URL)

    if mode == "pat":
        auth: _Auth = PATAuth(auth_raw["pat"])
    else:
        app_conf = auth_raw["app"]
        key_path = Path(app_conf["private_key_path"])
        if not key_path.is_absolute():
            key_path = (_SCRIPT_DIR / key_path).resolve()
        if not key_path.exists():
            raise FileNotFoundError(
                f"GitHub App private key not found: {key_path}\n"
                "Ensure 'auth.app.private_key_path' points to the .pem file."
            )
        private_key = key_path.read_text(encoding="utf-8")
        auth = GitHubAppAuth(
            app_id=int(app_conf["app_id"]),
            private_key_pem=private_key,
            installation_id=int(app_conf["installation_id"]),
            api_base=api_base,
        )

    log.info(f"Auth mode : {auth.mode_label}")

    default_org = gh.get("default_org", "").strip()

    branches_conf  = raw.get("branches", {})
    branch_include = [str(p) for p in branches_conf.get("include", [])]
    branch_exclude = [str(p) for p in branches_conf.get("exclude", [])]

    repos = load_repos_csv(repos_csv_path, default_org, branch_include, branch_exclude)

    # Clamp configurable integers to safe operating bounds.
    # max_workers=0  → ThreadPoolExecutor raises ValueError immediately.
    # max_workers>20 → likely to exhaust file descriptors / overwhelm the source.
    # git_max_retries<1 → the for-loop body never executes (no attempt is made).
    raw_workers = int(mig.get("max_workers", _DEFAULT_MAX_WORKERS))
    if not (1 <= raw_workers <= 20):
        raise ValueError(
            f"mirror-config.json: migration.max_workers must be between 1 and 20 "
            f"(got {raw_workers})"
        )

    raw_retries = int(mig.get("git_retry_count", _DEFAULT_GIT_RETRIES))
    if raw_retries < 1:
        raise ValueError(
            f"mirror-config.json: migration.git_retry_count must be at least 1 "
            f"(got {raw_retries})"
        )

    dry_run = bool(mig.get("dry_run", False))
    if dry_run:
        log.info("DRY RUN mode enabled -- no repos will be created or pushed")

    rename_default_branch = bool(mig.get("rename_default_branch_to_main", False))
    if rename_default_branch:
        log.info("rename_default_branch_to_main enabled -- source default branch will be renamed to 'main'")

    min_free_disk_gb = float(mig.get("min_free_disk_gb", _DEFAULT_MIN_FREE_DISK_GB))
    lfs_enabled = bool(mig.get("lfs_enabled", True))
    detailed_commit_count = bool(mig.get("detailed_commit_count", False))
    check_oversized_files = bool(mig.get("check_oversized_files", True))

    # CI skeleton configuration.
    ci_skeleton: CiSkeletonConfig | None = None
    ci_skel_raw = raw.get("ci_skeleton", {})
    if ci_skel_raw.get("enabled", False):
        ci_templates_dir_raw = ci_skel_raw.get("templates_dir", "ci-templates")
        ci_templates_dir = Path(ci_templates_dir_raw)
        if not ci_templates_dir.is_absolute():
            ci_templates_dir = (_SCRIPT_DIR / ci_templates_dir).resolve()

        ci_skeleton = CiSkeletonConfig(
            enabled=True,
            templates_dir=ci_templates_dir,
            target_path=ci_skel_raw.get("target_path", ".github/workflows"),
            commit_message=ci_skel_raw.get("commit_message", "chore: add CI skeleton workflow"),
            branches=[str(b) for b in ci_skel_raw.get("branches", ["main", "master", "develop"])],
            skip_if_exists=bool(ci_skel_raw.get("skip_if_exists", True)),
        )

        ci_problems = _validate_ci_skeleton_config(ci_skeleton, repos)
        if ci_problems:
            for problem in ci_problems:
                log.warning(f"CI skeleton config issue: {problem}")
            # Disable if the templates directory or any file is missing -- operator must fix first.
            log.warning(
                "CI skeleton DISABLED due to the configuration issue(s) above. "
                "Fix them and re-run."
            )
            ci_skeleton = None
        else:
            log.info(
                f"CI skeleton enabled  |  "
                f"templates_dir={ci_skeleton.templates_dir.name}  |  "
                f"target_path={ci_skeleton.target_path}  |  "
                f"branches={ci_skeleton.branches}  |  "
                f"skip_if_exists={ci_skeleton.skip_if_exists}"
            )

    if lfs_enabled and not _LFS_AVAILABLE:
        log.warning(
            "migration.lfs_enabled is true but git-lfs is not installed or not on PATH. "
            "LFS objects will NOT be migrated.  Install git-lfs and re-run, or set "
            "lfs_enabled: false to suppress this warning."
        )

    return MigrationConfig(
        auth=auth,
        github_default_org=default_org,
        github_api_url=api_base,
        github_url=gh.get("url", _GITHUB_URL),
        gitlab_url=gl["url"],
        gitlab_pat=gl["pat"],
        max_workers=raw_workers,
        clone_timeout=int(mig.get("clone_timeout_seconds", _DEFAULT_CLONE_TIMEOUT)),
        push_timeout=int(mig.get("push_timeout_seconds", _DEFAULT_PUSH_TIMEOUT)),
        git_max_retries=raw_retries,
        dry_run=dry_run,
        rename_default_branch=rename_default_branch,
        branch_include=branch_include,
        branch_exclude=branch_exclude,
        min_free_disk_gb=min_free_disk_gb,
        git_http_post_buffer=int(mig.get("git_http_post_buffer_bytes", _DEFAULT_GIT_HTTP_POST_BUFFER)),
        git_push_batch_size=int(mig.get("git_push_batch_size", _DEFAULT_PUSH_BATCH_SIZE)),
        lfs_enabled=lfs_enabled,
        detailed_commit_count=detailed_commit_count,
        check_oversized_files=check_oversized_files,
        ci_skeleton=ci_skeleton,
        repo_custom_properties=_load_repo_properties_for_creation(config_path.parent),
        repos=repos,
        notification=_parse_email_config(raw.get("notification", {})),
    )


def _parse_email_config(notif: dict) -> "EmailConfig | None":
    """Build an EmailConfig from the 'notification' block of mirror-config.json.

    Returns None when the block is absent or 'enabled' is false so callers can
    treat a missing config and a disabled config identically.
    """
    if not notif or not notif.get("enabled", False):
        return None

    smtp_port = int(notif.get("smtp_port", 587))
    # Auto-detect TLS mode from port when not explicitly set.
    use_ssl  = bool(notif.get("use_ssl",  smtp_port == 465))
    use_tls  = bool(notif.get("use_tls",  smtp_port == 587 and not use_ssl))

    to_raw = notif.get("to", [])
    to_list = [a.strip() for a in (to_raw if isinstance(to_raw, list) else to_raw.split(",")) if a.strip()]
    cc_raw = notif.get("cc", [])
    cc_list = [a.strip() for a in (cc_raw if isinstance(cc_raw, list) else cc_raw.split(",")) if a.strip()]

    if not to_list:
        log.warning("[email] notification.to is empty -- email notifications disabled")
        return None

    cfg = EmailConfig(
        enabled=True,
        smtp_host=notif["smtp_host"],
        smtp_port=smtp_port,
        smtp_user=notif.get("smtp_user", ""),
        smtp_password=notif.get("smtp_password", ""),
        use_tls=use_tls,
        use_ssl=use_ssl,
        from_address=notif.get("from", ""),
        to_addresses=to_list,
        cc_addresses=cc_list,
        subject_prefix=notif.get("subject_prefix", "[GitMigration]"),
        attach_report=bool(notif.get("attach_report", True)),
        attach_log_tail_lines=int(notif.get("attach_log_tail_lines", 200)),
        timeout=int(notif.get("smtp_timeout_seconds", 30)),
    )
    log.info(
        f"[email] Notifications enabled  |  "
        f"smtp={cfg.smtp_host}:{cfg.smtp_port}  |  "
        f"mode={'SSL' if cfg.use_ssl else 'STARTTLS' if cfg.use_tls else 'plain'}  |  "
        f"to={cfg.to_addresses}"
    )
    return cfg


# ===========================================================================
# Signal handling -- graceful shutdown on Ctrl+C / SIGTERM
# ===========================================================================

_shutdown_event = threading.Event()
_disk_lock = threading.Lock()  # Serialises concurrent per-repo free-disk checks.


def _handle_signal(signum: int, frame: object) -> None:
    log.warning(
        f"Signal {signum} received -- finishing active repositories then shutting down. "
        "Re-run the script to resume from checkpoint."
    )
    _shutdown_event.set()


signal.signal(signal.SIGINT, _handle_signal)
if sys.platform != "win32":
    signal.signal(signal.SIGTERM, _handle_signal)


# ===========================================================================
# Pre-flight connectivity and environment checks
# ===========================================================================

def _check_github_connectivity(client: GitHubClient) -> None:
    """Validate GitHub credentials and connectivity via GET /rate_limit (zero quota cost)."""
    status, resp = client.request("GET", "/rate_limit")
    if status == 401:
        raise RuntimeError("GitHub authentication failed: token is invalid or expired")
    if status != 200:
        raise RuntimeError(f"GitHub API unreachable (HTTP {status})")
    core = resp.get("resources", {}).get("core", {})
    log.info(
        f"GitHub connectivity OK  |  rate limit: "
        f"{core.get('remaining', '?')}/{core.get('limit', '?')} requests remaining"
    )


def _check_gitlab_connectivity(gitlab_url: str, gitlab_pat: str) -> None:
    """Validate GitLab credentials and connectivity via GET /api/v4/user."""
    url = f"{gitlab_url.rstrip('/')}/api/v4/user"
    req = urllib.request.Request(url, headers={"PRIVATE-TOKEN": gitlab_pat})
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read())
        log.info(f"GitLab connectivity OK  |  authenticated as: {data.get('username', '?')}")
    except urllib.error.HTTPError as exc:
        if exc.code == 401:
            raise RuntimeError("GitLab authentication failed: PAT is invalid or expired")
        raise RuntimeError(f"GitLab API error (HTTP {exc.code}) at {gitlab_url}")
    except (urllib.error.URLError, OSError) as exc:
        raise RuntimeError(f"Cannot reach GitLab at {gitlab_url}: {exc}")


def _check_disk_space(min_free_gb: float, max_workers: int = 1) -> None:
    """Ensure sufficient free disk space in the system temp directory.

    The effective minimum is scaled by *max_workers* because every concurrent
    worker may be cloning a repository up to min_free_disk_gb in size at the
    same time.  Also checks free inode count on Linux/macOS.
    """
    tmp = Path(tempfile.gettempdir())
    free_gb = shutil.disk_usage(tmp).free / (1024 ** 3)
    effective_min_gb = min_free_gb * max(1, max_workers)
    if free_gb < effective_min_gb:
        log.warning(
            f"Insufficient disk space: {free_gb:.1f} GB free in {tmp}, "
            f"need at least {effective_min_gb:.1f} GB "
            f"({min_free_gb:.1f} GB \u00d7 {max_workers} concurrent worker(s)). "
            "Free up space, lower migration.min_free_disk_gb, or reduce migration.max_workers. "
            "Continuing anyway -- individual repos may fail if disk fills up."
        )
    # Check free inodes on Linux/macOS -- repos with many small files can exhaust
    # inodes even when gigabytes of block space remain.
    if sys.platform != "win32" and hasattr(os, "statvfs"):
        try:
            vfs = os.statvfs(tmp)
            free_inodes = vfs.f_favail
            _MIN_FREE_INODES = 100_000
            if free_inodes < _MIN_FREE_INODES:
                log.warning(
                    f"Low inode count: {free_inodes:,} free in {tmp}, "
                    f"need at least {_MIN_FREE_INODES:,}. "
                    "Repositories with many small files may exhaust inodes "
                    "independently of available disk space. Continuing anyway."
                )
        except OSError:
            pass


def _check_tmpfs() -> None:
    """Warn if the system temp directory is on a tmpfs (RAM-backed) filesystem.

    On many Linux distributions /tmp is a tmpfs mount backed by RAM + swap.
    Cloning large repositories there consumes RAM, not disk, and can trigger
    the OOM killer on migration servers.  This is a warning so operators can
    decide whether to point TMPDIR at a disk-backed path.
    """
    if sys.platform == "win32":
        return
    tmp = str(Path(tempfile.gettempdir()).resolve())
    fstype = ""

    # Method 1: findmnt (util-linux, available on most Linux distros).
    try:
        result = subprocess.run(
            ["findmnt", "--target", tmp, "--output", "FSTYPE", "--noheadings"],
            capture_output=True, text=True, timeout=5,
        )
        if result.returncode == 0:
            fstype = result.stdout.strip().lower()
    except (FileNotFoundError, subprocess.TimeoutExpired, OSError):
        pass

    # Method 2: /proc/mounts fallback (Linux-only).
    if not fstype:
        try:
            mounts = Path("/proc/mounts").read_text(encoding="utf-8", errors="replace")
            best_prefix, best_fstype = "", ""
            for line in mounts.splitlines():
                parts = line.split()
                if len(parts) >= 3:
                    mount_point = parts[1]
                    if tmp.startswith(mount_point) and len(mount_point) > len(best_prefix):
                        best_prefix = mount_point
                        best_fstype = parts[2]
            fstype = best_fstype.lower()
        except OSError:
            pass

    if "tmpfs" in fstype:
        log.warning(
            f"Temp directory ({tmp}) is on a tmpfs (RAM-backed) filesystem. "
            "Cloning large repos here consumes RAM, not disk, and can trigger the OOM killer. "
            "Set TMPDIR to a disk-backed path before running "
            "(e.g. export TMPDIR=/var/migration-tmp && mkdir -p $TMPDIR)."
        )


def _check_file_descriptor_limit(max_workers: int) -> None:
    """Warn if the open-file-descriptor limit is too low for concurrent git workers.

    Each concurrent worker spawns a git subprocess plus helper processes
    (git-remote-https, etc.), each consuming ~20 file descriptors.  The Linux
    default of 1 024 is easily exhausted with 10+ concurrent workers.
    """
    if sys.platform == "win32":
        return
    try:
        import resource
        soft, _ = resource.getrlimit(resource.RLIMIT_NOFILE)
        # 50 FDs per worker + 200 for the Python process and log handles.
        recommended = max_workers * 50 + 200
        if soft < recommended:
            log.warning(
                f"Open-file-descriptor limit (ulimit -n) is {soft}, "
                f"but {recommended} is recommended for {max_workers} concurrent worker(s). "
                "Low limits cause 'Too many open files' errors during migration. "
                f"Run before starting:  ulimit -n {max(65535, recommended)}"
            )
    except (ImportError, ValueError, OSError):
        pass


def _check_gitlab_token_expiry(gitlab_url: str, gitlab_pat: str, warn_days: int = 7) -> None:
    """Warn if the GitLab PAT expires soon or has already expired.

    Uses GET /api/v4/personal_access_tokens/self (GitLab 14.10+).
    Falls back silently when the endpoint is unavailable or the token has no expiry.
    """
    url = f"{gitlab_url.rstrip('/')}/api/v4/personal_access_tokens/self"
    req = urllib.request.Request(url, headers={"PRIVATE-TOKEN": gitlab_pat})
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read())
    except Exception:
        return  # endpoint not supported or network error -- skip expiry check

    expires_at = data.get("expires_at")  # "2026-06-01" or null
    if not expires_at:
        return  # token has no expiry date

    try:
        expiry = datetime.datetime.strptime(expires_at, "%Y-%m-%d").date()
    except ValueError:
        return

    today = datetime.date.today()
    days_left = (expiry - today).days
    if days_left < 0:
        raise RuntimeError(
            f"GitLab PAT expired on {expires_at}. "
            "Generate a new token: GitLab \u2192 User Settings \u2192 Access Tokens."
        )
    if days_left < warn_days:
        log.warning(
            f"GitLab PAT expires in {days_left} day(s) (on {expires_at}). "
            "Renew it before the migration completes to avoid mid-run authentication failures."
        )


def _preflight_checks(config: MigrationConfig, client: GitHubClient) -> None:
    """Run all pre-flight checks before any migration work begins."""
    _check_disk_space(config.min_free_disk_gb, config.max_workers)
    _check_tmpfs()
    _check_file_descriptor_limit(config.max_workers)
    _check_github_connectivity(client)
    if not config.dry_run:
        _check_gitlab_connectivity(config.gitlab_url, config.gitlab_pat)
        _check_gitlab_token_expiry(config.gitlab_url, config.gitlab_pat)


# ===========================================================================
# Timing metrics
# ===========================================================================

def _compute_metrics(results: list[_RepoResult]) -> dict:
    """Compute basic timing stats for all processed (non-skipped) repos."""
    durations = sorted(
        r.duration_seconds for r in results
        if r.status not in ("skipped", "dry-run") and r.duration_seconds > 0
    )
    if not durations:
        return {}
    n = len(durations)
    return {
        "count": n,
        "min_seconds": round(durations[0], 1),
        "max_seconds": round(durations[-1], 1),
        "mean_seconds": round(sum(durations) / n, 1),
    }


# ===========================================================================
# Email notification
# ===========================================================================

@dataclass
class EmailConfig:
    """SMTP notification settings loaded from the 'notification' block in mirror-config.json."""
    enabled: bool
    smtp_host: str
    smtp_port: int                    # 465=SSL, 587=STARTTLS, 25=plain
    smtp_user: str
    smtp_password: str = field(repr=False)
    use_tls: bool = True              # STARTTLS on port 587 (recommended)
    use_ssl: bool = False             # Implicit SSL on port 465
    from_address: str = ""
    to_addresses: list[str] = field(default_factory=list)
    cc_addresses: list[str] = field(default_factory=list)
    subject_prefix: str = "[GitMigration]"
    attach_report: bool = True        # attach the XLSX/CSV report
    attach_log_tail_lines: int = 200  # last N lines of the log attached as .txt
    timeout: int = 30                 # SMTP connection timeout seconds


def _send_migration_email(
    cfg: EmailConfig,
    results: list[_RepoResult],
    run_started: str,
    elapsed_seconds: float,
    executed_by: str,
    report_path: Path | None,
    log_path: Path,
) -> None:
    """Compose and send a migration-complete notification via SMTP.

    Uses stdlib only (smtplib + email.mime).  Sends a multipart/alternative
    HTML + plain-text message with the XLSX/CSV report and a log tail attached.
    Any exception is caught and logged as a WARNING so a mail failure never
    aborts or marks the migration as failed.
    """
    import smtplib
    import email.mime.multipart
    import email.mime.text
    import email.mime.base
    import email.mime.application
    import email.encoders
    import email.utils

    if not cfg.enabled or not cfg.to_addresses:
        return

    succeeded = [r for r in results if r.status == "succeeded"]
    partial   = [r for r in results if r.status == "partial"]
    failed    = [r for r in results if r.status == "failed"]
    skipped   = [r for r in results if r.status == "skipped"]
    processed = len(succeeded) + len(partial) + len(failed)
    success_rate = (
        f"{len(succeeded)/processed*100:.1f}%" if processed else "N/A"
    )

    # Overall run outcome for subject line
    if failed:
        outcome = "FAILED"
    elif partial:
        outcome = "PARTIAL"
    else:
        outcome = "SUCCEEDED"

    subject = (
        f"{cfg.subject_prefix} Migration {outcome} "
        f"-- {len(succeeded)}/{processed} repos "
        f"| {run_started}"
    )

    # ---- failed / partial repo tables for HTML body ----
    def _repo_rows(repo_list: list[_RepoResult], color: str) -> str:
        if not repo_list:
            return ""
        rows = "".join(
            f"<tr><td style='padding:4px 8px;border:1px solid #ddd'>{r.target}</td>"
            f"<td style='padding:4px 8px;border:1px solid #ddd;color:{color}'>"
            f"{r.status}</td>"
            f"<td style='padding:4px 8px;border:1px solid #ddd;font-size:12px'>"
            f"{(r.error or r.custom_properties_error or r.ci_skeleton_error)[:120]}</td></tr>"
            for r in repo_list
        )
        return rows

    failed_rows  = _repo_rows(failed,  "#c0392b")
    partial_rows = _repo_rows(partial, "#e67e22")
    detail_table = ""
    if failed_rows or partial_rows:
        detail_table = f"""
<h3 style='color:#c0392b'>Repositories Requiring Attention</h3>
<table style='border-collapse:collapse;width:100%;font-family:monospace;font-size:13px'>
  <thead>
    <tr style='background:#f2f2f2'>
      <th style='padding:4px 8px;border:1px solid #ddd;text-align:left'>Repository</th>
      <th style='padding:4px 8px;border:1px solid #ddd;text-align:left'>Status</th>
      <th style='padding:4px 8px;border:1px solid #ddd;text-align:left'>Error (truncated)</th>
    </tr>
  </thead>
  <tbody>{failed_rows}{partial_rows}</tbody>
</table>"""

    html_body = f"""\
<!DOCTYPE html>
<html><head><meta charset='utf-8'></head>
<body style='font-family:Arial,Helvetica,sans-serif;color:#333;max-width:800px'>
<h2 style='border-bottom:2px solid #2c3e50;padding-bottom:6px'>
  GitLab &#8594; GitHub Migration Run Complete
</h2>
<table style='border-collapse:collapse;width:100%;margin-bottom:16px'>
  <tbody>
    <tr><td style='padding:4px 12px;font-weight:bold;width:200px'>Run Started</td>
        <td style='padding:4px 12px'>{run_started}</td></tr>
    <tr style='background:#f9f9f9'>
        <td style='padding:4px 12px;font-weight:bold'>Executed By</td>
        <td style='padding:4px 12px'>{executed_by or 'N/A'}</td></tr>
    <tr><td style='padding:4px 12px;font-weight:bold'>Elapsed</td>
        <td style='padding:4px 12px'>{_fmt_duration(elapsed_seconds)}</td></tr>
    <tr style='background:#f9f9f9'>
        <td style='padding:4px 12px;font-weight:bold'>Outcome</td>
        <td style='padding:4px 12px;font-weight:bold;color:{"#27ae60" if outcome=="SUCCEEDED" else "#c0392b" if outcome=="FAILED" else "#e67e22"}'>
          {outcome}</td></tr>
  </tbody>
</table>
<h3 style='color:#2c3e50'>Summary</h3>
<table style='border-collapse:collapse;font-size:14px'>
  <tbody>
    <tr><td style='padding:3px 12px'>&#9989; Succeeded</td>
        <td style='padding:3px 12px;font-weight:bold;color:#27ae60'>{len(succeeded)}</td></tr>
    <tr><td style='padding:3px 12px'>&#9888; Partial</td>
        <td style='padding:3px 12px;font-weight:bold;color:#e67e22'>{len(partial)}</td></tr>
    <tr><td style='padding:3px 12px'>&#10060; Failed</td>
        <td style='padding:3px 12px;font-weight:bold;color:#c0392b'>{len(failed)}</td></tr>
    <tr><td style='padding:3px 12px'>&#9193; Skipped</td>
        <td style='padding:3px 12px'>{len(skipped)}</td></tr>
    <tr><td style='padding:3px 12px'>&#128202; Success Rate</td>
        <td style='padding:3px 12px;font-weight:bold'>{success_rate}</td></tr>
  </tbody>
</table>
{detail_table}
<p style='margin-top:20px;font-size:12px;color:#888'>
  Full details are attached as a report file and log excerpt.
  Re-run the migration script to automatically retry any failed or partial repositories.
</p>
</body></html>"""

    # Plain-text fallback
    plain_body = (
        f"GitLab -> GitHub Migration Run Complete\n"
        f"{'='*50}\n"
        f"Run Started  : {run_started}\n"
        f"Executed By  : {executed_by or 'N/A'}\n"
        f"Elapsed      : {_fmt_duration(elapsed_seconds)}\n"
        f"Outcome      : {outcome}\n\n"
        f"Summary\n"
        f"-------\n"
        f"  Succeeded   : {len(succeeded)}\n"
        f"  Partial     : {len(partial)}\n"
        f"  Failed      : {len(failed)}\n"
        f"  Skipped     : {len(skipped)}\n"
        f"  Success Rate: {success_rate}\n\n"
    )
    if failed or partial:
        plain_body += "Repositories requiring attention:\n"
        for r in failed + partial:
            err = r.error or r.custom_properties_error or r.ci_skeleton_error
            plain_body += f"  [{r.status.upper()}] {r.target}  -- {err[:120]}\n"

    # ---- assemble MIME message ----
    msg = email.mime.multipart.MIMEMultipart("mixed")
    msg["Subject"] = subject
    msg["From"]    = email.utils.formataddr(("GitMigration Bot", cfg.from_address or cfg.smtp_user))
    msg["To"]      = ", ".join(cfg.to_addresses)
    if cfg.cc_addresses:
        msg["Cc"]  = ", ".join(cfg.cc_addresses)
    msg["Date"]    = email.utils.formatdate(localtime=True)
    msg["X-Mailer"] = "GitMirrorMigration"

    # multipart/alternative for HTML + plain
    alt = email.mime.multipart.MIMEMultipart("alternative")
    alt.attach(email.mime.text.MIMEText(plain_body, "plain", "utf-8"))
    alt.attach(email.mime.text.MIMEText(html_body,  "html",  "utf-8"))
    msg.attach(alt)

    # Attach XLSX or CSV report
    if cfg.attach_report and report_path and report_path.exists():
        try:
            report_data = report_path.read_bytes()
            mime_type = (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                if report_path.suffix == ".xlsx" else "text/csv"
            )
            att = email.mime.base.MIMEBase(*mime_type.split("/", 1))
            att.set_payload(report_data)
            email.encoders.encode_base64(att)
            att.add_header(
                "Content-Disposition", "attachment",
                filename=report_path.name,
            )
            msg.attach(att)
        except Exception as att_exc:
            log.warning(f"[email] Could not attach report: {att_exc}")

    # Attach log tail
    if cfg.attach_log_tail_lines > 0 and log_path.exists():
        try:
            lines = log_path.read_text(encoding="utf-8", errors="replace").splitlines()
            tail  = "\n".join(lines[-cfg.attach_log_tail_lines:])
            log_att = email.mime.text.MIMEText(tail, "plain", "utf-8")
            log_att.add_header(
                "Content-Disposition", "attachment",
                filename=f"migration-log-tail-{run_started.replace(' ', '_').replace(':', '-')}.txt",
            )
            msg.attach(log_att)
        except Exception as log_exc:
            log.warning(f"[email] Could not attach log tail: {log_exc}")

    # ---- SMTP send ----
    all_recipients = cfg.to_addresses + cfg.cc_addresses
    try:
        if cfg.use_ssl:
            # Implicit TLS (port 465)
            import ssl as _ssl
            ctx = _ssl.create_default_context()
            with smtplib.SMTP_SSL(cfg.smtp_host, cfg.smtp_port, timeout=cfg.timeout, context=ctx) as srv:
                if cfg.smtp_user and cfg.smtp_password:
                    srv.login(cfg.smtp_user, cfg.smtp_password)
                srv.sendmail(cfg.smtp_user, all_recipients, msg.as_bytes())
        elif cfg.use_tls:
            # STARTTLS (port 587)
            import ssl as _ssl
            ctx = _ssl.create_default_context()
            with smtplib.SMTP(cfg.smtp_host, cfg.smtp_port, timeout=cfg.timeout) as srv:
                srv.ehlo()
                srv.starttls(context=ctx)
                srv.ehlo()
                if cfg.smtp_user and cfg.smtp_password:
                    srv.login(cfg.smtp_user, cfg.smtp_password)
                srv.sendmail(cfg.smtp_user, all_recipients, msg.as_bytes())
        else:
            # Plain SMTP (port 25, no encryption -- internal relay only)
            with smtplib.SMTP(cfg.smtp_host, cfg.smtp_port, timeout=cfg.timeout) as srv:
                srv.ehlo()
                if cfg.smtp_user and cfg.smtp_password:
                    srv.login(cfg.smtp_user, cfg.smtp_password)
                srv.sendmail(cfg.smtp_user, all_recipients, msg.as_bytes())

        log.info(
            f"[email] Notification sent to {', '.join(cfg.to_addresses)} "
            f"(subject: {subject!r})"
        )
    except Exception as exc:
        log.warning(f"[email] Failed to send notification: {exc}")


# ===========================================================================
# Report writing
# ===========================================================================

def _write_reports(
    results: list[_RepoResult],
    run_started: str,
    config: MigrationConfig,
    elapsed_seconds: float = 0.0,
    executed_by: str = "",
) -> None:
    """Write timestamped JSON and XLSX (or CSV fallback) reports to the reports folder."""
    succeeded = [r for r in results if r.status == "succeeded"]
    partial   = [r for r in results if r.status == "partial"]
    failed    = [r for r in results if r.status == "failed"]
    skipped   = [r for r in results if r.status == "skipped"]
    dry_run   = [r for r in results if r.status == "dry-run"]

    target_orgs = sorted({r.target.split("/")[0] for r in results if "/" in r.target})

    # ---- JSON ----------------------------------------------------------
    report: dict = {
        "run_timestamp": run_started,
        "executed_by": executed_by,
        "auth_mode": config.auth.mode_label,
        "dry_run": config.dry_run,
        "target_orgs": target_orgs,
        "gitlab_url": config.gitlab_url,
        "total_elapsed_seconds": round(elapsed_seconds, 1),
        "summary": {
            "total_in_csv": len(config.repos),
            "processed_this_run": len(succeeded) + len(partial) + len(failed) + len(dry_run),
            "succeeded": len(succeeded),
            "partial": len(partial),
            "failed": len(failed),
            "skipped_already_done": len(skipped),
            "dry_run": len(dry_run),
            "success_rate": (
                f"{len(succeeded)/( len(succeeded)+len(partial)+len(failed) )*100:.1f}%"
                if (len(succeeded) + len(partial) + len(failed)) else "N/A"
            ),
        },
        "timing_metrics": _compute_metrics(results),
        "repos": [
            {
                "source": r.source,
                "target": r.target,
                "gh_repo_url": r.gh_repo_url,
                "status": r.status,
                "visibility": r.visibility,
                "default_branch": r.default_branch,
                "default_branch_renamed": r.default_branch_renamed,
                "branch_count": r.branch_count,
                "head_commit_sha": r.head_commit_sha,
                "gh_branch_count": r.gh_branch_count,
                "gh_head_commit_sha": r.gh_head_commit_sha,
                "tag_count": r.tag_count,
                "gh_tag_count": r.gh_tag_count,
                "validation_status": r.validation_status,
                "validation_notes": r.validation_notes,
                "missing_branches": r.missing_branches,
                "missing_tags": r.missing_tags,
                "branch_sha_mismatches": r.branch_sha_mismatches,
                "lfs_detected": r.lfs_detected,
                "lfs_object_count": r.lfs_object_count,
                "ci_skeleton_status": r.ci_skeleton_status,
                "ci_skeleton_branches_created": r.ci_skeleton_branches_created,
                "ci_skeleton_branches_skipped": r.ci_skeleton_branches_skipped,
                "ci_skeleton_error": r.ci_skeleton_error,
                "custom_properties_status": r.custom_properties_status,
                "custom_properties_applied": r.custom_properties_applied,
                "custom_properties_error": r.custom_properties_error,
                "error": r.error,
                "duration_seconds": round(r.duration_seconds, 1),
                "duration_minutes": round(r.duration_seconds / 60, 2),
                "completed_at": r.completed_at,
            }
            for r in results
        ],
    }
    _RESULTS_JSON.write_text(json.dumps(report, indent=2), encoding="utf-8")
    log.info(f"JSON report : {_RESULTS_JSON}")

    # ---- XLSX / CSV ----------------------------------------------------
    if _XLSX_AVAILABLE:
        try:
            _write_xlsx(results, report, config, elapsed_seconds, executed_by)
            log.info(f"XLSX report : {_RESULTS_XLSX}")
        except Exception as xlsx_exc:
            log.warning(
                f"XLSX report generation failed ({xlsx_exc}) -- "
                "writing CSV fallback instead"
            )
            _write_csv_fallback(results)
            log.info(f"CSV report  : {_RESULTS_CSV}")
    else:
        _write_csv_fallback(results)
        log.info(f"CSV report  : {_RESULTS_CSV}")
        log.warning(
            "openpyxl not installed -- plain CSV written (no multi-sheet support). "
            "Install with:  pip install openpyxl"
        )

    log.info(f"Log file    : {_LOG_FILE}")

    # ---- Email notification ----------------------------------------
    if config.notification and config.notification.enabled:
        _report_path = _RESULTS_XLSX if _XLSX_AVAILABLE and _RESULTS_XLSX.exists() else (
            _RESULTS_CSV if _RESULTS_CSV.exists() else None
        )
        _send_migration_email(
            cfg=config.notification,
            results=results,
            run_started=run_started,
            elapsed_seconds=elapsed_seconds,
            executed_by=executed_by,
            report_path=_report_path,
            log_path=_LOG_FILE,
        )


def _write_csv_fallback(results: list[_RepoResult]) -> None:
    """Write a plain CSV report when openpyxl is not available."""
    with _RESULTS_CSV.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow([
            "source", "target", "gh_repo_url", "status", "visibility",
            "default_branch", "default_branch_renamed", "duration_min",
            "branch_count", "head_commit_sha", "tag_count",
            "gh_branch_count", "gh_head_commit_sha", "gh_tag_count",
            "validation_status", "validation_notes",
            "missing_branches", "missing_tags", "branch_sha_mismatches",
            "lfs_detected", "lfs_object_count",
            "custom_properties_status", "custom_properties_applied", "custom_properties_error",
            "ci_skeleton_status", "ci_skeleton_branches_created",
            "ci_skeleton_branches_skipped", "ci_skeleton_error",
            "error", "completed_at",
        ])
        for r in results:
            _cp_str = (
                "; ".join(f"{k}={v!r}" for k, v in r.custom_properties_applied.items())
                if r.custom_properties_applied else ""
            )
            writer.writerow([
                r.source, r.target, r.gh_repo_url, r.status, r.visibility,
                r.default_branch, r.default_branch_renamed,
                round(r.duration_seconds / 60, 3) if r.duration_seconds else "",
                r.branch_count or "", r.head_commit_sha, r.tag_count or "",
                r.gh_branch_count or "", r.gh_head_commit_sha, r.gh_tag_count or "",
                r.validation_status, r.validation_notes,
                "; ".join(r.missing_branches) if r.missing_branches else "",
                "; ".join(r.missing_tags) if r.missing_tags else "",
                "; ".join(r.branch_sha_mismatches) if r.branch_sha_mismatches else "",
                "Yes" if r.lfs_detected else "No",
                r.lfs_object_count or "",
                r.custom_properties_status,
                _cp_str,
                r.custom_properties_error,
                r.ci_skeleton_status,
                "; ".join(r.ci_skeleton_branches_created) if r.ci_skeleton_branches_created else "",
                "; ".join(r.ci_skeleton_branches_skipped) if r.ci_skeleton_branches_skipped else "",
                r.ci_skeleton_error,
                r.error, r.completed_at,
            ])


def _write_xlsx(
    results: list[_RepoResult],
    report: dict,
    config: MigrationConfig,
    elapsed_seconds: float,
    executed_by: str = "",
) -> None:
    """Write a multi-sheet XLSX report: Repositories, Run Summary, Validation."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    def _fill(hex_color: str) -> PatternFill:
        return PatternFill("solid", fgColor=hex_color)

    HDR_FILL     = _fill("1F4E79")
    HDR_FONT     = Font(bold=True, color="FFFFFF", size=10)
    STATUS_FILLS = {
        "succeeded": _fill("C6EFCE"),
        "partial":   _fill("FFEB9C"),
        "failed":    _fill("FFC7CE"),
        "skipped":   _fill("F2F2F2"),
        "dry-run":   _fill("BDD7EE"),
    }
    VAL_FILLS = {
        "match":    _fill("C6EFCE"),
        "mismatch": _fill("FFC7CE"),
        "unknown":  _fill("FFEB9C"),
    }
    LEFT   = Alignment(horizontal="left",   vertical="center")
    CENTER = Alignment(horizontal="center", vertical="center")
    WRAP   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    def _header_row(ws, headers: list) -> None:
        for ci, hdr in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=hdr)
            c.font = HDR_FONT
            c.fill = HDR_FILL
            c.alignment = CENTER
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 22

    def _col_widths(ws, widths: list) -> None:
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

    def _hyperlink_cell(ws, row: int, col: int, url: str, label: str) -> None:
        c = ws.cell(row=row, column=col, value=label)
        c.hyperlink = url
        c.font = Font(color="0563C1", underline="single", size=10)
        c.alignment = LEFT

    def _match_icon(src_val: object, gh_val: object) -> str:
        if not src_val or not gh_val:
            return ""
        return "\u2713" if src_val == gh_val else "\u2717"

    def _recommended_action(r: _RepoResult) -> str:
        if r.status == "failed":
            err = r.error.lower()
            if "not found" in err or "404" in err:
                return "Verify source repo exists on GitLab and namespace is correct"
            if "permission" in err or "403" in err or "401" in err:
                return "Check GitLab PAT has 'read_repository' scope; GitHub PAT needs 'repo' scope"
            if "timed out" in err:
                return "Repo may be large -- increase clone_timeout_seconds / push_timeout_seconds"
            if "create github" in err or "github repo" in err:
                return "Check GitHub PAT has 'repo' scope and target org exists"
            return "Re-run script to retry -- check log file for detailed error"
        if r.validation_status == "mismatch":
            return f"Investigate mismatch: {r.validation_notes}"
        return ""

    succeeded  = [r for r in results if r.status == "succeeded"]
    partial    = [r for r in results if r.status == "partial"]
    failed     = [r for r in results if r.status == "failed"]
    skipped    = [r for r in results if r.status == "skipped"]
    dry_run_rs = [r for r in results if r.status == "dry-run"]
    mismatches = [r for r in results if r.validation_status == "mismatch"]
    validated  = [r for r in results if r.status not in ("skipped", "dry-run")]
    val_passed = sum(1 for r in validated if r.validation_status == "match")
    val_total  = len(validated)
    target_orgs = sorted({r.target.split("/")[0] for r in results if "/" in r.target})

    hostname = socket.gethostname()
    m = report.get("timing_metrics", {})
    wb = openpyxl.Workbook()

    # =========================================================================
    # Sheet 1: Migration Summary (management view)
    # =========================================================================
    ws_sum = wb.active
    ws_sum.title = "Migration Summary"
    ws_sum.column_dimensions["A"].width = 28
    ws_sum.column_dimensions["B"].width = 48

    def _sum_row(label: str, value: str, is_section: bool = False) -> None:
        ri = ws_sum.max_row + 1
        cl = ws_sum.cell(row=ri, column=1, value=label)
        cv = ws_sum.cell(row=ri, column=2, value=value)
        cl.alignment = LEFT
        cv.alignment = LEFT
        if is_section:
            cl.font = Font(bold=True, color="1F4E79", size=10)
            cl.fill = _fill("DDEEFF")
            cv.fill = _fill("DDEEFF")
        elif label:
            cl.font = Font(bold=True, size=10)

    _sum_row("RUN INFORMATION", "", is_section=True)
    _sum_row("Run Date / Time",        report["run_timestamp"])
    _sum_row("Executed By",            executed_by or "\u2014")
    _sum_row("Executed On",            hostname)
    _sum_row("Auth Mode",              config.auth.mode_label)
    _sum_row("Source (GitLab)",        config.gitlab_url)
    _sum_row("Target (GitHub)",        config.github_url)
    _sum_row("Target Org(s)",          ", ".join(target_orgs) if target_orgs else "\u2014")
    _sum_row("Dry Run",                "Yes" if config.dry_run else "No")
    _sum_row("Rename to main",         "Yes" if config.rename_default_branch else "No")
    _sum_row("", "")
    _sum_row("MIGRATION RESULTS", "", is_section=True)
    _sum_row("Total Repos in CSV",         str(len(config.repos)))
    _sum_row("Succeeded (fully)",          str(len(succeeded)))
    _sum_row("Partial (git ok, sub-step failed)", str(len(partial)))
    _sum_row("Failed",                     str(len(failed)))
    _processed = len(succeeded) + len(partial) + len(failed)
    _srate = f"{len(succeeded)}/{_processed} ({len(succeeded)/_processed*100:.1f}%)" if _processed else "N/A"
    _sum_row("Full-Success Rate",          _srate)
    _sum_row("Skipped (already migrated)", str(len(skipped)))
    _sum_row("Dry-Run Simulated",          str(len(dry_run_rs)))
    _sum_row("", "")
    _sum_row("VALIDATION SUMMARY", "", is_section=True)
    pass_rate = (
        f"{val_passed}/{val_total} ({val_passed/val_total*100:.1f}%)" if val_total else "N/A"
    )
    _sum_row("Validation Pass Rate",   pass_rate)
    _sum_row("Mismatches",             str(len(mismatches)))
    _sum_row("", "")
    _sum_row("TIMING", "", is_section=True)
    _sum_row("Total Duration", _fmt_duration(elapsed_seconds))
    if m:
        _sum_row("Min / Max per Repo",
                 f"{m.get('min_seconds', 0)}s  /  {m.get('max_seconds', 0)}s")
        _sum_row("Mean per Repo", f"{m.get('mean_seconds', 0)}s")
    if failed or partial:
        _sum_row("", "")
        _sum_row("FAILED / PARTIAL REPOSITORIES", "", is_section=True)
        for r in failed:
            _sum_row(r.source, f"FAILED: {r.error[:110]}" if r.error else "FAILED")
        for r in partial:
            parts = []
            if r.custom_properties_status == "failed":
                parts.append(f"props: {r.custom_properties_error[:60]}")
            if r.ci_skeleton_status == "failed":
                parts.append(f"ci: {r.ci_skeleton_error[:60]}")
            _sum_row(r.source, "PARTIAL: " + "; ".join(parts) if parts else "PARTIAL")

    # =========================================================================
    # Sheet 2: Repositories (team leads -- full list with GitHub URLs)
    # =========================================================================
    ws_repos = wb.create_sheet("Repositories")
    _header_row(ws_repos, [
        "Source (GitLab)", "Target (GitHub)", "GitHub URL",
        "Status", "Visibility", "Default Branch", "Renamed to main",
        "Branches Pushed", "Tags Pushed", "LFS Detected", "LFS Objects",
        "Custom Props Status", "Custom Props Applied", "Custom Props Error",
        "CI Skeleton", "CI Branches Created", "CI Branches Skipped", "CI Error",
        "Duration (min)", "Completed At", "Git Error",
    ])
    CP_STATUS_FILLS = {
        "applied":        _fill("C6EFCE"),
        "skipped":        _fill("F2F2F2"),
        "not_configured": _fill("F2F2F2"),
        "failed":         _fill("FFC7CE"),
        "dry-run":        _fill("BDD7EE"),
    }
    for ri, r in enumerate(results, 2):
        rf = STATUS_FILLS.get(r.status, _fill("FFFFFF"))
        cp_fill = CP_STATUS_FILLS.get(r.custom_properties_status, rf)
        ci_fill = (
            _fill("C6EFCE") if r.ci_skeleton_status == "created" else
            _fill("FFEB9C") if r.ci_skeleton_status in ("partial", "skipped_all") else
            _fill("FFC7CE") if r.ci_skeleton_status == "failed" else
            rf
        )
        _cp_applied_str = (
            ", ".join(f"{k}={v!r}" for k, v in r.custom_properties_applied.items())
            if r.custom_properties_applied else ""
        )
        vals = [
            r.source, r.target, None,
            r.status,
            r.visibility or "\u2014",
            r.default_branch or "\u2014",
            "Yes" if r.default_branch_renamed else ("No" if r.status in ("succeeded", "partial") else ""),
            r.branch_count or "",
            r.tag_count or "",
            "Yes" if r.lfs_detected else ("No" if r.status in ("succeeded", "partial") else ""),
            r.lfs_object_count if r.lfs_detected else "",
            r.custom_properties_status or "",
            _cp_applied_str,
            r.custom_properties_error or "",
            r.ci_skeleton_status or "",
            "; ".join(r.ci_skeleton_branches_created) if r.ci_skeleton_branches_created else "",
            "; ".join(r.ci_skeleton_branches_skipped) if r.ci_skeleton_branches_skipped else "",
            r.ci_skeleton_error or "",
            round(r.duration_seconds / 60, 2) if r.duration_seconds else "",
            r.completed_at,
            r.error or "",
        ]
        # Columns that should wrap: Custom Props Applied(13), CP Error(14), CI Error(18), Git Error(21)
        _wrap_cols_repos = {13, 14, 18, 21}
        # Column fill: cp cols 12-14 use cp_fill; ci cols 15-18 use ci_fill; rest use rf
        _cp_cols  = {12, 13, 14}
        _ci_cols  = {15, 16, 17, 18}
        for ci, val in enumerate(vals, 1):
            if ci == 3:
                if r.gh_repo_url:
                    _hyperlink_cell(ws_repos, ri, ci, r.gh_repo_url, r.gh_repo_url)
                    ws_repos.cell(ri, ci).fill = rf
                continue
            c = ws_repos.cell(row=ri, column=ci, value=val)
            c.fill = cp_fill if ci in _cp_cols else ci_fill if ci in _ci_cols else rf
            c.alignment = WRAP if ci in _wrap_cols_repos else LEFT
    _col_widths(ws_repos, [42, 38, 52, 12, 12, 18, 16, 14, 12, 12, 12, 18, 42, 40, 14, 30, 30, 40, 14, 22, 45])

    # =========================================================================
    # Sheet 3: Failed & Mismatches (engineers -- actionable only)
    # =========================================================================
    ws_fail = wb.create_sheet("Failed & Mismatches")
    _header_row(ws_fail, [
        "Source (GitLab)", "Target (GitHub)", "GitHub URL",
        "Issue Type", "Sub-step", "Error / Mismatch Detail", "Recommended Action",
    ])
    action_rows: list[tuple] = []
    for r in results:
        if r.status == "failed":
            action_rows.append((r, "Failed", "git", r.error))
        elif r.status == "partial":
            if r.custom_properties_status == "failed":
                action_rows.append((r, "Partial", "custom-properties", r.custom_properties_error))
            if r.ci_skeleton_status == "failed":
                action_rows.append((r, "Partial", "ci-skeleton", r.ci_skeleton_error))
        elif r.validation_status == "mismatch":
            action_rows.append((r, "Validation Mismatch", "git-validate", r.validation_notes))
    for ri, (r, issue_type, substep, detail) in enumerate(action_rows, 2):
        rf = STATUS_FILLS.get(r.status, VAL_FILLS.get(r.validation_status, _fill("FFFFFF")))
        vals = [r.source, r.target, None, issue_type, substep, detail, _recommended_action(r)]
        for ci, val in enumerate(vals, 1):
            if ci == 3:
                if r.gh_repo_url:
                    _hyperlink_cell(ws_fail, ri, ci, r.gh_repo_url, r.gh_repo_url)
                    ws_fail.cell(ri, ci).fill = rf
                continue
            c = ws_fail.cell(row=ri, column=ci, value=val)
            c.fill = rf
            c.alignment = WRAP if ci in (6, 7) else LEFT
    _col_widths(ws_fail, [42, 38, 52, 16, 20, 55, 55])
    if not action_rows:
        ws_fail.cell(row=2, column=1,
            value="\u2705  No failures or mismatches. All repositories migrated successfully.")

    # =========================================================================
    # Sheet 4: Validation Detail (DevOps / audit sign-off)
    # =========================================================================
    ws_val = wb.create_sheet("Validation Detail")
    _header_row(ws_val, [
        "Source (GitLab)", "Target (GitHub)", "GitHub URL", "Migration Status",
        "Src Branches", "GH Branches", "Branches \u2713/\u2717",
        "Src Tags", "GH Tags", "Tags \u2713/\u2717",
        # Column K: GitLab HEAD at time of push
        # Column L: GitHub HEAD at time of push (should equal K; may differ if
        #           .github CI skeleton was added after push -- see column M note)
        "GitLab HEAD (push)", "GitHub HEAD (push)", "HEAD \u2713/\u2717",
        "Default Branch", "Renamed to main",
        "Validation Status", "Validation Notes",
        "Missing Branches", "Missing Tags", "Branch SHA Mismatches",
    ])
    # Tooltip / note fills for the ahead-case HEAD match cell
    _AHEAD_FILL = _fill("E2EFDA")   # soft green -- "match, but GitHub is ahead"
    for ri, r in enumerate(
        (r for r in results if r.status not in ("skipped", "dry-run")), 2
    ):
        rf = VAL_FILLS.get(r.validation_status, _fill("FFFFFF"))
        # HEAD match logic: if GitHub is strictly ahead of GitLab (e.g. .github
        # CI skeleton commit was added), the migration is still correct -- GitLab
        # history is fully present.  Show ✓* instead of ✗.
        _head_match: str
        if r.gh_head_is_ahead:
            _head_match = "\u2713*"  # ✓* = match with extra commits
        else:
            _head_match = _match_icon(r.head_commit_sha, r.gh_head_commit_sha)
        vals = [
            r.source, r.target, None,
            r.status,
            r.branch_count or "",
            r.gh_branch_count or "",
            _match_icon(r.branch_count, r.gh_branch_count),
            r.tag_count or "",
            r.gh_tag_count or "",
            _match_icon(r.tag_count, r.gh_tag_count),
            r.head_commit_sha[:10] if r.head_commit_sha else "",
            r.gh_head_commit_sha[:10] if r.gh_head_commit_sha else "",
            _head_match,
            r.default_branch or "\u2014",
            "Yes" if r.default_branch_renamed else "No",
            r.validation_status,
            r.validation_notes,
            "; ".join(r.missing_branches) if r.missing_branches else "",
            "; ".join(r.missing_tags) if r.missing_tags else "",
            "; ".join(r.branch_sha_mismatches) if r.branch_sha_mismatches else "",
        ]
        # Column indices that should wrap: notes(17), missing_branches(18),
        # missing_tags(19), branch_sha_mismatches(20)
        _wrap_cols = {17, 18, 19, 20}
        for ci, val in enumerate(vals, 1):
            if ci == 3:
                if r.gh_repo_url:
                    _hyperlink_cell(ws_val, ri, ci, r.gh_repo_url, r.gh_repo_url)
                    ws_val.cell(ri, ci).fill = rf
                continue
            c = ws_val.cell(row=ri, column=ci, value=val)
            # Column 13 = "HEAD ✓/✗": use soft-green fill when GitHub is ahead
            # (migration correct -- GitLab history present -- GitHub just has
            # extra .github CI skeleton commits on top).
            if ci == 13 and r.gh_head_is_ahead:
                c.fill = _AHEAD_FILL
            else:
                c.fill = rf
            c.alignment = WRAP if ci in _wrap_cols else LEFT
    _col_widths(ws_val, [42, 38, 52, 14, 14, 14, 14, 12, 12, 12, 14, 14, 14, 18, 16, 18, 50, 45, 45, 55])

    # =========================================================================
    # Sheet 5: Run Metrics (internal ops -- not for client)
    # =========================================================================
    ws_met = wb.create_sheet("Run Metrics")
    ws_met.column_dimensions["A"].width = 28
    ws_met.column_dimensions["B"].width = 42

    def _met_row(label: str, value: str, is_section: bool = False) -> None:
        ri = ws_met.max_row + 1
        cl = ws_met.cell(row=ri, column=1, value=label)
        cv = ws_met.cell(row=ri, column=2, value=value)
        cl.alignment = LEFT
        cv.alignment = LEFT
        if is_section:
            cl.font = Font(bold=True, color="1F4E79", size=10)
            cl.fill = _fill("DDEEFF")
            cv.fill = _fill("DDEEFF")
        elif label:
            cl.font = Font(bold=True, size=10)

    _met_row("CONFIGURATION", "", is_section=True)
    _met_row("Auth Mode",          config.auth.mode_label)
    _met_row("Concurrent Workers", str(config.max_workers))
    _met_row("Clone Timeout",      f"{config.clone_timeout}s")
    _met_row("Push Timeout",       f"{config.push_timeout}s")
    _met_row("Git Retries",        str(config.git_max_retries))
    _met_row("Min Free Disk",      f"{config.min_free_disk_gb} GB")
    _met_row("LFS Enabled",        str(config.lfs_enabled))
    _met_row("LFS Available",      "Yes" if _LFS_AVAILABLE else "No (git-lfs not installed)")
    _met_row("CI Skeleton",        "Enabled" if (config.ci_skeleton and config.ci_skeleton.enabled) else "Disabled")
    if config.ci_skeleton and config.ci_skeleton.enabled:
        _met_row("  Templates Dir",  str(config.ci_skeleton.templates_dir.name))
        _met_row("  Target Path",    config.ci_skeleton.target_path)
        _met_row("  Branches",       ", ".join(config.ci_skeleton.branches))
        _met_row("  Skip If Exists", str(config.ci_skeleton.skip_if_exists))
    _met_row("Detailed Commit Count", str(config.detailed_commit_count))
    _met_row("Dry Run",            str(config.dry_run))
    _met_row("Rename to main",     str(config.rename_default_branch))
    _met_row("", "")
    _met_row("TIMING", "", is_section=True)
    _met_row("Total Duration", _fmt_duration(elapsed_seconds))
    if m:
        _met_row("Min per Repo",   f"{m.get('min_seconds', 0)}s")
        _met_row("Max per Repo",   f"{m.get('max_seconds', 0)}s")
        _met_row("Mean per Repo",  f"{m.get('mean_seconds', 0)}s")

    # =========================================================================
    # Sheet 6: Custom Properties Results
    # =========================================================================
    ws_cp = wb.create_sheet("Custom Properties")
    _header_row(ws_cp, [
        "Target (GitHub)", "GitHub URL",
        "Status", "Properties Applied", "Error",
    ])
    cp_rows = [r for r in results if r.status not in ("skipped",)]
    for ri, r in enumerate(cp_rows, 2):
        rf = CP_STATUS_FILLS.get(r.custom_properties_status, STATUS_FILLS.get(r.status, _fill("FFFFFF")))
        _cp_str = (
            "\n".join(f"{k} = {v!r}" for k, v in r.custom_properties_applied.items())
            if r.custom_properties_applied else "\u2014"
        )
        cp_vals = [
            r.target, None,
            r.custom_properties_status or "not_configured",
            _cp_str,
            r.custom_properties_error or "",
        ]
        for ci, val in enumerate(cp_vals, 1):
            if ci == 2:
                if r.gh_repo_url:
                    _hyperlink_cell(ws_cp, ri, ci, r.gh_repo_url, r.gh_repo_url)
                    ws_cp.cell(ri, ci).fill = rf
                continue
            c = ws_cp.cell(row=ri, column=ci, value=val)
            c.fill = rf
            c.alignment = WRAP if ci in (4, 5) else LEFT
    _col_widths(ws_cp, [45, 52, 18, 55, 50])
    if not cp_rows:
        ws_cp.cell(row=2, column=1, value="\u2139\ufe0f  No repos processed this run.")

    # =========================================================================
    # Sheet 7: CI Skeleton Results
    # =========================================================================
    ws_ci = wb.create_sheet("CI Skeleton")
    _header_row(ws_ci, [
        "Target (GitHub)", "GitHub URL", "Template File",
        "CI Status", "Branches Created", "Branches Skipped (exists)", "Error",
    ])
    CI_STATUS_FILLS = {
        "created":     _fill("C6EFCE"),
        "partial":     _fill("FFEB9C"),
        "skipped_all": _fill("F2F2F2"),
        "failed":      _fill("FFC7CE"),
        "dry-run":     _fill("FFEB9C"),
        "not_configured": _fill("F2F2F2"),
    }
    ci_rows = [r for r in results if r.status not in ("skipped",)]
    for ri, r in enumerate(ci_rows, 2):
        rf = CI_STATUS_FILLS.get(r.ci_skeleton_status, _fill("FFFFFF"))
        ci_vals = [
            r.target, None,
            "",   # template filename -- enriched below if available
            r.ci_skeleton_status or "not_configured",
            "; ".join(r.ci_skeleton_branches_created) if r.ci_skeleton_branches_created else "\u2014",
            "; ".join(r.ci_skeleton_branches_skipped) if r.ci_skeleton_branches_skipped else "\u2014",
            r.ci_skeleton_error or "",
        ]
        for ci, val in enumerate(ci_vals, 1):
            if ci == 2:
                if r.gh_repo_url:
                    _hyperlink_cell(ws_ci, ri, ci, r.gh_repo_url, r.gh_repo_url)
                    ws_ci.cell(ri, ci).fill = rf
                continue
            c = ws_ci.cell(row=ri, column=ci, value=val)
            c.fill = rf
            c.alignment = WRAP if ci in (5, 6, 7) else LEFT
    _col_widths(ws_ci, [45, 52, 22, 16, 40, 40, 45])
    if not ci_rows:
        ws_ci.cell(row=2, column=1, value="\u2139\ufe0f  CI skeleton was not enabled during this run.")

    wb.save(str(_RESULTS_XLSX))


# ===========================================================================
# Per-repo worker (runs in thread pool)
# ===========================================================================

def _migrate_one(
    spec: RepoSpec,
    config: MigrationConfig,
    client: GitHubClient,
    state: MigrationState,
) -> _RepoResult:
    """Migrate one repository. Designed to be submitted to a ThreadPoolExecutor."""
    key = f"{spec.namespace}/{spec.project}"
    target_key = f"{spec.target_org}/{spec.target_name}"
    name = spec.target_name
    t_start = time.monotonic()

    # ── Step 1: git clone + push + validate ─────────────────────────────────
    mr = _mirror_repo(spec, config, client)

    # ── Step 2: Custom properties ────────────────────────────────────────────
    cp_status = "not_configured"
    cp_applied: dict = {}
    cp_error = ""

    if config.dry_run:
        cp_status = "dry-run"
    elif mr.ok:
        _raw_props = config.repo_custom_properties.get(target_key)
        if _raw_props:
            _schema = config._org_property_schemas.get(spec.target_org, {})
            _ok, _err = _set_repo_properties(
                client, spec.target_org, name, _raw_props, target_key,
                schema_map=_schema,
            )
            if _ok:
                cp_status = "applied"
                # Coerce values for display (mirrors what _set_repo_properties sends)
                cp_applied = {
                    k: _coerce_property_value(v, _schema.get(k, "string"), prop_name=k)
                    for k, v in _raw_props.items()
                }
                log.info(
                    f"[{name}] Custom properties applied ({len(cp_applied)}): "
                    + ", ".join(
                        f"{k}={v!r}" for k, v in cp_applied.items()
                    )
                )
            else:
                cp_status = "failed"
                cp_error = _err
                log.warning(f"[{name}] Custom properties FAILED: {_err}")
        else:
            cp_status = "skipped"
            log.debug(f"[{name}] Custom properties: not in repo-properties.csv -- skipped")

    # ── Step 3: CI skeleton ──────────────────────────────────────────────────
    ci_result = _CiSkeletonResult(status="not_configured")
    if config.ci_skeleton and config.ci_skeleton.enabled:
        if config.dry_run or mr.ok:
            ci_result = _create_ci_skeleton(spec, config, client, mr.default_branch)
            if ci_result.status == "created":
                log.info(
                    f"[{name}] CI skeleton created on: "
                    + ", ".join(ci_result.branches_created)
                )
            elif ci_result.status == "skipped_all":
                log.info(
                    f"[{name}] CI skeleton: already exists on all target branches -- skipped"
                )
            elif ci_result.status == "partial":
                log.warning(
                    f"[{name}] CI skeleton partial: created={ci_result.branches_created}, "
                    f"skipped={ci_result.branches_skipped}"
                    + (f", error={ci_result.error}" if ci_result.error else "")
                )
            elif ci_result.status == "failed":
                log.warning(
                    f"[{name}] CI skeleton FAILED: {ci_result.error}"
                )

    # ── Determine overall status ─────────────────────────────────────────────
    # "succeeded"  : git ok AND (no props configured OR props applied) AND (CI not failed)
    # "partial"    : git ok BUT props or CI had a non-blocking failure
    # "failed"     : git push/validate failed (hard failure)
    if config.dry_run:
        status = "dry-run"
    elif not mr.ok:
        status = "failed"
    else:
        props_ok = cp_status in ("applied", "skipped", "not_configured", "dry-run")
        ci_ok = ci_result.status not in ("failed",)
        if props_ok and ci_ok:
            status = "succeeded"
        else:
            status = "partial"
            partial_reasons: list[str] = []
            if not props_ok:
                partial_reasons.append(f"custom-properties({cp_error[:80]})")
            if not ci_ok:
                partial_reasons.append(f"ci-skeleton({ci_result.error[:80]})")
            log.warning(
                f"[{name}] Migration PARTIAL -- git succeeded but: "
                + "; ".join(partial_reasons)
            )

    if not config.dry_run:
        # Checkpoint only marks fully succeeded or failed -- partial stays unresolved
        # so a re-run can retry the failed sub-steps.
        state.record(key, "succeeded" if status == "succeeded" else
                          "failed" if status == "failed" else "partial")

    return _RepoResult(
        source=key,
        target=target_key,
        status=status,
        default_branch=mr.default_branch,
        error=mr.error,
        duration_seconds=time.monotonic() - t_start,
        completed_at=_fmt_ts(),
        visibility=spec.visibility,
        gh_repo_url=mr.gh_repo_url,
        branch_count=mr.branch_count,
        head_commit_sha=mr.head_commit_sha,
        gh_branch_count=mr.gh_branch_count,
        gh_head_commit_sha=mr.gh_head_commit_sha,
        gh_head_is_ahead=mr.gh_head_is_ahead,
        tag_count=mr.tag_count,
        gh_tag_count=mr.gh_tag_count,
        commit_count=mr.commit_count,
        gh_commit_count=mr.gh_commit_count,
        default_branch_renamed=mr.default_branch_renamed,
        validation_status=mr.validation_status,
        validation_notes=mr.validation_notes,
        missing_branches=mr.missing_branches,
        missing_tags=mr.missing_tags,
        branch_sha_mismatches=mr.branch_sha_mismatches,
        lfs_detected=mr.lfs_detected,
        lfs_object_count=mr.lfs_object_count,
        ci_skeleton_status=ci_result.status,
        ci_skeleton_branches_created=ci_result.branches_created,
        ci_skeleton_branches_skipped=ci_result.branches_skipped,
        ci_skeleton_error=ci_result.error,
        custom_properties_status=cp_status,
        custom_properties_applied=cp_applied,
        custom_properties_error=cp_error,
    )


# ===========================================================================
# Pre-flight preview and confirmation
# ===========================================================================

def _print_migration_preview(
    config: MigrationConfig,
    pending: list[RepoSpec],
    already_done: list[RepoSpec],
    config_path: Path | None = None,
    repos_csv_path: Path | None = None,
) -> bool:

    W = 72
    BORDER  = "\u2550" * W   # ═══
    DIVIDER = "\u2500" * W   # ───

    def _header(title: str) -> None:
        pad = (W - len(title) - 2) // 2
        print(f"\n\u2554{BORDER}\u2557")
        print(f"\u2551{' ' * pad} {title} {' ' * (W - pad - len(title) - 1)}\u2551")
        print(f"\u255a{BORDER}\u255d")

    def _kv(icon: str, label: str, value: str, width: int = 20) -> None:
        print(f"  {icon}  {label:<{width}} {value}")

    def _divider() -> None:
        print(f"  {DIVIDER}")

    _header("GitLab → GitHub Mirror Migration  │  Pre-flight Preview")

    # ---- configuration panel ------------------------------------------
    print()
    if config.dry_run:
        print("  \u26a0\ufe0f  DRY RUN MODE \u2014 no repositories will be created or modified")
        print()

    _kv("\U0001f99a", "Source  (GitLab)",   config.gitlab_url)
    _kv("\U0001f431", "Target  (GitHub)",   config.github_url)
    _kv("\U0001f511", "Auth mode",          config.auth.mode_label)
    _kv("\U0001f9f5", "Concurrent workers", str(config.max_workers))
    _kv("\u23f1\ufe0f",  "Clone timeout",     f"{config.clone_timeout} s")
    _kv("\u23f1\ufe0f",  "Push timeout",      f"{config.push_timeout} s")
    _kv("\U0001f504", "Git retries",        str(config.git_max_retries))
    if config.branch_include:
        _kv("\U0001f33f", "Branch include",  ', '.join(config.branch_include))
    if config.branch_exclude:
        _kv("\u274c",    "Branch exclude",   ', '.join(config.branch_exclude))

    # ---- checkpoint summary -------------------------------------------
    print()
    _divider()
    if already_done:
        _kv("\u2714\ufe0f", "Already done",
            f"{len(already_done)} repo(s) \u2014 will be skipped (checkpoint)")
    _kv("\U0001f4e6", "Pending", f"{len(pending)} repo(s) to migrate")
    _divider()

    if not pending:
        print()
        print("  \u2705  All repositories have already been migrated. Nothing to do.")
        print()
        return True

    # ---- repository preview table -------------------------------------
    show     = pending[:_PREVIEW_MAX_ROWS]
    overflow = len(pending) - len(show)

    HDR_SRC = "\U0001f99a Source (GitLab namespace/project)"
    HDR_TGT = "\U0001f431 Target (GitHub org/repo)"
    HDR_VIS = "\U0001f512 Visibility"
    HDR_BR  = "\U0001f33f Effective Branch Filter"

    any_branch_filter = (
        bool(config.branch_include or config.branch_exclude)
        or any(s.branch_include or s.branch_exclude for s in pending)
    )

    def _fmt_pats(pats: list[str], glyph: str) -> str:
        shown = [p[:17] + "\u2026" if len(p) > 17 else p for p in pats[:2]]
        extra = len(pats) - 2
        return f"{glyph} " + "; ".join(shown) + (f" +{extra}\u2026" if extra > 0 else "")

    def _branch_label(s: RepoSpec) -> str:
        if not s.branch_include and not s.branch_exclude:
            return "\u2714 all"
        src = "\U0001f310" if s.branch_from_global else "\U0001f4cb"
        parts = [
            p for p in [
                _fmt_pats(s.branch_include, "\u2795") if s.branch_include else "",
                _fmt_pats(s.branch_exclude, "\u2796") if s.branch_exclude else "",
            ] if p
        ]
        return f"{src} " + "  ".join(parts)


    src_w = min(max(len(HDR_SRC), max(len(f"{s.namespace}/{s.project}") for s in show)), 72)
    tgt_w = min(max(len(HDR_TGT), max(len(f"{s.target_org}/{s.target_name}") for s in show)), 56)
    vis_w = max(len(HDR_VIS), max(len(s.visibility) + 3 for s in show))  # +3 for vis icon

    cols_w = [src_w, tgt_w, vis_w]
    if any_branch_filter:
        cols_w.append(max(len(HDR_BR), max(len(_branch_label(s)) for s in show)))

    def _tsep(top: bool = False, bottom: bool = False) -> str:
        lc, mc, rc = (
            ("\u250c", "\u252c", "\u2510") if top else
            ("\u2514", "\u2534", "\u2518") if bottom else
            ("\u251c", "\u253c", "\u2524")
        )
        return "  " + lc + (mc).join("\u2500" * (w + 2) for w in cols_w) + rc

    def _trow(*cells: str) -> str:
        parts = []
        for cell, w in zip(cells, cols_w):
            c = cell[:w - 3] + "\u2026" if len(cell) > w else cell
            parts.append(f" {c:<{w}} ")
        return "  \u2502" + "\u2502".join(parts) + "\u2502"

    headers = [HDR_SRC, HDR_TGT, HDR_VIS]
    if any_branch_filter:
        headers.append(HDR_BR)

    print()
    print(_tsep(top=True))
    print(_trow(*headers))
    print(_tsep())
    for spec in show:
        cells = [
            f"{spec.namespace}/{spec.project}",
            f"{spec.target_org}/{spec.target_name}",
            f"{_VIS_ICONS.get(spec.visibility, '')} {spec.visibility}",
        ]
        if any_branch_filter:
            cells.append(_branch_label(spec))
        print(_trow(*cells))
    print(_tsep(bottom=True))

    if any_branch_filter:
        print(
            "  \U0001f4cb per-repo (repos.csv)  "
            "\U0001f310 global fallback (mirror-config.json)  "
            "\u2714 all  \u2795 include  \u2796 exclude"
        )

    if overflow:
        print(f"\n  \u2139\ufe0f  ... and {overflow} more repo(s) not shown. See repos.csv for the full list.")

    # ---- confirmation prompt ------------------------------------------
    if not sys.stdin.isatty():
        print("  \U0001f916  Non-interactive mode \u2014 proceeding automatically.")
        print()
        return True

    try:
        prompt = (
            "  \u26a0\ufe0f  Proceed with DRY RUN? [y/N]: "
            if config.dry_run
            else "  \u25b6\ufe0f  Proceed with migration? [y/N]: "
        )
        answer = input(prompt).strip().lower()
    except (EOFError, KeyboardInterrupt):
        print()
        return False

    print()
    if answer in ("y", "yes"):
        return True

    return False


# ===========================================================================
# Orchestration
# ===========================================================================

def migrate_all(config: MigrationConfig, config_path: Path, repos_csv_path: Path, batch_size: int = 0) -> None:
    """Migrate all repos concurrently with checkpoint, rate limiting, and graceful shutdown."""
    if shutil.which("git") is None:
        raise RuntimeError(
            "'git' is not installed or not on PATH. "
            "Install: https://git-scm.com/downloads"
        )

    # Single shared client and state object -- both are thread-safe.
    client = GitHubClient(
        auth=config.auth,
        api_base=config.github_api_url,
        max_retries=_DEFAULT_API_RETRIES,
    )
    _preflight_checks(config, client)
    gh_actor = _get_github_actor(client, config.auth)
    if gh_actor:
        log.info(f"Authenticated as: {gh_actor}")

    # Pre-fetch property schemas for orgs that have custom properties to inject at creation.
    # Done once here (before threads launch) so _mirror_repo workers only do dict lookups.
    if config.repo_custom_properties:
        orgs_with_props = {
            key.split("/", 1)[0] for key in config.repo_custom_properties
        }
        for org in sorted(orgs_with_props):
            config._org_property_schemas[org] = _fetch_org_property_schema(client, org)

    state = MigrationState(_CHECKPOINT_FILE)

    # Split repos into already-done (skip) and pending (process).
    already_done: list[RepoSpec] = []
    pending: list[RepoSpec] = []
    for spec in config.repos:
        key = f"{spec.namespace}/{spec.project}"
        (already_done if state.is_succeeded(key) else pending).append(spec)

    if batch_size > 0 and len(pending) > batch_size:
        log.info(
            f"Batch mode: capping this run at {batch_size} repo(s) "
            f"({len(pending) - batch_size} remaining will be processed in future runs)."
        )
        pending = pending[:batch_size]

    skipped_results: list[_RepoResult] = [
        _RepoResult(
            source=f"{s.namespace}/{s.project}",
            target=f"{s.target_org}/{s.target_name}",
            status="skipped",
            default_branch="",
            error="",
            duration_seconds=0.0,
            completed_at=_fmt_ts(),
        )
        for s in already_done
    ]

    if not _print_migration_preview(config, pending, already_done, config_path, repos_csv_path):
        log.info("Migration cancelled at pre-flight preview.")
        return

    total = len(pending)
    if total == 0:
        log.info("All repositories have already been migrated. Nothing to do.")
        _write_reports(skipped_results, _fmt_ts(), config, elapsed_seconds=0.0, executed_by=gh_actor)
        return

    run_started = _fmt_ts()
    wall_start  = time.monotonic()
    results: list[_RepoResult] = list(skipped_results)
    completed_count = 0

    with ThreadPoolExecutor(
        max_workers=config.max_workers,
        thread_name_prefix="gitmirror",
    ) as pool:
        all_futures: dict[Future, RepoSpec] = {
            pool.submit(_migrate_one, spec, config, client, state): spec
            for spec in pending
        }

        for future in as_completed(all_futures):
            spec = all_futures[future]
            try:
                result = future.result()
            except Exception as exc:
                # Defensive catch -- _migrate_one should never raise, but just in case.
                key = f"{spec.namespace}/{spec.project}"
                state.record(key, "failed")
                result = _RepoResult(
                    source=key,
                    target=f"{spec.target_org}/{spec.target_name}",
                    status="failed",
                    default_branch="",
                    error=str(exc),
                    duration_seconds=0.0,
                    completed_at=_fmt_ts(),
                )

            results.append(result)
            completed_count += 1
            pct = completed_count / total * 100
            elapsed = time.monotonic() - wall_start
            eta_secs = int((total - completed_count) / (completed_count / elapsed)) if completed_count < total and elapsed > 0 else 0
            eta_str = f" | ETA {eta_secs // 60}m {eta_secs % 60:02d}s" if eta_secs else ""
            status_icon = (
                "\u2705" if result.status == "succeeded" else
                "\u26a0\ufe0f" if result.status == "partial" else
                "\u274c"
            )
            _counter = _c(f"[{completed_count}/{total} {pct:.1f}%{eta_str}]", 2, 37)
            _src     = _c(result.source, 1, 36)   # bold cyan
            _tgt     = _c(result.target, 1, 96)   # bold bright-cyan
            _arrow   = _c("\u2192", 2)             # dim arrow
            _branch  = (
                f" \U0001f33f {_c(result.default_branch, 92)}"
                if result.default_branch else ""
            )
            _stats_parts: list[str] = []
            if result.branch_count:
                _stats_parts.append(f"{result.branch_count}b")
            if result.tag_count:
                _stats_parts.append(f"{result.tag_count}t")
            if result.commit_count:
                _stats_parts.append(f"{result.commit_count:,}c")
            _stats = _c(" [" + " ".join(_stats_parts) + "]", 2, 36) if _stats_parts else ""
            log.info(f"{status_icon} {_counter} {_src} {_arrow} {_tgt}{_branch}{_stats}")
            if result.status == "failed":
                log.error(f"  \u26a0\ufe0f  {result.error}")
            elif result.status == "partial":
                # Show exactly which sub-steps passed and which failed.
                def _step_tag(label: str, ok: bool, detail: str = "") -> str:
                    icon   = _c("\u2705", 92) if ok else _c("\u274c", 91)
                    suffix = _c(f" ({detail[:70]})", 2) if detail and not ok else ""
                    return f"{icon} {label}{suffix}"
                _cp_ok  = result.custom_properties_status in ("applied", "skipped", "not_configured", "dry-run")
                _ci_ok  = result.ci_skeleton_status not in ("failed",)
                _git_tag = _step_tag("git", True)
                _cp_tag  = _step_tag("custom-props",  _cp_ok,  result.custom_properties_error)
                _ci_tag  = _step_tag(".github",       _ci_ok,  result.ci_skeleton_error)
                log.warning(f"  \u26a0\ufe0f  Partial  {_git_tag}  {_cp_tag}  {_ci_tag}")

            # Honour shutdown request: cancel any futures that haven't started yet.
            if _shutdown_event.is_set():
                cancelled = sum(1 for f in all_futures if f.cancel())
                log.warning(
                    f"Shutdown: cancelled {cancelled} pending task(s). "
                    "Run the script again to resume from checkpoint."
                )
                break

    # Final summary
    succeeded = [r for r in results if r.status == "succeeded"]
    partial   = [r for r in results if r.status == "partial"]
    failed    = [r for r in results if r.status == "failed"]
    skipped   = [r for r in results if r.status == "skipped"]
    elapsed   = time.monotonic() - wall_start

    processed = len(succeeded) + len(partial) + len(failed)
    success_rate = f"{len(succeeded)}/{processed} ({len(succeeded)/processed*100:.1f}%)" if processed else "N/A"

    _div1  = _c("\u2550" * 60, 2, 37)   # dim gray double-rule
    _div2  = _c("\u2500" * 60, 2, 37)   # dim gray single-rule
    log.info(_div1)
    log.info(_c("\U0001f3c1  Migration Run Complete", 1, 97))  # bold bright-white
    log.info(_div2)
    log.info(f"  \u2705  Succeeded  :  {_c(str(len(succeeded)), 1, 92)}")   # bold bright-green
    if partial:
        log.info(
            f"  \u26a0\ufe0f  Partial    :  {_c(str(len(partial)), 1, 93)}  "
            + _c("(git ok; custom-props or CI skeleton failed)", 2)
        )  # bold bright-yellow count, dim explanation
    log.info(f"  \u274c  Failed     :  {_c(str(len(failed)),    1, 91)}")   # bold bright-red
    log.info(f"  \u23ed\ufe0f  Skipped    :  {_c(str(len(skipped)),  2)}  {_c('(already done)', 2)}")  # dim
    log.info(f"  \U0001f4ca  Success Rate:  {_c(success_rate, 1, 97)}")     # bold bright-white
    log.info(f"  \u23f1\ufe0f  Elapsed    :  {_c(_fmt_duration(elapsed), 1)}")
    # Custom properties summary
    cp_applied  = sum(1 for r in results if r.custom_properties_status == "applied")
    cp_failed   = sum(1 for r in results if r.custom_properties_status == "failed")
    cp_skipped  = sum(1 for r in results if r.custom_properties_status == "skipped")
    cp_none     = sum(1 for r in results if r.custom_properties_status == "not_configured")
    if cp_applied or cp_failed:
        log.info(_div2)
        log.info(_c("  Custom Properties:", 1, 4))   # bold underline section header
        log.info(f"    Applied    : {_c(str(cp_applied), 92)}")   # bright-green
        if cp_failed:
            log.info(f"    Failed     : {_c(str(cp_failed), 91)}")  # bright-red
        if cp_skipped:
            log.info(f"    Skipped    : {_c(str(cp_skipped), 2)}  {_c('(not in repo-properties.csv)', 2)}")
        if cp_none:
            log.info(f"    N/A        : {_c(str(cp_none), 2)}  {_c('(no CSV configured)', 2)}")
    # CI skeleton (.github) summary — shown whenever CI skeleton is configured.
    ci_created  = sum(1 for r in results if r.ci_skeleton_status == "created")
    ci_partial  = sum(1 for r in results if r.ci_skeleton_status == "partial")
    ci_skipped  = sum(1 for r in results if r.ci_skeleton_status == "skipped_all")
    ci_failed   = sum(1 for r in results if r.ci_skeleton_status == "failed")
    ci_not_cfg  = sum(1 for r in results if r.ci_skeleton_status == "not_configured")
    if config.ci_skeleton is not None or ci_created or ci_failed or ci_partial:
        log.info(_div2)
        log.info(_c("  .github Folder (CI Skeleton):", 1, 4))  # bold underline section header
        if ci_created:
            log.info(f"    Created    : {_c(str(ci_created), 92)}")   # bright-green
        if ci_partial:
            log.info(f"    Partial    : {_c(str(ci_partial), 93)}")   # bright-yellow
        if ci_skipped:
            log.info(f"    Skipped    : {_c(str(ci_skipped), 2)}  {_c('(file already exists on branch)', 2)}")
        if ci_failed:
            log.info(f"    Failed     : {_c(str(ci_failed), 91)}")    # bright-red
        if ci_not_cfg:
            log.info(f"    N/A        : {_c(str(ci_not_cfg), 2)}  {_c('(ci_skeleton not configured)', 2)}")
    log.info(_div1)
    if failed:
        log.warning(f"  \u26a0\ufe0f  Failed repos: {_c(', '.join(r.target for r in failed), 1, 91)}")
    if partial:
        log.warning(f"  \u26a0\ufe0f  Partial repos: {_c(', '.join(r.target for r in partial), 1, 93)}")
    if failed or partial:
        log.info("  \U0001f504  Re-run the script to retry failed/partial sub-steps automatically.")

    _write_reports(results, run_started, config, elapsed_seconds=elapsed, executed_by=gh_actor)


# ===========================================================================
# Post-migration: apply GitHub custom repository properties
# ===========================================================================

_REPO_PROPERTIES_CSV_FILENAME = "repo-properties.csv"


def _load_repo_properties_csv(csv_path: Path) -> dict[str, dict[str, str]]:
    """Read repo-properties.csv and return a mapping of 'org/repo' -> {prop: val}.

    Expected CSV format (header row required):
        target_org,target_name,<property1>,<property2>,...

    Rows where both target_org and target_name are non-blank are included.
    """
    if not csv_path.exists():
        raise FileNotFoundError(
            f"Repo properties CSV not found: {csv_path}\n"
            "Create it with columns: target_org,target_name,<prop1>,<prop2>,..."
        )

    result: dict[str, dict[str, str]] = {}
    with csv_path.open(encoding="utf-8", newline="") as fh:
        reader = csv.DictReader(fh)
        if reader.fieldnames is None:
            raise ValueError(f"repo-properties.csv appears to be empty: {csv_path}")
        required = {"target_org", "target_name"}
        missing_cols = required - {c.strip() for c in reader.fieldnames if c}
        if missing_cols:
            raise ValueError(
                f"repo-properties.csv is missing required columns: {sorted(missing_cols)}"
            )
        prop_cols = [
            c.strip() for c in reader.fieldnames
            if c and c.strip() not in ("target_org", "target_name")
        ]
        if not prop_cols:
            raise ValueError(
                "repo-properties.csv has no property columns besides target_org/target_name."
            )
        for lineno, row in enumerate(reader, 2):
            org = (row.get("target_org") or "").strip()
            name = (row.get("target_name") or "").strip()
            if not org or not name:
                log.debug(f"repo-properties.csv line {lineno}: skipping blank org/name")
                continue
            key = f"{org}/{name}"
            props = {p: (row.get(p) or "").strip() for p in prop_cols}
            # Keep the null sentinel (__null__) to allow intentional clearing of a property.
            # Drop plain empty strings -- no value means "leave as-is", not "clear".
            props = {k: v for k, v in props.items() if v}
            if props:
                result[key] = props
    return result


def _fetch_org_property_schema(client: GitHubClient, org: str) -> dict[str, str]:
    """Return {property_name: value_type} for all properties in the org schema.

    Returns {} when the schema cannot be fetched (insufficient permissions, network error, etc.)
    so callers degrade gracefully rather than blocking the migration.
    """
    try:
        status, body = client.request("GET", f"/orgs/{org}/properties/schema")
        if status != 200:
            log.debug(f"Could not fetch property schema for org '{org}' (HTTP {status})")
            return {}
        entries: list[dict] = body if isinstance(body, list) else []
        return {
            e["property_name"]: e.get("value_type", "string")
            for e in entries
            if "property_name" in e
        }
    except Exception as exc:
        log.debug(f"Property schema fetch for org '{org}' failed: {exc}")
        return {}


_GITHUB_PROPERTY_TYPES = frozenset({"string", "single_select", "multi_select", "true_false", "url"})

# Sentinel string in CSV that explicitly clears a property (sets API value to null).
_PROP_NULL_SENTINEL = "__null__"


def _coerce_property_value(
    raw: str,
    value_type: str,
    prop_name: str = "",
) -> "str | list[str] | None":
    """Convert a raw CSV cell to the exact value type the GitHub API expects.

    GitHub custom property value types and their wire format:
      string        -> str  (any text; no further coercion)
      single_select -> str  (must match one of the schema's allowed_values; passed as-is)
      true_false    -> str  "true" | "false"  (normalised to lowercase)
      url           -> str  (validated to start with http:// or https://)
      multi_select  -> list[str]  (semicolon-separated in CSV: "v1;v2" -> ["v1", "v2"])

    Clearing a property:
      Use the sentinel value "__null__" in the CSV cell to send null to the API,
      which removes / unsets the property from the repository.

    Returns None for the null sentinel so the caller can emit {"value": null}.
    Logs a warning for unrecognised value_type and falls back to plain string.
    """
    if raw == _PROP_NULL_SENTINEL:
        return None

    if value_type == "multi_select":
        return [v.strip() for v in raw.split(";") if v.strip()]

    if value_type == "true_false":
        normalised = raw.strip().lower()
        if normalised not in ("true", "false"):
            log.warning(
                f"Property '{prop_name}' (true_false): invalid value {raw!r}. "
                "Expected 'true' or 'false'. Sending as-is; GitHub may reject with HTTP 422."
            )
            return raw
        return normalised

    if value_type == "url":
        stripped = raw.strip()
        if not (stripped.startswith("http://") or stripped.startswith("https://")):
            log.warning(
                f"Property '{prop_name}' (url): value {raw!r} does not start with "
                "http:// or https://. GitHub may reject with HTTP 422."
            )
        return stripped

    if value_type not in _GITHUB_PROPERTY_TYPES:
        log.debug(
            f"Property '{prop_name}': unknown value_type {value_type!r}. "
            "Treating as string."
        )

    # string, single_select, and unknown types: pass through as-is.
    return raw


def _validate_org_property_schema(
    client: GitHubClient,
    org: str,
    property_names: set[str],
) -> list[str]:
    """Return property names not defined in the org schema, or [] if validation cannot run."""
    schema_map = _fetch_org_property_schema(client, org)
    if not schema_map:
        log.warning(
            f"[post-migration] Could not read property schema for org '{org}'. "
            "Skipping validation -- GitHub will reject unknown property names."
        )
        return []
    return sorted(property_names - set(schema_map))


def _set_repo_properties(
    client: GitHubClient,
    org: str,
    repo: str,
    properties: dict[str, str],
    label: str,
    schema_map: "dict[str, str] | None" = None,
) -> tuple[bool, str]:
    """PATCH /repos/{org}/{repo}/properties/values. Returns (ok, error_message).

    *schema_map* is {property_name: value_type}. When not supplied it is fetched
    from the org schema so that values are coerced to the correct wire type.
    Supports all GitHub property types: string, single_select, multi_select,
    true_false, url. Use the CSV sentinel '__null__' to clear a property (null).
    """
    if not properties:
        return True, ""
    if schema_map is None:
        schema_map = _fetch_org_property_schema(client, org)
    coerced = {
        k: _coerce_property_value(v, schema_map.get(k, "string"), prop_name=k)
        for k, v in properties.items()
    }
    payload = {
        "properties": [
            {"property_name": k, "value": v}  # None serialises to JSON null -- intentional
            for k, v in coerced.items()
        ]
    }
    try:
        status, body = client.request("PATCH", f"/repos/{org}/{repo}/properties/values", body=payload)
        if status in (200, 204):
            log.info(
                f"[{label}] Custom properties set: "
                + ", ".join(f"{k}={v!r}" for k, v in properties.items())
            )
            return True, ""
        raw_msg = body.get('message', str(body))[:200]
        if status == 403:
            err = (
                f"HTTP 403: {raw_msg} -- "
                "Token lacks permission to write org custom properties. "
                "For a classic PAT: ensure the 'admin:org' scope is granted. "
                "For a fine-grained PAT: enable 'Custom properties' (read/write) "
                "at the organization level. "
                "For a GitHub App: add the 'custom_properties:write' permission."
            )
        else:
            err = f"HTTP {status}: {raw_msg}"
        log.warning(f"[{label}] Failed to set custom properties -- {err}")
        return False, err
    except Exception as exc:
        err = str(exc)
        log.warning(f"[{label}] Exception setting custom properties -- {err}")
        return False, err


def run_post_migration(
    config: MigrationConfig,
    config_path: Path,
    repos_csv_path: Path,
    properties_csv_path: Path,
) -> None:
    """Apply GitHub custom repository properties to previously migrated repos.

    Reads the checkpoint file to find repos with 'succeeded' status, then
    applies custom properties from *properties_csv_path* for each matched repo.
    A separate post-migration report (JSON) is written when complete.
    """
    log.info("=" * 60)
    log.info("POST-MIGRATION: Applying custom repository properties")
    log.info("=" * 60)

    # ---- Load properties CSV -------------------------------------------
    try:
        props_map = _load_repo_properties_csv(properties_csv_path)
    except (FileNotFoundError, ValueError) as exc:
        log.error(f"[post-migration] Cannot load properties CSV: {exc}")
        sys.exit(1)

    if not props_map:
        log.warning("[post-migration] repo-properties.csv has no usable rows. Nothing to do.")
        return

    log.info(f"[post-migration] Loaded {len(props_map)} repo property mappings from {properties_csv_path.name}")

    # ---- Validate property names against org schema --------------------
    client = GitHubClient(
        auth=config.auth,
        api_base=config.github_api_url,
        max_retries=_DEFAULT_API_RETRIES,
    )

    # Fetch schema per org: used for both validation and value coercion (multi_select needs arrays).
    orgs_in_csv: set[str] = {key.split("/", 1)[0] for key in props_map}
    all_prop_names: set[str] = {p for props in props_map.values() for p in props}
    org_schemas: dict[str, dict[str, str]] = {}

    for org in sorted(orgs_in_csv):
        schema = _fetch_org_property_schema(client, org)
        org_schemas[org] = schema
        if schema:
            undefined = sorted(all_prop_names - set(schema))
            if undefined:
                log.warning(
                    f"[post-migration] Org '{org}': property name(s) not in schema -- {undefined}. "
                    "Define them in GitHub → Org Settings → Custom Properties before proceeding."
                )
        else:
            log.warning(
                f"[post-migration] Could not read property schema for org '{org}'. "
                "Value coercion will default to string; multi_select may fail."
            )

    # ---- Load checkpoint to find succeeded repos -----------------------
    state = MigrationState(_CHECKPOINT_FILE)
    succeeded_keys = {k for k, v in state._state.items() if v == "succeeded"}

    if not succeeded_keys:
        log.warning("[post-migration] No repos with 'succeeded' status in checkpoint. Nothing to do.")
        return

    log.info(f"[post-migration] {len(succeeded_keys)} succeeded repo(s) in checkpoint")

    # ---- Apply properties ----------------------------------------------
    applied: list[dict] = []
    skipped: list[dict] = []
    failed: list[dict] = []
    t_start = time.monotonic()

    for key in sorted(succeeded_keys):
        # key format: "namespace/project"
        # We need the target org/name -- read from repos.csv mapping via config
        spec = next(
            (r for r in config.repos if f"{r.namespace}/{r.project}" == key),
            None,
        )
        if spec is None:
            log.debug(f"[post-migration] No RepoSpec for checkpoint key '{key}' -- skipping")
            skipped.append({"source": key, "target": "", "reason": "not in repos.csv"})
            continue

        target_key = f"{spec.target_org}/{spec.target_name}"
        props = props_map.get(target_key)
        if not props:
            log.debug(f"[post-migration] No properties defined for '{target_key}' -- skipping")
            skipped.append({"source": key, "target": target_key, "reason": "not in repo-properties.csv"})
            continue

        ok, err = _set_repo_properties(
            client, spec.target_org, spec.target_name, props, target_key,
            schema_map=org_schemas.get(spec.target_org),
        )
        if ok:
            applied.append({"source": key, "target": target_key, "properties": props})
        else:
            failed.append({"source": key, "target": target_key, "properties": props, "error": err})

    elapsed = time.monotonic() - t_start

    # ---- Summary -------------------------------------------------------
    log.info("=" * 60)
    log.info(f"[post-migration] Applied : {len(applied)} repo(s)")
    log.info(f"[post-migration] Skipped : {len(skipped)} repo(s) (no properties configured)")
    log.info(f"[post-migration] Failed  : {len(failed)} repo(s)")
    log.info(f"[post-migration] Elapsed : {_fmt_duration(elapsed)}")
    log.info("=" * 60)

    # ---- Write post-migration report -----------------------------------
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_path = _SCRIPT_DIR / "reports" / f"post-migration-properties-{ts}.json"
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_data = {
        "run_type": "post_migration_properties",
        "completed_at": _fmt_ts(),
        "elapsed_seconds": round(elapsed, 1),
        "summary": {
            "applied": len(applied),
            "skipped": len(skipped),
            "failed": len(failed),
        },
        "applied": applied,
        "skipped": skipped,
        "failed": failed,
    }
    report_path.write_text(json.dumps(report_data, indent=2), encoding="utf-8")
    log.info(f"[post-migration] Report written: {report_path}")

    if failed:
        log.warning(
            f"[post-migration] {len(failed)} repo(s) failed. "
            f"Review {report_path.name} and retry."
        )
        sys.exit(1)


# ===========================================================================
# Entry point
# ===========================================================================

if __name__ == "__main__":
    import argparse as _argparse

    _parser = _argparse.ArgumentParser(
        description="GitLab \u2192 GitHub Mirror Migration",
        formatter_class=_argparse.RawDescriptionHelpFormatter,
    )
    _parser.add_argument(
        "--batch", type=int, default=0, metavar="N",
        help=(
            "Process at most N pending repositories in this run "
            "(0 = all pending). Use this to control blast radius "
            "when migrating in multiple runs (e.g. --batch 500)."
        ),
    )
    _parser.add_argument(
        "--post-migration", action="store_true",
        help=(
            "Run post-migration step: apply GitHub custom repository properties "
            "from repo-properties.csv to all previously succeeded repos. "
            "Does NOT re-run the migration."
        ),
    )
    _args = _parser.parse_args()

    config_path = _SCRIPT_DIR / _CONFIG_FILENAME
    repos_csv_path = _SCRIPT_DIR / _REPOS_CSV_FILENAME
    properties_csv_path = _SCRIPT_DIR / _REPO_PROPERTIES_CSV_FILENAME

    try:
        migration_config = load_config(config_path, repos_csv_path)

        if _args.post_migration:
            run_post_migration(migration_config, config_path, repos_csv_path, properties_csv_path)
        else:
            migrate_all(migration_config, config_path, repos_csv_path, batch_size=_args.batch)

    except _MissingConfigError as exc:
        log.warning(str(exc))
        sys.exit(1)
    except (ValueError, RuntimeError, FileNotFoundError, KeyError) as exc:
        log.error(str(exc))
        sys.exit(1)
    except Exception as exc:
        log.error(f"Unexpected error: {exc}")
        sys.exit(1)
