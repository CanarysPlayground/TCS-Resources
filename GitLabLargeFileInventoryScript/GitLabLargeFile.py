import subprocess
import importlib
import importlib.util

_REQUIRED_PACKAGES = {
    "requests": "requests>=2.31.0",
    "openpyxl": "openpyxl>=3.1.2",
}

def _ensure_packages() -> None:
    """Install any missing third-party packages via pip at startup."""
    missing = [
        pip_spec
        for import_name, pip_spec in _REQUIRED_PACKAGES.items()
        if importlib.util.find_spec(import_name) is None
    ]
    if not missing:
        return

    print(f"[Bootstrap] Installing missing packages: {', '.join(missing)}")
    result = subprocess.run(
        [subprocess.sys.executable, "-m", "pip", "install", *missing],
        check=False,
    )
    if result.returncode != 0:
        print(
            "[Bootstrap] ERROR: Could not install required packages.\n"
            f"Please run manually:  pip install {' '.join(missing)}",
            file=subprocess.sys.stderr,
        )
        raise SystemExit(1)

    # Invalidate importlib caches so newly installed packages are importable
    importlib.invalidate_caches()
    print("[Bootstrap] Packages installed successfully.\n")

_ensure_packages()
# ---------------------------------------------------------------------------

import base64
import csv
import json
import logging
import sys
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from logging.handlers import RotatingFileHandler
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import requests
import requests.adapters

# ---------------------------------------------------------------------------
# Optional dependency: openpyxl (for Excel output)
# ---------------------------------------------------------------------------
try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
SCRIPT_DIR = Path(__file__).parent.resolve()
CONFIG_FILE = SCRIPT_DIR / "config.json"
REPOS_FILE = SCRIPT_DIR / "repos.csv"

REQUIRED_CONFIG_KEYS = ["gitlab_url", "private_token"]

# Output column headers
REPORT_HEADERS = [
    "Org / Group Name",
    "Repository Name",
    "Repository Path",
    "Branch Name",
    "File Path",
    "Folder Path",
    "File Name",
    "File Size (MB)",
    "File Size (Bytes)",
    "Last Commit ID",
    "File URL",
    "Scanned At",
]

# Files smaller than this are checked for LFS pointer content via GET
_LFS_CANDIDATE_THRESHOLD = 1024

# ---------------------------------------------------------------------------
# Logging setup
# ---------------------------------------------------------------------------

def setup_logging(log_level: str, log_file: str) -> logging.Logger:
    """Configure console + rotating file logging."""
    logger = logging.getLogger("gitlab_scanner")
    logger.setLevel(getattr(logging, log_level.upper(), logging.INFO))

    fmt = logging.Formatter(
        "%(asctime)s | %(levelname)-8s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    # Console handler
    ch = logging.StreamHandler(sys.stdout)
    ch.setFormatter(fmt)
    logger.addHandler(ch)

    # Rotating file handler (10 MB × 5 backups)
    log_path = SCRIPT_DIR / log_file
    fh = RotatingFileHandler(log_path, maxBytes=10 * 1024 * 1024, backupCount=5, encoding="utf-8")
    fh.setFormatter(fmt)
    logger.addHandler(fh)

    return logger


# ---------------------------------------------------------------------------
# Config & repos loading
# ---------------------------------------------------------------------------

def load_config(path: Path) -> Dict[str, Any]:
    """Load and validate config.json."""
    if not path.exists():
        raise FileNotFoundError(f"Config file not found: {path}")

    with open(path, "r", encoding="utf-8") as f:
        cfg = json.load(f)

    missing = [k for k in REQUIRED_CONFIG_KEYS if not cfg.get(k)]
    if missing:
        raise ValueError(f"Missing required config keys: {missing}")

    if cfg.get("private_token", "").startswith("YOUR_"):
        raise ValueError(
            "Please replace 'YOUR_GITLAB_PRIVATE_TOKEN_HERE' in config.json with a real token."
        )

    # Apply defaults
    cfg.setdefault("size_threshold_mb", 100)
    cfg.setdefault("output_format", "excel")
    cfg.setdefault("output_file", "large_files_report")
    cfg.setdefault("branches_to_scan", [])
    cfg.setdefault("scan_all_branches", False)
    cfg.setdefault("default_branch_only", True)
    cfg.setdefault("request_timeout_seconds", 30)
    cfg.setdefault("max_retries", 3)
    cfg.setdefault("retry_delay_seconds", 2)
    cfg.setdefault("log_level", "INFO")
    cfg.setdefault("log_file", "gitlab_scanner.log")
    cfg.setdefault("concurrent_workers", 2)           # repos in parallel
    cfg.setdefault("branch_workers", 5)               # branches per repo in parallel
    cfg.setdefault("file_workers", 16)                # file probes per branch in parallel
    cfg.setdefault("max_concurrent_requests", 20)     # total in-flight HTTP calls
    cfg.setdefault("per_page", 100)

    return cfg


def load_repos(path: Path) -> List[Dict[str, str]]:
    """Load repos.csv. Returns list of dicts with 'org_name' and 'repo_path'."""
    if not path.exists():
        raise FileNotFoundError(f"Repos file not found: {path}")

    repos: List[Dict[str, str]] = []
    with open(path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames is None:
            raise ValueError("repos.csv is empty or has no headers.")

        headers_lower = [h.strip().lower() for h in reader.fieldnames]
        if "repo_path" not in headers_lower:
            raise ValueError(
                "repos.csv must contain a 'repo_path' column "
                "(GitLab namespace/project, e.g. my-group/my-project)."
            )

        for row_num, row in enumerate(reader, start=2):
            # Normalise keys
            normalised = {k.strip().lower(): (v or "").strip() for k, v in row.items()}
            repo_path = normalised.get("repo_path", "")
            org_name = normalised.get("org_name", repo_path.split("/")[0] if "/" in repo_path else repo_path)

            if not repo_path:
                continue  # skip blank rows

            repos.append(
                {
                    "org_name": org_name,
                    "repo_path": repo_path,
                }
            )

    return repos


# ---------------------------------------------------------------------------
# GitLab API client
# ---------------------------------------------------------------------------

class GitLabClient:
    """
    Thread-safe GitLab REST API v4 client.

    Key optimisations:
    - HEAD requests for file size: reads X-Gitlab-Size + X-Gitlab-Last-Commit-Id
      headers without downloading base64 content.
    - Blob SHA cache: identical blobs on different branches are probed once.
    - Semaphore caps total concurrent in-flight HTTP calls.
    - Properly-sized urllib3 connection pool.
    - Exponential back-off with semaphore released during sleep.
    """

    def __init__(self, cfg: Dict[str, Any], logger: logging.Logger) -> None:
        self.base_url = cfg["gitlab_url"].rstrip("/") + "/api/v4"
        self.timeout = cfg["request_timeout_seconds"]
        self.max_retries = cfg["max_retries"]
        self.retry_delay = cfg["retry_delay_seconds"]
        self.per_page = cfg["per_page"]
        self.logger = logger

        max_conn = cfg.get("max_concurrent_requests", 20)
        self.session = requests.Session()
        adapter = requests.adapters.HTTPAdapter(
            pool_connections=max_conn,
            pool_maxsize=max_conn * 2,
            max_retries=0,
        )
        self.session.mount("https://", adapter)
        self.session.mount("http://", adapter)
        self.session.headers.update({"PRIVATE-TOKEN": cfg["private_token"]})

        # Caps total concurrent in-flight HTTP calls across all worker threads.
        self._semaphore = threading.Semaphore(max_conn)

        # blob_sha -> size_bytes; prevents re-fetching the same blob on multiple branches.
        self._blob_size_cache: Dict[str, int] = {}
        self._cache_lock = threading.Lock()

    # ------------------------------------------------------------------
    # Core HTTP execution
    # ------------------------------------------------------------------

    def _execute(self, method: str, url: str, params: Optional[Dict] = None) -> requests.Response:
        """Single HTTP call with semaphore + exponential back-off retry."""
        params = params or {}
        for attempt in range(1, self.max_retries + 1):
            try:
                with self._semaphore:
                    resp = self.session.request(method, url, params=params, timeout=self.timeout)
            except (requests.exceptions.ConnectionError, requests.exceptions.Timeout) as exc:
                if attempt < self.max_retries:
                    delay = self.retry_delay * (2 ** (attempt - 1))
                    self.logger.warning(
                        "Network error attempt %d/%d for %s: %s — retry in %.0fs.",
                        attempt, self.max_retries, url, exc, delay,
                    )
                    time.sleep(delay)
                    continue
                raise

            if resp.status_code == 429:
                if attempt < self.max_retries:
                    wait = int(resp.headers.get("Retry-After", self.retry_delay * (2 ** attempt)))
                    self.logger.warning(
                        "Rate limited (attempt %d/%d). Waiting %ds.", attempt, self.max_retries, wait
                    )
                    time.sleep(wait)
                    continue
                resp.raise_for_status()

            return resp

        raise RuntimeError(f"Exhausted {self.max_retries} retries for {url}")

    # ------------------------------------------------------------------
    # Public request helpers
    # ------------------------------------------------------------------

    def _get(self, endpoint: str, params: Optional[Dict] = None) -> Any:
        resp = self._execute("GET", f"{self.base_url}{endpoint}", params)
        resp.raise_for_status()
        return resp.json()

    def _get_paginated(self, endpoint: str, params: Optional[Dict] = None) -> List[Any]:
        params = dict(params or {})
        params["per_page"] = self.per_page
        url = f"{self.base_url}{endpoint}"
        items: List[Any] = []
        page = 1

        while True:
            params["page"] = page
            resp = self._execute("GET", url, params)
            resp.raise_for_status()
            batch: List[Any] = resp.json()
            if not batch:
                break
            items.extend(batch)
            if page >= int(resp.headers.get("X-Total-Pages", 1)):
                break
            page += 1

        return items

    # ------------------------------------------------------------------
    # Domain helpers
    # ------------------------------------------------------------------

    def get_project(self, repo_path: str) -> Optional[Dict]:
        encoded = requests.utils.quote(repo_path, safe="")
        try:
            return self._get(f"/projects/{encoded}")
        except requests.exceptions.HTTPError as exc:
            if exc.response is not None and exc.response.status_code == 404:
                return None
            raise

    def get_branches(self, project_id: int) -> List[str]:
        return [b["name"] for b in self._get_paginated(f"/projects/{project_id}/repository/branches")]

    def get_blobs(self, project_id: int, branch: str) -> List[Dict]:
        tree = self._get_paginated(
            f"/projects/{project_id}/repository/tree",
            params={"ref": branch, "recursive": True},
        )
        return [item for item in tree if item.get("type") == "blob"]

    # ------------------------------------------------------------------
    # File size probing (hot path)
    # ------------------------------------------------------------------

    def probe_file_size(
        self,
        project_id: int,
        file_path: str,
        branch: str,
        blob_sha: str,
        threshold_bytes: int,
    ) -> Optional[Tuple[int, str]]:
        """
        Return (size_bytes, last_commit_id) if size >= threshold, else None.

        Strategy (fastest to slowest):
          1. Blob SHA cache hit  — zero API calls.
          2. HEAD /repository/files — reads X-Gitlab-Size + X-Gitlab-Last-Commit-Id
             headers without downloading any file content.
          3. Full GET — only for files < 1 KB (potential LFS pointer) or when
             HEAD is unavailable/fails on the server.
        """
        # 1. Cache check (blob SHA is content-addressed; same SHA == same size).
        if blob_sha:
            with self._cache_lock:
                cached = self._blob_size_cache.get(blob_sha)
            if cached is not None:
                if cached < threshold_bytes:
                    return None
                # Large blob in cache; still need per-branch commit ID via HEAD.
                commit_id = self._head_commit_id(project_id, file_path, branch)
                return cached, commit_id

        # 2. HEAD request — no content download.
        encoded = requests.utils.quote(file_path, safe="")
        url = f"{self.base_url}/projects/{project_id}/repository/files/{encoded}"
        try:
            resp = self._execute("HEAD", url, {"ref": branch})
        except requests.exceptions.RequestException as exc:
            self.logger.debug("HEAD failed for %s@%s (%s); falling back to GET.", file_path, branch, exc)
            return self._probe_via_get(project_id, file_path, branch, blob_sha, threshold_bytes)

        if resp.status_code == 404:
            return None
        if resp.status_code in (405, 501):
            return self._probe_via_get(project_id, file_path, branch, blob_sha, threshold_bytes)
        try:
            resp.raise_for_status()
        except requests.exceptions.HTTPError:
            return self._probe_via_get(project_id, file_path, branch, blob_sha, threshold_bytes)

        size_str = resp.headers.get("X-Gitlab-Size", "")
        if not size_str:
            return self._probe_via_get(project_id, file_path, branch, blob_sha, threshold_bytes)

        size_bytes = int(size_str)
        last_commit = (
            resp.headers.get("X-Gitlab-Last-Commit-Id")
            or resp.headers.get("X-Gitlab-Commit-Id")
            or ""
        )

        if blob_sha:
            with self._cache_lock:
                self._blob_size_cache[blob_sha] = size_bytes

        # 3. LFS pointer candidate: git object is tiny but may reference a huge object.
        if size_bytes < _LFS_CANDIDATE_THRESHOLD:
            return self._resolve_lfs(project_id, file_path, branch, blob_sha, threshold_bytes, last_commit)

        return (size_bytes, last_commit) if size_bytes >= threshold_bytes else None

    def _head_commit_id(self, project_id: int, file_path: str, branch: str) -> str:
        """Lightweight HEAD to fetch last_commit_id for a blob already in cache."""
        encoded = requests.utils.quote(file_path, safe="")
        url = f"{self.base_url}/projects/{project_id}/repository/files/{encoded}"
        try:
            resp = self._execute("HEAD", url, {"ref": branch})
            if resp.ok:
                return (
                    resp.headers.get("X-Gitlab-Last-Commit-Id")
                    or resp.headers.get("X-Gitlab-Commit-Id")
                    or ""
                )
        except Exception:
            pass
        return ""

    def _resolve_lfs(
        self,
        project_id: int,
        file_path: str,
        branch: str,
        blob_sha: str,
        threshold_bytes: int,
        fallback_commit: str,
    ) -> Optional[Tuple[int, str]]:
        """GET the file and parse the LFS pointer to obtain the actual object size."""
        meta = self._get_file_raw(project_id, file_path, branch)
        if meta is None:
            return None

        size_bytes: int = meta.get("size", 0)
        last_commit: str = meta.get("last_commit_id", fallback_commit)

        if size_bytes < _LFS_CANDIDATE_THRESHOLD and meta.get("encoding") == "base64":
            try:
                content = base64.b64decode(meta.get("content", "")).decode("utf-8", errors="replace")
                if content.startswith("version https://git-lfs.github.com/spec/"):
                    for line in content.splitlines():
                        if line.startswith("size "):
                            size_bytes = int(line.split(" ", 1)[1])
                            break
            except Exception:
                pass

        if blob_sha:
            with self._cache_lock:
                self._blob_size_cache[blob_sha] = size_bytes

        return (size_bytes, last_commit) if size_bytes >= threshold_bytes else None

    def _probe_via_get(
        self,
        project_id: int,
        file_path: str,
        branch: str,
        blob_sha: str,
        threshold_bytes: int,
    ) -> Optional[Tuple[int, str]]:
        """Full GET fallback: handles both regular files and LFS pointers."""
        meta = self._get_file_raw(project_id, file_path, branch)
        if meta is None:
            return None

        size_bytes: int = meta.get("size", 0)
        last_commit: str = meta.get("last_commit_id", "")

        if size_bytes < _LFS_CANDIDATE_THRESHOLD and meta.get("encoding") == "base64":
            try:
                content = base64.b64decode(meta.get("content", "")).decode("utf-8", errors="replace")
                if content.startswith("version https://git-lfs.github.com/spec/"):
                    for line in content.splitlines():
                        if line.startswith("size "):
                            size_bytes = int(line.split(" ", 1)[1])
                            break
            except Exception:
                pass

        if blob_sha:
            with self._cache_lock:
                self._blob_size_cache[blob_sha] = size_bytes

        return (size_bytes, last_commit) if size_bytes >= threshold_bytes else None

    def _get_file_raw(self, project_id: int, file_path: str, branch: str) -> Optional[Dict]:
        encoded = requests.utils.quote(file_path, safe="")
        try:
            return self._get(
                f"/projects/{project_id}/repository/files/{encoded}",
                params={"ref": branch},
            )
        except requests.exceptions.HTTPError as exc:
            if exc.response is not None and exc.response.status_code == 404:
                return None
            raise


# ---------------------------------------------------------------------------
# Scanner logic
# ---------------------------------------------------------------------------

def determine_branches(
    client: GitLabClient,
    project_id: int,
    default_branch: str,
    cfg: Dict[str, Any],
    logger: logging.Logger,
) -> List[str]:
    """Determine which branches to scan based on config."""
    if cfg.get("scan_all_branches"):
        branches = client.get_branches(project_id)
        logger.info("  Found %d branch(es) to scan (all branches mode).", len(branches))
        return branches

    if cfg.get("branches_to_scan"):
        return list(cfg["branches_to_scan"])

    # Default: only the project's default branch
    return [default_branch]


def scan_branch(
    client: GitLabClient,
    project_info: Dict,
    org_name: str,
    repo_path: str,
    branch: str,
    threshold_bytes: int,
    file_workers: int,
    scanned_at: str,
    logger: logging.Logger,
) -> List[Dict[str, Any]]:
    """Scan a single branch. File probing is parallelised across file_workers threads."""
    project_id: int = project_info["id"]
    repo_name: str = project_info["name"]
    web_url: str = project_info.get("web_url", "")

    logger.info("    Scanning branch: %s", branch)

    try:
        blobs = client.get_blobs(project_id, branch)
    except Exception as exc:
        logger.error("    Could not fetch tree for branch '%s': %s", branch, exc)
        return []

    total = len(blobs)
    logger.info("    Branch '%s' — %d file(s) in tree. Probing sizes …", branch, total)

    large_files: List[Dict[str, Any]] = []
    results_lock = threading.Lock()
    checked_count = [0]

    def probe(blob: Dict) -> None:
        file_path: str = blob.get("path", "")
        blob_sha: str = blob.get("id", "")
        if not file_path:
            return

        try:
            result = client.probe_file_size(project_id, file_path, branch, blob_sha, threshold_bytes)
        except Exception as exc:
            logger.debug("    Skipping %s@%s: %s", file_path, branch, exc)
            return

        with results_lock:
            checked_count[0] += 1
            done = checked_count[0]

        if done % 100 == 0 or done == total:
            logger.info("    [%s] Progress: %d/%d files checked.", branch, done, total)

        if result is None:
            return

        size_bytes, last_commit = result
        folder_path = str(Path(file_path).parent)
        if folder_path == ".":
            folder_path = "(root)"

        row: Dict[str, Any] = {
            "Org / Group Name": org_name,
            "Repository Name": repo_name,
            "Repository Path": repo_path,
            "Branch Name": branch,
            "File Path": file_path,
            "Folder Path": folder_path,
            "File Name": Path(file_path).name,
            "File Size (MB)": round(size_bytes / (1024 * 1024), 3),
            "File Size (Bytes)": size_bytes,
            "Last Commit ID": last_commit,
            "File URL": f"{web_url}/-/blob/{branch}/{file_path}",
            "Scanned At": scanned_at,
        }

        with results_lock:
            large_files.append(row)
            logger.info("    * Large file: %s  (%.2f MB)", file_path, row["File Size (MB)"])

    with ThreadPoolExecutor(
        max_workers=file_workers, thread_name_prefix=f"file-{branch[:20]}"
    ) as ex:
        list(ex.map(probe, blobs))

    logger.info(
        "    Branch '%s' done — %d large file(s) found.", branch, len(large_files)
    )
    return large_files


def scan_repo(
    client: GitLabClient,
    org_name: str,
    repo_path: str,
    cfg: Dict[str, Any],
    scanned_at: str,
    logger: logging.Logger,
) -> Tuple[str, List[Dict[str, Any]]]:
    """Scan all configured branches of one repo. Branches are scanned in parallel."""
    threshold_bytes = int(cfg["size_threshold_mb"] * 1024 * 1024)
    branch_workers: int = cfg.get("branch_workers", 5)
    file_workers: int = cfg.get("file_workers", 16)

    logger.info("Scanning repo: %s", repo_path)

    project = client.get_project(repo_path)
    if project is None:
        logger.warning("  Repo not found or no access: %s — skipping.", repo_path)
        return repo_path, []

    default_branch: str = project.get("default_branch") or "main"
    branches = determine_branches(client, project["id"], default_branch, cfg, logger)

    all_rows: List[Dict[str, Any]] = []
    rows_lock = threading.Lock()

    def scan_one(branch: str) -> None:
        rows = scan_branch(
            client, project, org_name, repo_path, branch,
            threshold_bytes, file_workers, scanned_at, logger,
        )
        with rows_lock:
            all_rows.extend(rows)

    with ThreadPoolExecutor(
        max_workers=min(branch_workers, len(branches)),
        thread_name_prefix="branch",
    ) as ex:
        futures = {ex.submit(scan_one, b): b for b in branches}
        for future in as_completed(futures):
            branch = futures[future]
            exc = future.exception()
            if exc:
                logger.error("  Branch '%s' raised: %s", branch, exc, exc_info=exc)

    logger.info("Finished repo: %s — total large files: %d", repo_path, len(all_rows))
    return repo_path, all_rows


# ---------------------------------------------------------------------------
# Output writers
# ---------------------------------------------------------------------------

def write_csv(rows: List[Dict[str, Any]], output_path: Path, logger: logging.Logger) -> None:
    """Write results to a CSV file."""
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=REPORT_HEADERS)
        writer.writeheader()
        writer.writerows(rows)
    logger.info("CSV report written: %s", output_path)


def write_excel(rows: List[Dict[str, Any]], output_path: Path, logger: logging.Logger) -> None:
    """Write results to a styled Excel (.xlsx) file."""
    if not OPENPYXL_AVAILABLE:
        logger.warning("openpyxl not installed. Falling back to CSV output.")
        csv_path = output_path.with_suffix(".csv")
        write_csv(rows, csv_path, logger)
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Large Files Report"

    # -- Header style
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=False)

    for col_idx, header in enumerate(REPORT_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    ws.row_dimensions[1].height = 22

    # -- Data rows
    even_fill = PatternFill("solid", fgColor="D6E4F0")

    for row_idx, data in enumerate(rows, start=2):
        fill = even_fill if row_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for col_idx, key in enumerate(REPORT_HEADERS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=data.get(key, ""))
            cell.fill = fill
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = left_align

    # -- Hyperlink the "File URL" column
    url_col = REPORT_HEADERS.index("File URL") + 1
    for row_idx in range(2, len(rows) + 2):
        cell = ws.cell(row=row_idx, column=url_col)
        url_val = cell.value
        if url_val:
            cell.hyperlink = url_val
            cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")

    # -- Auto-fit column widths (rough heuristic)
    for col_idx, header in enumerate(REPORT_HEADERS, start=1):
        max_len = len(header)
        for row_idx in range(2, len(rows) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, min(len(str(val)), 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 4

    # -- Freeze header row
    ws.freeze_panes = "A2"

    # -- Auto-filter
    ws.auto_filter.ref = ws.dimensions

    # -- Summary sheet
    ws_summary = wb.create_sheet(title="Summary")
    summary_data = [
        ("Report Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Total Large Files Found", len(rows)),
        ("Size Threshold", f"{rows[0].get('File Size (MB)', 'N/A')} MB threshold" if rows else "N/A"),
        ("Unique Repositories Scanned", len({r["Repository Path"] for r in rows})),
        ("Unique Branches Scanned", len({(r["Repository Path"], r["Branch Name"]) for r in rows})),
    ]
    for r_idx, (label, value) in enumerate(summary_data, start=1):
        ws_summary.cell(row=r_idx, column=1, value=label).font = Font(bold=True, name="Calibri")
        ws_summary.cell(row=r_idx, column=2, value=value).font = Font(name="Calibri")
    ws_summary.column_dimensions["A"].width = 35
    ws_summary.column_dimensions["B"].width = 40

    wb.save(output_path)
    logger.info("Excel report written: %s", output_path)


def write_report(rows: List[Dict[str, Any]], cfg: Dict[str, Any], logger: logging.Logger) -> None:
    """Dispatch to the correct writer based on config."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = f"{cfg['output_file']}_{timestamp}"
    output_format = cfg["output_format"].lower()

    if output_format in ("excel", "xlsx"):
        output_path = SCRIPT_DIR / f"{base_name}.xlsx"
        write_excel(rows, output_path, logger)
    else:
        output_path = SCRIPT_DIR / f"{base_name}.csv"
        write_csv(rows, output_path, logger)


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def main() -> None:
    # -- Load config (pre-logger so errors go to stderr)
    try:
        cfg = load_config(CONFIG_FILE)
    except (FileNotFoundError, ValueError) as exc:
        sys.exit(f"[ERROR] {exc}")

    logger = setup_logging(cfg["log_level"], cfg["log_file"])

    logger.info("=" * 70)
    logger.info("GitLab Large File Inventory Script")
    logger.info("=" * 70)
    logger.info("GitLab URL            : %s", cfg["gitlab_url"])
    logger.info("Size Threshold        : %s MB", cfg["size_threshold_mb"])
    logger.info("Concurrent repos      : %s", cfg["concurrent_workers"])
    logger.info("Branch workers        : %s", cfg["branch_workers"])
    logger.info("File workers          : %s", cfg["file_workers"])
    logger.info("Max concurrent reqs   : %s", cfg["max_concurrent_requests"])
    logger.info("Output format         : %s", cfg["output_format"])
    logger.info("Config file           : %s", CONFIG_FILE)
    logger.info("Repos file            : %s", REPOS_FILE)
    logger.info("=" * 70)

    # -- Load repos
    try:
        repos = load_repos(REPOS_FILE)
    except (FileNotFoundError, ValueError) as exc:
        logger.error(exc)
        sys.exit(1)

    if not repos:
        logger.error("No valid repos found in %s. Exiting.", REPOS_FILE)
        sys.exit(1)

    logger.info("Loaded %d repo(s) from repos.csv.", len(repos))

    # -- Initialize GitLab client
    client = GitLabClient(cfg, logger)

    # -- Verify connectivity
    try:
        user_info = client._get("/user")
        logger.info("Authenticated as: %s (%s)", user_info.get("name"), user_info.get("username"))
    except Exception as exc:  # pylint: disable=broad-except
        logger.error("Failed to authenticate with GitLab: %s", exc)
        sys.exit(1)

    scanned_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    all_large_files: List[Dict[str, Any]] = []
    workers = min(cfg["concurrent_workers"], len(repos))

    # -- Parallel repo scanning
    with ThreadPoolExecutor(max_workers=workers) as executor:
        futures = {
            executor.submit(
                scan_repo, client, r["org_name"], r["repo_path"], cfg, scanned_at, logger
            ): r
            for r in repos
        }

        for future in as_completed(futures):
            repo = futures[future]
            try:
                _, rows = future.result()
                all_large_files.extend(rows)
            except Exception as exc:  # pylint: disable=broad-except
                logger.error(
                    "Unexpected error scanning '%s': %s", repo["repo_path"], exc, exc_info=True
                )

    # -- Write output
    logger.info("=" * 70)
    logger.info("Scan complete. Total large files found: %d", len(all_large_files))

    if all_large_files:
        # Sort: org → repo → branch → file path
        all_large_files.sort(
            key=lambda r: (
                r["Org / Group Name"],
                r["Repository Path"],
                r["Branch Name"],
                r["File Path"],
            )
        )
        write_report(all_large_files, cfg, logger)
    else:
        logger.info("No files exceeded the %s MB threshold. No report generated.", cfg["size_threshold_mb"])

    logger.info("Done.")


if __name__ == "__main__":
    main()
